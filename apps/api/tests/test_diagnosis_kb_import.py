"""Knowledge-base import: what gets in, what gets refused, and what never happens.

The interesting assertions are the negative ones. An importer that accepts a
good file is easy; the value is in refusing a row that would put wrong advice
about a car in front of a driver, without also refusing the 399 good rows
beside it.
"""

import io

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from models.diagnosis_kb import (
    DiagnosisImportRun,
    DiagnosisMaster,
    DiagnosisSolution,
    DiagnosisSymptomAlias,
    RecordStatus,
    SourceType,
    VerificationStatus,
)
from services.diagnosis_kb_import import (
    ImportError_,
    normalise_phrase,
    parse_and_import,
)

MASTER_COLS = [
    "diagnosis_code", "manufacturer", "model", "variant", "engine_code", "transmission",
    "model_year_from", "model_year_to", "fuel_type", "odometer_from_km", "odometer_to_km",
    "system", "subsystem", "error_code", "related_error_codes", "canonical_symptom",
    "symptom", "user_keywords", "possible_cause", "diagnostic_steps", "confirms_when",
    "rule_out", "severity", "safety_critical", "can_drive", "recommended_action",
    "requires_professional", "estimated_cost_min", "estimated_cost_max",
    "source_type", "source_name", "source_url", "confidence_score",
    "verification_status", "last_verified", "status", "notes",
]
SOLUTION_COLS = [
    "solution_code", "diagnosis_code", "sequence", "solution_title", "solution_type",
    "difficulty", "is_temporary_fix", "resolves_root_cause", "steps", "expected_outcome",
    "verification_check", "tools_required", "parts_required", "oem_part_numbers",
    "consumables", "labour_hours_est", "cost_parts_min", "cost_parts_max",
    "cost_labour_min", "cost_labour_max", "success_rate_pct", "safety_warning",
    "prerequisites", "do_not_attempt_if", "warranty_impact", "environmental_note",
    "source_type", "source_name", "source_url", "confidence_score",
    "verification_status", "status", "notes",
]
ALIAS_COLS = ["canonical_symptom", "user_phrase", "language", "status", "notes"]


def master_row(**over):
    row = {
        "diagnosis_code": "TEST-001", "manufacturer": "Tata", "model": "Nexon",
        "variant": None, "engine_code": "1.2 Revotron", "transmission": "Manual",
        "model_year_from": 2017, "model_year_to": 2023, "fuel_type": "Petrol",
        "odometer_from_km": None, "odometer_to_km": None,
        "system": "Engine", "subsystem": "Ignition", "error_code": "P0301",
        "related_error_codes": "P0300|P0302", "canonical_symptom": "ENGINE_MISFIRE",
        "symptom": "Misfire under load on cylinder 1",
        "user_keywords": "shaking|jerking|vibrating",
        "possible_cause": "Faulty ignition coil|Worn spark plug",
        "diagnostic_steps": "Read live data|Swap coil to another cylinder",
        "confirms_when": "Misfire follows the swapped coil", "rule_out": "No misfire at idle",
        "severity": "MEDIUM", "safety_critical": "FALSE", "can_drive": "LIMITED",
        "recommended_action": "Have the ignition system checked.",
        "requires_professional": "TRUE",
        "estimated_cost_min": 1500, "estimated_cost_max": 6000,
        "source_type": "TECHNICAL", "source_name": "Test manual",
        "source_url": "https://example.com/manual", "confidence_score": 0.8,
        "verification_status": "VERIFIED", "last_verified": "2026-01-15",
        "status": "ACTIVE", "notes": None,
    }
    row.update(over)
    return row


def solution_row(**over):
    row = {
        "solution_code": "TEST-001-S1", "diagnosis_code": "TEST-001", "sequence": 1,
        "solution_title": "Replace ignition coil", "solution_type": "PART_REPLACEMENT",
        "difficulty": "MECHANIC", "is_temporary_fix": "FALSE", "resolves_root_cause": "TRUE",
        "steps": "Disconnect battery|Remove coil|Fit new coil|Clear codes",
        "expected_outcome": "Misfire clears", "verification_check": "No P0301 after 20 km",
        "tools_required": "10mm socket|Torque wrench", "parts_required": "Ignition coil",
        "oem_part_numbers": None, "consumables": None, "labour_hours_est": 0.5,
        "cost_parts_min": 1200, "cost_parts_max": 3000,
        "cost_labour_min": 300, "cost_labour_max": 800, "success_rate_pct": 85,
        "safety_warning": "Disconnect the battery before working on ignition.",
        "prerequisites": "Engine cold", "do_not_attempt_if": "You cannot isolate the battery",
        "warranty_impact": "NONE", "environmental_note": None,
        "source_type": "TECHNICAL", "source_name": "Test manual", "source_url": None,
        "confidence_score": 0.8, "verification_status": "VERIFIED", "status": "ACTIVE",
        "notes": None,
    }
    row.update(over)
    return row


def build_workbook(masters=None, solutions=None, aliases=None, *, drop_master_col=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "diagnosis_master"
    cols = [c for c in MASTER_COLS if c != drop_master_col]
    ws.append(cols)
    for m in masters or []:
        ws.append([m.get(c) for c in cols])

    ws2 = wb.create_sheet("diagnosis_solutions")
    ws2.append(SOLUTION_COLS)
    for s in solutions or []:
        ws2.append([s.get(c) for c in SOLUTION_COLS])

    ws3 = wb.create_sheet("symptom_aliases")
    ws3.append(ALIAS_COLS)
    for a in aliases or []:
        ws3.append([a.get(c) for c in ALIAS_COLS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def run(db, content, *, dry_run=False, filename="kb.xlsx"):
    return await parse_and_import(
        db, content=content, filename=filename, imported_by="admin@test", dry_run=dry_run
    )


# ── the happy path ──────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_valid_workbook_imports_master_solutions_and_aliases(db_session):
    content = build_workbook(
        [master_row()],
        [solution_row(), solution_row(solution_code="TEST-001-S2", sequence=2,
                                      solution_title="Replace spark plug")],
        [{"canonical_symptom": "ENGINE_MISFIRE", "user_phrase": "Engine shaking!",
          "language": "en", "status": "ACTIVE"}],
    )
    result = await run(db_session, content)

    assert result.master_created == 1
    assert result.solution_created == 2
    assert result.alias_created == 1
    assert result.errors == []

    row = (await db_session.execute(
        select(DiagnosisMaster).where(DiagnosisMaster.diagnosis_code == "TEST-001")
    )).scalar_one()
    assert row.is_servable
    assert row.canonical_symptom == "ENGINE_MISFIRE"

    # Queried rather than reached through row.solutions: a lazy relationship
    # load inside an async session raises MissingGreenlet.
    sols = (await db_session.execute(
        select(DiagnosisSolution).where(DiagnosisSolution.diagnosis_id == row.id)
        .order_by(DiagnosisSolution.sequence)
    )).scalars().all()
    assert [s.sequence for s in sols] == [1, 2]

    alias = (await db_session.execute(select(DiagnosisSymptomAlias))).scalar_one()
    # Stored normalised so lookup is an index hit and punctuation cannot fork a row.
    assert alias.normalised_phrase == "engine shaking"


@pytest.mark.asyncio
async def test_reimport_updates_rather_than_duplicating(db_session):
    await run(db_session, build_workbook([master_row()], [solution_row()]))
    result = await run(
        db_session,
        build_workbook([master_row(symptom="Revised symptom text")], [solution_row()]),
    )

    assert result.master_created == 0
    assert result.master_updated == 1
    total = (await db_session.execute(
        select(func.count()).select_from(DiagnosisMaster)
    )).scalar_one()
    assert total == 1
    row = (await db_session.execute(select(DiagnosisMaster))).scalar_one()
    assert row.symptom == "Revised symptom text"


# ── refusals ────────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_one_bad_row_does_not_block_the_good_ones(db_session):
    """The property that decides whether anyone ever uses the importer."""
    content = build_workbook([
        master_row(),
        master_row(diagnosis_code="TEST-002", severity="CATASTROPHIC"),   # invalid enum
        master_row(diagnosis_code="TEST-003"),
    ])
    result = await run(db_session, content)

    assert result.master_created == 2
    assert result.master_rejected == 1
    assert any(e.column == "severity" and e.row == 3 for e in result.errors)

    codes = set((await db_session.execute(select(DiagnosisMaster.diagnosis_code))).scalars().all())
    assert codes == {"TEST-001", "TEST-003"}


@pytest.mark.asyncio
async def test_orphan_solution_is_rejected_not_silently_dropped(db_session):
    content = build_workbook(
        [master_row()],
        [solution_row(solution_code="ORPHAN-S1", diagnosis_code="DOES-NOT-EXIST")],
    )
    result = await run(db_session, content)

    assert result.solution_created == 0
    assert result.solution_rejected == 1
    assert any("no diagnosis" in e.message for e in result.errors)


@pytest.mark.asyncio
async def test_temporary_fix_cannot_claim_to_resolve_root_cause(db_session):
    """A bypass sold as a repair is how somebody breaks down twice."""
    content = build_workbook(
        [master_row()],
        [solution_row(is_temporary_fix="TRUE", resolves_root_cause="TRUE")],
    )
    result = await run(db_session, content)

    assert result.solution_rejected == 1
    assert any("bypass is not a repair" in e.message for e in result.errors)


@pytest.mark.asyncio
async def test_ai_generated_rows_cannot_promote_themselves(db_session):
    """The review queue must not be bypassable by editing a spreadsheet cell."""
    content = build_workbook(
        [master_row(source_type="AI_GENERATED", verification_status="VERIFIED")]
    )
    result = await run(db_session, content)

    assert result.master_created == 1
    row = (await db_session.execute(select(DiagnosisMaster))).scalar_one()
    assert row.source_type == SourceType.ai_generated
    assert row.verification_status == VerificationStatus.pending_review
    assert not row.is_servable


@pytest.mark.asyncio
async def test_malformed_url_rejects_the_row(db_session):
    content = build_workbook([master_row(source_url="not-a-url")])
    result = await run(db_session, content)
    assert result.master_rejected == 1
    assert any(e.column == "source_url" for e in result.errors)


@pytest.mark.asyncio
async def test_duplicate_code_within_one_file_is_caught(db_session):
    content = build_workbook([master_row(), master_row(symptom="Second copy")])
    result = await run(db_session, content)
    assert result.master_rejected == 1
    assert any("duplicate" in e.message for e in result.errors)


@pytest.mark.asyncio
async def test_inverted_year_range_is_rejected(db_session):
    content = build_workbook([master_row(model_year_from=2023, model_year_to=2017)])
    result = await run(db_session, content)
    assert result.master_rejected == 1
    assert any(e.column == "model_year_to" for e in result.errors)


@pytest.mark.asyncio
async def test_confidence_outside_zero_to_one_is_rejected(db_session):
    content = build_workbook([master_row(confidence_score=85)])   # 85 not 0.85
    result = await run(db_session, content)
    assert result.master_rejected == 1
    assert any(e.column == "confidence_score" for e in result.errors)


@pytest.mark.asyncio
async def test_lowercase_canonical_symptom_is_rejected(db_session):
    content = build_workbook([master_row(canonical_symptom="engine misfire")])
    result = await run(db_session, content)
    assert result.master_rejected == 1
    assert any(e.column == "canonical_symptom" for e in result.errors)


@pytest.mark.asyncio
async def test_missing_required_column_fails_the_whole_file(db_session):
    """A structural fault is different from a row fault: nothing can be trusted."""
    content = build_workbook([master_row()], drop_master_col="can_drive")
    with pytest.raises(ImportError_, match="missing required columns"):
        await run(db_session, content)


@pytest.mark.asyncio
async def test_non_excel_file_is_refused(db_session):
    with pytest.raises(ImportError_, match="Only .xlsx"):
        await run(db_session, b"not a workbook", filename="kb.csv")


@pytest.mark.asyncio
async def test_corrupt_workbook_is_refused_cleanly(db_session):
    with pytest.raises(ImportError_, match="Could not read"):
        await run(db_session, b"PK\x03\x04 garbage", filename="kb.xlsx")


# ── dry run ─────────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_dry_run_reports_without_writing(db_session):
    content = build_workbook([master_row()], [solution_row()])
    result = await run(db_session, content, dry_run=True)

    assert result.master_created == 1
    assert result.solution_created == 1

    assert (await db_session.execute(
        select(func.count()).select_from(DiagnosisMaster)
    )).scalar_one() == 0
    assert (await db_session.execute(
        select(func.count()).select_from(DiagnosisSolution)
    )).scalar_one() == 0


@pytest.mark.asyncio
async def test_the_preview_itself_is_recorded(db_session):
    """An admin previewing a file is a fact worth keeping — an earlier version
    rolled the audit row back along with the staged writes."""
    await run(db_session, build_workbook([master_row()]), dry_run=True)

    run_row = (await db_session.execute(select(DiagnosisImportRun))).scalar_one()
    assert run_row.dry_run is True
    assert run_row.master_created == 1
    assert run_row.imported_by == "admin@test"


@pytest.mark.asyncio
async def test_import_never_deletes_existing_records(db_session):
    """A wrong file must not be able to empty the corpus."""
    await run(db_session, build_workbook([master_row(), master_row(diagnosis_code="KEEP-ME")]))
    await run(db_session, build_workbook([master_row(symptom="only this one now")]))

    codes = set((await db_session.execute(select(DiagnosisMaster.diagnosis_code))).scalars().all())
    assert "KEEP-ME" in codes


# ── normalisation ───────────────────────────────────────────────────────────

def test_normalise_phrase_collapses_case_punctuation_and_space():
    assert normalise_phrase("Engine  SHAKING!!") == "engine shaking"
    assert normalise_phrase("  car is jerking, badly ") == "car is jerking badly"
    assert normalise_phrase("") == ""


@pytest.mark.asyncio
async def test_alias_variants_collapse_to_one_row(db_session):
    content = build_workbook(
        [master_row()],
        aliases=[
            {"canonical_symptom": "ENGINE_MISFIRE", "user_phrase": "engine shaking"},
            {"canonical_symptom": "ENGINE_MISFIRE", "user_phrase": "Engine  Shaking!"},
        ],
    )
    result = await run(db_session, content)
    assert result.alias_created == 1


@pytest.mark.asyncio
async def test_any_scope_is_stored_canonically(db_session):
    content = build_workbook([master_row(manufacturer="any", model="Any")])
    await run(db_session, content)
    row = (await db_session.execute(select(DiagnosisMaster))).scalar_one()
    assert row.manufacturer == "ANY"
    assert row.model == "ANY"


@pytest.mark.asyncio
async def test_draft_rows_are_not_servable(db_session):
    content = build_workbook([
        master_row(diagnosis_code="D-1", status="DRAFT", verification_status="VERIFIED"),
        master_row(diagnosis_code="D-2", status="ACTIVE", verification_status="PENDING_REVIEW"),
        master_row(diagnosis_code="D-3", status="ACTIVE", verification_status="VERIFIED"),
    ])
    await run(db_session, content)

    rows = {
        r.diagnosis_code: r
        for r in (await db_session.execute(select(DiagnosisMaster))).scalars().all()
    }
    assert rows["D-1"].is_servable is False   # published but unreviewed
    assert rows["D-2"].is_servable is False   # reviewed but unpublished
    assert rows["D-3"].is_servable is True    # both gates passed
    assert rows["D-1"].status == RecordStatus.draft
