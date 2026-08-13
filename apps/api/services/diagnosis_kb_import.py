"""Excel → knowledge base import: validate every row, reject the bad ones, keep the rest.

Excel is an authoring format, never the runtime store. This reads a workbook,
checks it row by row, and writes what passes into Postgres.

THE RULES THAT MATTER

  * A bad row never stops a good one. Editors work in large batches; failing
    the whole file on row 340 means the other 339 have to be re-imported, and
    in practice means nobody imports anything.
  * Nothing is ever deleted. Retirement is `status = INACTIVE`, which is
    reversible; a DELETE is not, and an import is exactly the kind of bulk
    operation where a wrong file would otherwise be catastrophic.
  * Upsert on the natural key — diagnosis_code, solution_code — so re-importing
    a corrected file updates rather than duplicates.
  * AI_GENERATED is forced to PENDING_REVIEW no matter what the file says.
    Otherwise the review queue can be bypassed by editing a spreadsheet cell.
  * A solution whose diagnosis_code does not exist is rejected rather than
    orphaned. An orphan solution is unreachable but still counted, which makes
    the KB look bigger than it is.

Dry-run first. `preview()` does every check and writes nothing, so an admin can
see the damage before committing.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.diagnosis_kb import (
    ANY_VEHICLE,
    CanDrive,
    DiagnosisImportRun,
    DiagnosisMaster,
    DiagnosisSolution,
    DiagnosisSymptomAlias,
    Difficulty,
    RecordStatus,
    Severity,
    SolutionType,
    SourceType,
    VerificationStatus,
    WarrantyImpact,
)

logger = logging.getLogger("gaadiiq.kb_import")

MASTER_SHEET = "diagnosis_master"
SOLUTION_SHEET = "diagnosis_solutions"
ALIAS_SHEET = "symptom_aliases"

MASTER_REQUIRED = [
    "diagnosis_code", "manufacturer", "model", "model_year_from", "model_year_to",
    "fuel_type", "system", "canonical_symptom", "symptom", "user_keywords",
    "possible_cause", "diagnostic_steps", "severity", "safety_critical",
    "can_drive", "recommended_action", "requires_professional",
    "source_type", "source_name", "confidence_score", "verification_status", "status",
]
SOLUTION_REQUIRED = [
    "solution_code", "diagnosis_code", "sequence", "solution_title", "solution_type",
    "difficulty", "is_temporary_fix", "resolves_root_cause", "steps",
    "source_type", "source_name", "confidence_score", "verification_status", "status",
]
ALIAS_REQUIRED = ["canonical_symptom", "user_phrase"]

_URL = re.compile(r"^https?://[^\s]+$", re.IGNORECASE)
_CANON = re.compile(r"^[A-Z][A-Z0-9_]{2,79}$")


@dataclass
class RowError:
    sheet: str
    row: int
    column: str
    message: str

    def as_dict(self) -> dict[str, Any]:
        return {"sheet": self.sheet, "row": self.row, "column": self.column, "message": self.message}


@dataclass
class ImportResult:
    dry_run: bool = True
    master_read: int = 0
    master_created: int = 0
    master_updated: int = 0
    master_rejected: int = 0
    solution_read: int = 0
    solution_created: int = 0
    solution_updated: int = 0
    solution_rejected: int = 0
    alias_read: int = 0
    alias_created: int = 0
    errors: list[RowError] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors

    def as_dict(self) -> dict[str, Any]:
        return {
            "dry_run": self.dry_run,
            "master": {
                "read": self.master_read, "created": self.master_created,
                "updated": self.master_updated, "rejected": self.master_rejected,
            },
            "solutions": {
                "read": self.solution_read, "created": self.solution_created,
                "updated": self.solution_updated, "rejected": self.solution_rejected,
            },
            "aliases": {"read": self.alias_read, "created": self.alias_created},
            "error_count": len(self.errors),
            # Capped: a wholly wrong file produces thousands, and an admin needs
            # the first page, not a payload that stalls the browser.
            "errors": [e.as_dict() for e in self.errors[:200]],
            "errors_truncated": max(0, len(self.errors) - 200),
        }


class ImportError_(Exception):
    """The file could not be read at all — wrong type, corrupt, missing sheets."""


def normalise_phrase(text: str) -> str:
    """Lower-case, strip punctuation, collapse whitespace.

    The alias table stores this so a lookup is an index hit rather than a scan,
    and so 'Engine  shaking!' and 'engine shaking' cannot become two rows.
    """
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", (text or "").lower())).strip()


# ── cell coercion ───────────────────────────────────────────────────────────
def _s(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _scope(value: Any) -> str | None:
    """A vehicle-scope field. 'any', 'Any' and 'ANY' all mean the sentinel."""
    text = _s(value)
    if text is None:
        return None
    return ANY_VEHICLE if text.upper() == ANY_VEHICLE else text


def _int(value: Any, *, field_name: str, row: int, sheet: str, errors: list[RowError]) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        errors.append(RowError(sheet, row, field_name, f"expected a whole number, got {value!r}"))
        return None


def _float(value: Any, *, field_name: str, row: int, sheet: str, errors: list[RowError]) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        errors.append(RowError(sheet, row, field_name, f"expected a number, got {value!r}"))
        return None


def _bool(value: Any, *, field_name: str, row: int, sheet: str, errors: list[RowError]) -> bool | None:
    text = str(value).strip().upper() if value is not None else ""
    if text in ("TRUE", "YES", "Y", "1"):
        return True
    if text in ("FALSE", "NO", "N", "0"):
        return False
    errors.append(RowError(sheet, row, field_name, f"expected TRUE or FALSE, got {value!r}"))
    return None


def _enum(enum_cls, value: Any, *, field_name: str, row: int, sheet: str, errors: list[RowError]):
    text = str(value).strip().upper() if value is not None else ""
    for member in enum_cls:
        if member.value == text:
            return member
    allowed = ", ".join(m.value for m in enum_cls)
    errors.append(RowError(sheet, row, field_name, f"got {value!r}; allowed: {allowed}"))
    return None


def _date(value: Any, *, field_name: str, row: int, sheet: str, errors: list[RowError]) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        errors.append(RowError(sheet, row, field_name, f"expected YYYY-MM-DD, got {value!r}"))
        return None


def _url(value: Any, *, field_name: str, row: int, sheet: str, errors: list[RowError]) -> str | None:
    text = _s(value)
    if text is None:
        return None
    if not _URL.match(text):
        errors.append(RowError(sheet, row, field_name, "must start with http:// or https://"))
        return None
    return text


def _read_sheet(wb, name: str) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    """Return (headers, [(excel_row_number, {header: value})]). Blank rows skipped."""
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    out: list[tuple[int, dict[str, Any]]] = []
    for offset, values in enumerate(rows, start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        out.append((offset, {headers[i]: v for i, v in enumerate(values) if i < len(headers)}))
    return headers, out


def _missing_columns(headers: list[str], required: list[str]) -> list[str]:
    present = {h for h in headers if h}
    return [c for c in required if c not in present]


async def parse_and_import(
    db: AsyncSession,
    *,
    content: bytes,
    filename: str,
    imported_by: str,
    dry_run: bool = True,
) -> ImportResult:
    """Validate the workbook and, unless dry_run, write what passes."""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ImportError_("Only .xlsx or .xlsm files can be imported.")

    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ImportError_(f"Could not read the workbook: {exc}") from exc

    result = ImportResult(dry_run=dry_run)
    errors = result.errors

    m_headers, m_rows = _read_sheet(wb, MASTER_SHEET)
    if not m_headers:
        raise ImportError_(f"Sheet '{MASTER_SHEET}' is missing or empty.")
    missing = _missing_columns(m_headers, MASTER_REQUIRED)
    if missing:
        raise ImportError_(f"'{MASTER_SHEET}' is missing required columns: {', '.join(missing)}")

    s_headers, s_rows = _read_sheet(wb, SOLUTION_SHEET)
    if s_headers:
        missing = _missing_columns(s_headers, SOLUTION_REQUIRED)
        if missing:
            raise ImportError_(f"'{SOLUTION_SHEET}' is missing required columns: {', '.join(missing)}")

    a_headers, a_rows = _read_sheet(wb, ALIAS_SHEET)
    if a_headers:
        missing = _missing_columns(a_headers, ALIAS_REQUIRED)
        if missing:
            raise ImportError_(f"'{ALIAS_SHEET}' is missing required columns: {', '.join(missing)}")

    # ── master ──────────────────────────────────────────────────────────────
    result.master_read = len(m_rows)
    seen_codes: dict[str, int] = {}
    staged_master: list[tuple[int, dict[str, Any]]] = []

    for row_no, row in m_rows:
        before = len(errors)
        code = _s(row.get("diagnosis_code"))
        if not code:
            errors.append(RowError(MASTER_SHEET, row_no, "diagnosis_code", "required"))
        elif code in seen_codes:
            errors.append(RowError(
                MASTER_SHEET, row_no, "diagnosis_code",
                f"duplicate of row {seen_codes[code]} in this file",
            ))
        else:
            seen_codes[code] = row_no

        for col in MASTER_REQUIRED:
            if _s(row.get(col)) is None and col != "diagnosis_code":
                errors.append(RowError(MASTER_SHEET, row_no, col, "required"))

        canon = _s(row.get("canonical_symptom"))
        if canon and not _CANON.match(canon):
            errors.append(RowError(
                MASTER_SHEET, row_no, "canonical_symptom",
                "must be UPPER_SNAKE_CASE, e.g. ENGINE_MISFIRE",
            ))

        y_from = _int(row.get("model_year_from"), field_name="model_year_from", row=row_no, sheet=MASTER_SHEET, errors=errors)
        y_to = _int(row.get("model_year_to"), field_name="model_year_to", row=row_no, sheet=MASTER_SHEET, errors=errors)
        if y_from is not None and (y_from < 1950 or y_from > 2100):
            errors.append(RowError(MASTER_SHEET, row_no, "model_year_from", "outside 1950-2100"))
        if y_from is not None and y_to is not None and y_to < y_from:
            errors.append(RowError(MASTER_SHEET, row_no, "model_year_to", "must be >= model_year_from"))

        c_min = _int(row.get("estimated_cost_min"), field_name="estimated_cost_min", row=row_no, sheet=MASTER_SHEET, errors=errors)
        c_max = _int(row.get("estimated_cost_max"), field_name="estimated_cost_max", row=row_no, sheet=MASTER_SHEET, errors=errors)
        if c_min is not None and c_max is not None and c_max < c_min:
            errors.append(RowError(MASTER_SHEET, row_no, "estimated_cost_max", "must be >= estimated_cost_min"))

        conf = _float(row.get("confidence_score"), field_name="confidence_score", row=row_no, sheet=MASTER_SHEET, errors=errors)
        if conf is not None and not (0.0 <= conf <= 1.0):
            errors.append(RowError(MASTER_SHEET, row_no, "confidence_score", "must be between 0.00 and 1.00"))

        severity = _enum(Severity, row.get("severity"), field_name="severity", row=row_no, sheet=MASTER_SHEET, errors=errors)
        can_drive = _enum(CanDrive, row.get("can_drive"), field_name="can_drive", row=row_no, sheet=MASTER_SHEET, errors=errors)
        src = _enum(SourceType, row.get("source_type"), field_name="source_type", row=row_no, sheet=MASTER_SHEET, errors=errors)
        vstat = _enum(
            VerificationStatus, row.get("verification_status"),
            field_name="verification_status", row=row_no, sheet=MASTER_SHEET, errors=errors,
        )
        rstat = _enum(RecordStatus, row.get("status"), field_name="status", row=row_no, sheet=MASTER_SHEET, errors=errors)
        safety = _bool(row.get("safety_critical"), field_name="safety_critical", row=row_no, sheet=MASTER_SHEET, errors=errors)
        prof = _bool(row.get("requires_professional"), field_name="requires_professional", row=row_no, sheet=MASTER_SHEET, errors=errors)
        url = _url(row.get("source_url"), field_name="source_url", row=row_no, sheet=MASTER_SHEET, errors=errors)
        verified_on = _date(row.get("last_verified"), field_name="last_verified", row=row_no, sheet=MASTER_SHEET, errors=errors)

        # An AI-written row may not promote itself out of the queue.
        if src == SourceType.ai_generated and vstat != VerificationStatus.pending_review:
            vstat = VerificationStatus.pending_review

        if len(errors) > before:
            result.master_rejected += 1
            continue

        staged_master.append((row_no, {
            "diagnosis_code": code,
            "manufacturer": _scope(row.get("manufacturer")),
            "model": _scope(row.get("model")),
            "variant": _s(row.get("variant")),
            "engine_code": _s(row.get("engine_code")),
            "transmission": _s(row.get("transmission")),
            "model_year_from": y_from,
            "model_year_to": y_to,
            "fuel_type": _s(row.get("fuel_type")),
            "odometer_from_km": _int(row.get("odometer_from_km"), field_name="odometer_from_km", row=row_no, sheet=MASTER_SHEET, errors=errors),
            "odometer_to_km": _int(row.get("odometer_to_km"), field_name="odometer_to_km", row=row_no, sheet=MASTER_SHEET, errors=errors),
            "system": _s(row.get("system")),
            "subsystem": _s(row.get("subsystem")),
            "error_code": _s(row.get("error_code")),
            "related_error_codes": _s(row.get("related_error_codes")),
            "canonical_symptom": canon,
            "symptom": _s(row.get("symptom")),
            "user_keywords": _s(row.get("user_keywords")),
            "possible_cause": _s(row.get("possible_cause")),
            "diagnostic_steps": _s(row.get("diagnostic_steps")),
            "confirms_when": _s(row.get("confirms_when")),
            "rule_out": _s(row.get("rule_out")),
            "severity": severity,
            "safety_critical": bool(safety),
            "can_drive": can_drive,
            "recommended_action": _s(row.get("recommended_action")),
            "requires_professional": bool(prof),
            "estimated_cost_min": c_min,
            "estimated_cost_max": c_max,
            "source_type": src,
            "source_name": _s(row.get("source_name")),
            "source_url": url,
            "confidence_score": conf or 0.0,
            "verification_status": vstat,
            "last_verified": verified_on,
            "status": rstat,
            "notes": _s(row.get("notes")),
        }))

    # ── write master ────────────────────────────────────────────────────────
    code_to_id: dict[str, Any] = {}
    if staged_master:
        codes = [d["diagnosis_code"] for _r, d in staged_master]
        existing = {
            m.diagnosis_code: m
            for m in (await db.execute(
                select(DiagnosisMaster).where(DiagnosisMaster.diagnosis_code.in_(codes))
            )).scalars().all()
        }
        for _row_no, data in staged_master:
            found = existing.get(data["diagnosis_code"])
            if found is not None:
                if not dry_run:
                    for k, v in data.items():
                        setattr(found, k, v)
                result.master_updated += 1
                code_to_id[data["diagnosis_code"]] = found.id
            else:
                result.master_created += 1
                if not dry_run:
                    record = DiagnosisMaster(**data)
                    db.add(record)
                    await db.flush()
                    code_to_id[data["diagnosis_code"]] = record.id
                else:
                    code_to_id[data["diagnosis_code"]] = None

    # ── solutions ───────────────────────────────────────────────────────────
    result.solution_read = len(s_rows)
    seen_solutions: dict[str, int] = {}
    staged_solutions: list[tuple[int, str, dict[str, Any]]] = []

    for row_no, row in s_rows:
        before = len(errors)
        s_code = _s(row.get("solution_code"))
        d_code = _s(row.get("diagnosis_code"))

        if not s_code:
            errors.append(RowError(SOLUTION_SHEET, row_no, "solution_code", "required"))
        elif s_code in seen_solutions:
            errors.append(RowError(
                SOLUTION_SHEET, row_no, "solution_code",
                f"duplicate of row {seen_solutions[s_code]} in this file",
            ))
        else:
            seen_solutions[s_code] = row_no

        for col in SOLUTION_REQUIRED:
            if _s(row.get(col)) is None and col not in ("solution_code", "diagnosis_code"):
                errors.append(RowError(SOLUTION_SHEET, row_no, col, "required"))

        # Orphan check: against this file first, then the database.
        known_here = d_code in code_to_id
        if not d_code:
            errors.append(RowError(SOLUTION_SHEET, row_no, "diagnosis_code", "required"))
        elif not known_here:
            found = (await db.execute(
                select(DiagnosisMaster).where(DiagnosisMaster.diagnosis_code == d_code)
            )).scalar_one_or_none()
            if found is None:
                errors.append(RowError(
                    SOLUTION_SHEET, row_no, "diagnosis_code",
                    f"no diagnosis '{d_code}' on the master sheet or in the database",
                ))
            else:
                code_to_id[d_code] = found.id

        seq = _int(row.get("sequence"), field_name="sequence", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        stype = _enum(SolutionType, row.get("solution_type"), field_name="solution_type", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        diff = _enum(Difficulty, row.get("difficulty"), field_name="difficulty", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        temp = _bool(row.get("is_temporary_fix"), field_name="is_temporary_fix", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        root = _bool(row.get("resolves_root_cause"), field_name="resolves_root_cause", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        s_src = _enum(SourceType, row.get("source_type"), field_name="source_type", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        s_vstat = _enum(
            VerificationStatus,
            row.get("verification_status"),
            field_name="verification_status",
            row=row_no,
            sheet=SOLUTION_SHEET,
            errors=errors,
        )
        s_rstat = _enum(RecordStatus, row.get("status"), field_name="status", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        s_conf = _float(row.get("confidence_score"), field_name="confidence_score", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        if s_conf is not None and not (0.0 <= s_conf <= 1.0):
            errors.append(RowError(SOLUTION_SHEET, row_no, "confidence_score", "must be between 0.00 and 1.00"))

        warranty = None
        if _s(row.get("warranty_impact")):
            warranty = _enum(
                WarrantyImpact,
                row.get("warranty_impact"),
                field_name="warranty_impact",
                row=row_no,
                sheet=SOLUTION_SHEET,
                errors=errors,
            )

        success = _int(row.get("success_rate_pct"), field_name="success_rate_pct", row=row_no, sheet=SOLUTION_SHEET, errors=errors)
        if success is not None and not (0 <= success <= 100):
            errors.append(RowError(SOLUTION_SHEET, row_no, "success_rate_pct", "must be 0-100"))

        # A temporary fix that also claims to resolve the root cause is a
        # contradiction, and the direction it fails in is the dangerous one:
        # the user is told a bypass is a repair.
        if temp is True and root is True:
            errors.append(RowError(
                SOLUTION_SHEET, row_no, "resolves_root_cause",
                "cannot be TRUE when is_temporary_fix is TRUE — a bypass is not a repair",
            ))

        if s_src == SourceType.ai_generated and s_vstat != VerificationStatus.pending_review:
            s_vstat = VerificationStatus.pending_review

        if len(errors) > before:
            result.solution_rejected += 1
            continue

        staged_solutions.append((row_no, d_code, {
            "solution_code": s_code,
            "sequence": seq or 1,
            "solution_title": _s(row.get("solution_title")),
            "solution_type": stype,
            "difficulty": diff,
            "is_temporary_fix": bool(temp),
            "resolves_root_cause": bool(root),
            "steps": _s(row.get("steps")),
            "expected_outcome": _s(row.get("expected_outcome")),
            "verification_check": _s(row.get("verification_check")),
            "tools_required": _s(row.get("tools_required")),
            "parts_required": _s(row.get("parts_required")),
            "oem_part_numbers": _s(row.get("oem_part_numbers")),
            "consumables": _s(row.get("consumables")),
            "labour_hours_est": _float(row.get("labour_hours_est"), field_name="labour_hours_est", row=row_no, sheet=SOLUTION_SHEET, errors=errors),
            "cost_parts_min": _int(row.get("cost_parts_min"), field_name="cost_parts_min", row=row_no, sheet=SOLUTION_SHEET, errors=errors),
            "cost_parts_max": _int(row.get("cost_parts_max"), field_name="cost_parts_max", row=row_no, sheet=SOLUTION_SHEET, errors=errors),
            "cost_labour_min": _int(row.get("cost_labour_min"), field_name="cost_labour_min", row=row_no, sheet=SOLUTION_SHEET, errors=errors),
            "cost_labour_max": _int(row.get("cost_labour_max"), field_name="cost_labour_max", row=row_no, sheet=SOLUTION_SHEET, errors=errors),
            "success_rate_pct": success,
            "safety_warning": _s(row.get("safety_warning")),
            "prerequisites": _s(row.get("prerequisites")),
            "do_not_attempt_if": _s(row.get("do_not_attempt_if")),
            "warranty_impact": warranty,
            "environmental_note": _s(row.get("environmental_note")),
            "source_type": s_src,
            "source_name": _s(row.get("source_name")),
            "source_url": _url(row.get("source_url"), field_name="source_url", row=row_no, sheet=SOLUTION_SHEET, errors=errors),
            "confidence_score": s_conf or 0.0,
            "verification_status": s_vstat,
            "status": s_rstat,
            "notes": _s(row.get("notes")),
        }))

    if staged_solutions and not dry_run:
        codes = [d["solution_code"] for _r, _d, d in staged_solutions]
        existing_sol = {
            s.solution_code: s
            for s in (await db.execute(
                select(DiagnosisSolution).where(DiagnosisSolution.solution_code.in_(codes))
            )).scalars().all()
        }
        for _row_no, d_code, data in staged_solutions:
            found = existing_sol.get(data["solution_code"])
            if found is not None:
                for k, v in data.items():
                    setattr(found, k, v)
                result.solution_updated += 1
            else:
                db.add(DiagnosisSolution(diagnosis_id=code_to_id[d_code], **data))
                result.solution_created += 1
    elif staged_solutions:
        codes = [d["solution_code"] for _r, _d, d in staged_solutions]
        existing_codes = set((await db.execute(
            select(DiagnosisSolution.solution_code).where(DiagnosisSolution.solution_code.in_(codes))
        )).scalars().all())
        for _row_no, _d_code, data in staged_solutions:
            if data["solution_code"] in existing_codes:
                result.solution_updated += 1
            else:
                result.solution_created += 1

    # ── aliases ─────────────────────────────────────────────────────────────
    result.alias_read = len(a_rows)
    seen_aliases: set[tuple[str, str]] = set()
    for row_no, row in a_rows:
        canon = _s(row.get("canonical_symptom"))
        phrase = _s(row.get("user_phrase"))
        lang = _s(row.get("language")) or "en"
        if not canon or not phrase:
            errors.append(RowError(ALIAS_SHEET, row_no, "user_phrase", "canonical_symptom and user_phrase are both required"))
            continue
        norm = normalise_phrase(phrase)
        if not norm:
            errors.append(RowError(ALIAS_SHEET, row_no, "user_phrase", "empty once punctuation is stripped"))
            continue
        key = (norm, lang)
        if key in seen_aliases:
            continue          # silently deduplicated: repetition here is harmless
        seen_aliases.add(key)

        if dry_run:
            result.alias_created += 1
            continue

        found = (await db.execute(
            select(DiagnosisSymptomAlias).where(
                DiagnosisSymptomAlias.normalised_phrase == norm,
                DiagnosisSymptomAlias.language == lang,
            )
        )).scalar_one_or_none()
        if found is not None:
            found.canonical_symptom = canon
            found.user_phrase = phrase
        else:
            db.add(DiagnosisSymptomAlias(
                canonical_symptom=canon, user_phrase=phrase,
                normalised_phrase=norm, language=lang,
                status=RecordStatus.active, notes=_s(row.get("notes")),
            ))
            result.alias_created += 1

    # ── audit ───────────────────────────────────────────────────────────────
    run = DiagnosisImportRun(
        filename=filename,
        imported_by=imported_by,
        dry_run=dry_run,
        master_rows_read=result.master_read,
        master_created=result.master_created,
        master_updated=result.master_updated,
        master_rejected=result.master_rejected,
        solution_rows_read=result.solution_read,
        solution_created=result.solution_created,
        solution_updated=result.solution_updated,
        solution_rejected=result.solution_rejected,
        alias_rows_read=result.alias_read,
        alias_created=result.alias_created,
        errors="\n".join(
            f"{e.sheet}!row{e.row} [{e.column}] {e.message}" for e in errors[:500]
        ) or None,
        completed_at=datetime.now(timezone.utc),
    )
    db.add(run)

    if dry_run:
        # Undo the staged writes but keep the audit row: an admin previewing a
        # file is a fact worth recording, and rollback would erase it.
        await db.flush()
        await db.rollback()
        db.add(run)

    await db.commit()

    logger.info(
        "kb import%s: %s master(+%d/~%d/-%d) solutions(+%d/~%d/-%d) aliases(+%d) errors=%d",
        " [dry run]" if dry_run else "",
        filename,
        result.master_created, result.master_updated, result.master_rejected,
        result.solution_created, result.solution_updated, result.solution_rejected,
        result.alias_created, len(errors),
    )
    return result
