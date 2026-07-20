#!/usr/bin/env python3
"""Generate GAADIIQ_AIDiagnosis_Test_Results.xlsx from ai-diagnosis-scenarios.json."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
QA = ROOT / "docs/qa"
ART = Path("/opt/cursor/artifacts/ai-diagnosis")
ART.mkdir(parents=True, exist_ok=True)

scenarios = json.loads((QA / "ai-diagnosis-scenarios.json").read_text())
pytest_note = "test_heuristic_*: 5 passed; endpoint tests may need bcrypt/passlib-compatible env"
ht = ART / "pytest-heuristic.txt"
if ht.exists():
    pytest_note = ht.read_text()[-200:].strip().split("\n")[-1] + " | endpoint auth tests blocked by bcrypt/passlib env mismatch"

wb = Workbook()
ws = wb.active
ws.title = "Summary"
ws["A1"] = "GAADIIQ AI Diagnosis / AI Valuation — Full Feature Retest"
ws["A1"].font = Font(bold=True, size=14)
ws["A3"] = "Scope"
ws["B3"] = scenarios["scope_note"]
ws["A4"] = "Generated"
ws["B4"] = scenarios["generated_at"]
ws["A5"] = "Scenario matrix"
c = scenarios["counts"]
ws["B5"] = (
    f"{c.get('PASS', 0)} PASS / {c.get('FAIL', 0)} FAIL / "
    f"{c.get('PARTIAL', 0)} PARTIAL (total {scenarios['total']})"
)
ws["A6"] = "API pytest"
ws["B6"] = pytest_note
ws["A7"] = "Verdict"
ws["B7"] = (
    "No mechanical AI Diagnosis. AI Valuation works via fallback/heuristic but over-claims AI, "
    "dual Claude/Ollama stacks, auth/copy contradiction, Home misroutes to AI Advisor."
)
ws["B7"].alignment = Alignment(wrap_text=True)
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 110

fills = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
    "PARTIAL": PatternFill("solid", fgColor="FFEB9C"),
}

ws2 = wb.create_sheet("Scenario Results")
for col, h in enumerate(["Suite", "ID", "Name", "Status", "Severity", "Detail"], 1):
    ws2.cell(1, col, h).font = Font(bold=True)
for i, r in enumerate(scenarios["results"], 2):
    ws2.cell(i, 1, r["suite"])
    ws2.cell(i, 2, r["id"])
    ws2.cell(i, 3, r["name"])
    cell = ws2.cell(i, 4, r["status"])
    cell.fill = fills.get(r["status"], PatternFill())
    ws2.cell(i, 5, r.get("severity", "—"))
    ws2.cell(i, 6, r["detail"])
for col, w in enumerate([16, 10, 55, 10, 10, 95], 1):
    ws2.column_dimensions[chr(64 + col)].width = w

ws3 = wb.create_sheet("Gaps for Claude")
gaps = [
    ("P0", "Clarify or build mechanical AI Diagnosis", "No OBD/symptom module — implement or drop claims"),
    ("P0", "Fix auth vs copy on /ai-valuation", "Hero no sign-up vs authGuard"),
    ("P0", "Show when heuristic fallback used", "Honest AI labeling"),
    ("P0", "Unify valuation engines", "One schema + one shared scorer"),
    ("P0", "Stop circular API heuristic", "Catalogue base not asking price"),
    ("P0", "Align List-car vs page fallbacks", "Same inputs → same mid"),
    ("P0", "Home AI Price Valuation → /ai-valuation", "Not /ai-advisor"),
    ("P0", "Rate-limit Edge ai-valuation", "Claude cost abuse"),
    ("P1", "Expose method/confidence/range in API+Next", "Float-only today"),
    ("P1", "Score condition + transmission", "Collected but ignored"),
    ("P1", "Fix ApiService.getAIValuation", "POST /listings/{id}/valuate"),
    ("P1", "Deterministic confidence", "No Math.random()"),
    ("P1", "Edge JSON fence stripping", "Before JSON.parse"),
    ("P1", "Next standalone valuation page", "Parity with Angular"),
    ("P1", "Honest marketing copy", "No fake 50,000+"),
    ("P2", "SEO + Playwright e2e", "/ai-valuation"),
]
for col, h in enumerate(["Priority", "Gap", "Notes"], 1):
    ws3.cell(1, col, h).font = Font(bold=True)
for i, row in enumerate(gaps, 2):
    for j, v in enumerate(row, 1):
        ws3.cell(i, j, v)
ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 48
ws3.column_dimensions["C"].width = 70

ws4 = wb.create_sheet("Claude Prompt")
md = QA / "AIDiagnosis-E2E-Gaps.md"
ws4["A1"] = "See AIDiagnosis-E2E-Gaps.md — Claude fix prompt section"
if md.exists():
    ws4["A2"] = md.read_text()[:30000]
    ws4["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[2].height = 420
ws4.column_dimensions["A"].width = 120

out = QA / "GAADIIQ_AIDiagnosis_Test_Results.xlsx"
wb.save(out)
wb.save(ART / out.name)
print("Wrote", out)
