#!/usr/bin/env python3
"""Full New Cars E2E results (Angular screenshot features + API + Next)."""
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(__file__).resolve().parent / "GAADIIQ_NewCars_Full_E2E_Results.xlsx"
ANG = Path("/opt/cursor/artifacts/new-cars/angular-e2e-scenarios.json")
API = Path("/opt/cursor/artifacts/new-cars/api-results-clean.json")

wb = Workbook()
header = PatternFill("solid", fgColor="1B3A6B")
pass_f = PatternFill("solid", fgColor="C6EFCE")
fail_f = PatternFill("solid", fgColor="FFC7CE")
partial_f = PatternFill("solid", fgColor="FFEB9C")
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ang = json.loads(ANG.read_text()) if ANG.exists() else {"passed": 0, "failed": 0, "partial": 0, "total": 0, "results": [], "gaps": []}
api = json.loads(API.read_text()) if API.exists() else {"passed": 0, "failed": 0, "total": 0, "results": [], "gaps": []}

ws = wb.active
ws.title = "Summary"
ws["A1"] = "GAADIIQ New Cars — Full Feature E2E Retest"
ws["A1"].font = Font(bold=True, size=16, color="1B3A6B")
ws["A2"] = "2026-07-18 | Angular /new-cars (screenshot features) + API + Next"
ws.merge_cells("A1:E1")

rows = [
    ("Suite", "PASS", "FAIL", "PARTIAL", "TOTAL"),
    ("Angular feature E2E (code+CDN probe)", ang["passed"], ang["failed"], ang.get("partial", 0), ang["total"]),
    ("API listing_type=new", api.get("passed", 0), api.get("failed", 0), 0, api.get("total", 0)),
    ("Next Playwright suite", 11, 0, 0, 11),
]
for r_i, row in enumerate(rows, start=4):
    for c_i, v in enumerate(row, start=1):
        cell = ws.cell(r_i, c_i, v)
        cell.border = thin
        if r_i == 4:
            cell.fill = header
            cell.font = Font(bold=True, color="FFFFFF")
        elif r_i > 4 and c_i == 2:
            cell.fill = pass_f
        elif r_i > 4 and c_i == 3 and v:
            cell.fill = fail_f
        elif r_i > 4 and c_i == 4 and v:
            cell.fill = partial_f

ws["A10"] = "Verdict"
ws["A10"].font = Font(bold=True)
ws["B10"] = (
    "Angular New Cars page sections RENDER, but E2E feature paths are largely broken: "
    "hotlinked images 404 + wrong placeholder path, View Details/Compare/Notify/Budget Above ₹30L/"
    "Electric-Luxury body filters miswired, launches/expert content static and inaccurate. "
    "API new filter works but seed is empty and km validation missing."
)
ws.merge_cells("B10:F10")
ws["B10"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[10].height = 60
for col, w in zip("ABCDE", [42, 10, 10, 12, 10]):
    ws.column_dimensions[col].width = w

# Angular results
aw = wb.create_sheet("Angular Feature E2E")
aw.append(["ID", "Feature / Scenario", "Status", "Detail", "Gap"])
for c in aw[1]:
    c.fill = header
    c.font = Font(bold=True, color="FFFFFF")
for r in ang.get("results", []):
    aw.append([r["id"], r["name"], r["status"], r.get("detail"), r.get("gap") or ""])
    st = aw.cell(aw.max_row, 3)
    st.fill = {"PASS": pass_f, "FAIL": fail_f, "PARTIAL": partial_f}.get(r["status"], partial_f)
for col, w in zip("ABCDE", [12, 42, 10, 55, 50]):
    aw.column_dimensions[col].width = w

# Section map from screenshots
sw = wb.create_sheet("Screenshot Sections")
sw.append(["Section (from user screenshots)", "Status", "Key findings"])
for c in sw[1]:
    c.fill = header
    c.font = Font(bold=True, color="FFFFFF")
sections = [
    ("Browse by Body Type", "PARTIAL", "UI OK; Electric/Luxury not handled on listings; icons misleading"),
    ("Browse by Budget", "FAIL", "Above ₹30L uses max=1Cr (shows almost all cars), not min≥30L"),
    ("Popular New Car Models + Filters", "FAIL", "Most images broken; badge overlap; min=max price; Compare unwired"),
    ("New Launches carousel", "FAIL", "CDN 404; 'last 3 months' false; date truncation; Swift missing NEW"),
    ("Upcoming Cars + Notify Me", "FAIL", "Images broken; Notify is local signal only — no email/DB"),
    ("Expert Recommendations", "FAIL", "Hardcoded; View Model only filters make, not model; not AI"),
    ("Hero By Brand / Budget / Body", "PASS", "Tabs present and navigate with carType=New"),
    ("Nav New Cars active + city Kolkata", "PASS", "Nav/chrome present in screenshots"),
]
for s in sections:
    sw.append(list(s))
    st = sw.cell(sw.max_row, 2)
    st.fill = {"PASS": pass_f, "FAIL": fail_f, "PARTIAL": partial_f}[s[1]]
for col, w in zip("ABC", [40, 10, 70]):
    sw.column_dimensions[col].width = w

# Gaps + Claude prompt sheet
gw = wb.create_sheet("Gaps + Claude Prompt")
gw.append(["Priority", "Gap", "Evidence", "Fix"])
for c in gw[1]:
    c.fill = header
    c.font = Font(bold=True, color="FFFFFF")
gaps = [
    ("P0", "Hotlinked aeplcdn images return 404", "CDN probe + screenshots", "Host images on own CDN/R2; stop third-party hotlinks"),
    ("P0", "Fallback path assets/placeholder.svg missing", "File not found; real is assets/cars/placeholder.svg", "Fix all (error) handlers + src fallbacks"),
    ("P0", "View Details ignores model query param", "new-cars passes model; listings never reads params['model']", "On init: if make+model set selectedModel or go /cars/:id"),
    ("P0", "Compare heart + Compare button unwired", "compareSet unused; /compare has no prefill", "Pass IDs via queryParams; hydrate compare slots"),
    ("P0", "Notify Me not persisted", "toggleNotify in-memory only", "Save to Supabase/API + require auth/email"),
    ("P0", "Budget Above ₹30L wrong", "max=100000000", "Use minPrice=3000000 (and support minPrice on listings)"),
    ("P0", "Electric/Luxury body types break filter", "listings.bodyTypes lacks them", "Map Electric→fuel=Electric; Luxury→minPrice; or add types"),
    ("P1", "New Launches date copy false", "May 2024 in 'last 3 months'", "Compute window or change subtitle; fix date pill CSS overflow"),
    ("P1", "Expert/Launch CTAs make-only", "queryParams make only", "Pass model; deep-link to variant list"),
    ("P1", "Price range ₹30L–₹30L", "min==max formatting", "Show '₹X onwards' when single price"),
    ("P1", "Card layout when image missing", "Badge overlaps title", "Fixed image aspect-ratio + reserved media box"),
    ("P1", "Expert picks not AI / not data-driven", "Hardcoded array", "Editorial CMS or recommend API; label honestly"),
    ("P1", "Body type iconography wrong", "SUV graph / MUV person in UI", "Use consistent vehicle icons/SVGs"),
    ("P2", "No free-text model search on New Cars", "Hero tabs only", "Add search → listings q= "),
    ("P2", "API seed has zero new listings", "seed.py used-only", "Seed new inventory for Next parity"),
]
for g in gaps:
    gw.append(list(g))
    gw.cell(gw.max_row, 1).fill = fail_f if g[0] == "P0" else partial_f
for col, w in zip("ABCD", [10, 45, 45, 50]):
    gw.column_dimensions[col].width = w

pw = wb.create_sheet("Claude Prompt")
prompt = """Fix ALL broken New Cars E2E features on Angular /new-cars (apps/gaadiiq-angular) based on docs/qa/NewCars-Full-E2E-Gaps.md and docs/qa/GAADIIQ_NewCars_Full_E2E_Results.xlsx.

P0 — must fix:
1. IMAGES: Stop hotlinking imgd.aeplcdn.com (returns 404). Use owned assets under assets/cars/ or R2 CDN. Fix every image error fallback from assets/placeholder.svg → assets/cars/placeholder.svg. Reserve fixed aspect-ratio media box so badges never overlap titles.
2. VIEW DETAILS: From Popular Models, open the correct model variant view. Either (a) navigate to /listings?carType=New&make=&model= AND make listings.component read params['model'] to call selectModel(), or (b) routerLink to /cars/{{representativeId}}.
3. COMPARE: Wire compare-heart selections into /compare?ids=... (or session storage) and prefill compare slots. Compare button must not open empty compare.
4. NOTIFY ME: Persist upcoming-car notifications (Supabase table or API) with user email; require login; do not use only an in-memory signal.
5. BUDGET: Fix "Above ₹30L" to filter minPrice >= 3000000 (add minPrice query support on listings). Keep other budget chips as maxPrice.
6. BODY TYPE: Electric should filter fuel=Electric; Luxury should filter minPrice (e.g. 30L+). Do not send bodyType=Electric/Luxury into a bodyType enum that only has Hatchback/Sedan/SUV/MUV.

P1:
7. New Launches: fix "last 3 months" filter/copy; fix date pill truncation; consistent NEW badge rules.
8. Expert Recommendations + Launch Explore: deep-link make+model (not make only). Label Expert as Editorial unless AI-backed.
9. Price display: if minPrice===maxPrice show "₹X.XL onwards".
10. Replace misleading body-type icons with vehicle SVGs.

P2: Add model search on New Cars; keep Playwright/static regression cases.

Acceptance:
- No broken images on Popular Models / Launches / Upcoming (placeholder OK if real asset missing).
- View Details lands on that model's variants.
- Compare from heart pre-fills cars.
- Notify Me survives refresh for logged-in user.
- Above ₹30L and Electric/Luxury filters return sensible results.
- Unit/e2e coverage for query-param wiring.
"""
pw["A1"] = prompt
pw["A1"].alignment = Alignment(wrap_text=True, vertical="top")
pw.column_dimensions["A"].width = 100
pw.row_dimensions[1].height = 420

wb.save(OUT)
print("Wrote", OUT)
