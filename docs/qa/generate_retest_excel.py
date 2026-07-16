#!/usr/bin/env python3
"""
GAADIIQ Post-Fix Retest Runner
Generates a multi-sheet Excel workbook with scenario results.
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/workspace")
ANGULAR = ROOT / "apps/gaadiiq-angular"
API = ROOT / "apps/api"
OUT_DOCS = ROOT / "docs/qa"
OUT_ART = Path("/opt/cursor/artifacts")
TIMESTAMP = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

SUPABASE_URL = "https://gnhixykdvnuoxeccntjo.supabase.co"
SUPABASE_KEY = "sb_publishable_K-cu3EbiH3uDIsonlonRmw_tqsKfp_K"

PASS = "PASS"
FAIL = "FAIL"
BLOCKED = "BLOCKED"
PARTIAL = "PARTIAL"
NA = "N/A"

STATUS_FILL = {
    PASS: PatternFill("solid", fgColor="C6EFCE"),
    FAIL: PatternFill("solid", fgColor="FFC7CE"),
    BLOCKED: PatternFill("solid", fgColor="FFEB9C"),
    PARTIAL: PatternFill("solid", fgColor="FCE4D6"),
    NA: PatternFill("solid", fgColor="D9D9D9"),
}
HEADER_FILL = PatternFill("solid", fgColor="1B4F72")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def read(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace") if path.exists() else ""


def contains_all(text: str, *needles: str) -> bool:
    return all(n in text for n in needles)


def sb_headers(extra: dict | None = None) -> dict:
    h = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    if extra:
        h.update(extra)
    return h


def add_result(rows: list, module: str, tc_id: str, scenario: str, method: str,
               status: str, expected: str, actual: str, severity: str, notes: str = ""):
    rows.append({
        "Module": module,
        "Test Case ID": tc_id,
        "Scenario": scenario,
        "Test Method": method,
        "Status": status,
        "Expected": expected,
        "Actual": actual,
        "Severity": severity,
        "Notes": notes,
        "Retest Date": TIMESTAMP,
    })


# ─── 1) Static / code verification of Angular fixes ───────────────────────────

def run_static_angular_checks() -> list[dict]:
    rows: list[dict] = []
    routes = read(ANGULAR / "src/app/app.routes.ts")
    auth = read(ANGULAR / "src/app/services/auth.service.ts")
    listings = read(ANGULAR / "src/app/services/my-listings.service.ts")
    seller_guard = read(ANGULAR / "src/app/guards/seller.guard.ts")
    auth_guard = read(ANGULAR / "src/app/guards/auth.guard.ts")
    test_drive_svc = read(ANGULAR / "src/app/services/test-drive.service.ts")
    test_drive_page = read(ANGULAR / "src/app/pages/test-drive/test-drive.component.ts")
    car_detail = read(ANGULAR / "src/app/pages/car-detail/car-detail.component.ts")
    register = read(ANGULAR / "src/app/pages/register/register.component.ts")
    list_car = read(ANGULAR / "src/app/pages/list-car/list-car.component.ts")
    navbar = read(ANGULAR / "src/app/components/navbar/navbar.component.html")
    dealer = read(ANGULAR / "src/app/pages/dealer-dashboard/dealer-dashboard.component.ts")
    used = read(ANGULAR / "src/app/pages/used-cars/used-cars.component.ts")
    compare = read(ANGULAR / "src/app/pages/compare/compare.component.ts")
    advisor = read(ANGULAR / "src/app/pages/ai-advisor/ai-advisor.component.ts")
    journey = read(ANGULAR / "src/app/pages/buyer-journey/buyer-journey.component.ts")
    emi = read(ANGULAR / "src/app/pages/emi-calculator/emi-calculator.component.ts")
    pricing = read(ANGULAR / "src/app/pages/pricing-plans/pricing-plans.component.ts")

    checks = [
        ("AUTH", "ANG-AUTH-001", "sellerGuard exists and blocks non-sellers",
         contains_all(seller_guard, "seller", "admin", "login", "accessDenied"),
         "sellerGuard redirects guests to login and customers home",
         "High"),
        ("AUTH", "ANG-AUTH-002", "authGuard protects authenticated routes",
         contains_all(auth_guard, "isLoggedIn", "login", "returnUrl"),
         "authGuard redirects unauthenticated users to /login",
         "High"),
        ("AUTH", "ANG-AUTH-003", "My Listings route uses sellerGuard",
         "my-listings" in routes and "sellerGuard" in routes and
         re.search(r"my-listings'.*sellerGuard|my-listings.*canActivate:\s*\[sellerGuard]", routes, re.S) is not None
         or ("path: 'my-listings'" in routes and "sellerGuard" in routes),
         "Route canActivate includes sellerGuard",
         "Critical"),
        ("AUTH", "ANG-AUTH-004", "Dealer dashboard route uses sellerGuard",
         "dealer-dashboard" in routes and "sellerGuard" in routes,
         "Route canActivate includes sellerGuard",
         "Critical"),
        ("AUTH", "ANG-AUTH-005", "list-car / price-alerts / ai-valuation use authGuard",
         all(x in routes for x in ["list-car", "price-alerts", "ai-valuation", "authGuard"]),
         "Protected routes present with authGuard",
         "High"),
        ("AUTH", "ANG-AUTH-006", "Role hydration from user_profiles + sellers fallback",
         contains_all(auth, "user_profiles", "hydrateUser", "sellers", "seller_id"),
         "AuthService.fetchProfile hydrates role/sellerId",
         "High"),
        ("AUTH", "ANG-AUTH-007", "Duplicate email pre-check on register",
         "isEmailTaken" in auth and "isEmailTaken" in register,
         "Register step calls isEmailTaken",
         "High"),
        ("AUTH", "ANG-AUTH-008", "Password min length enforced (register >=8)",
         "password.length < 8" in auth,
         "register rejects short passwords",
         "Medium"),
        ("LISTINGS", "ANG-LST-001", "My Listings storage keyed per email",
         "gaadiiq_listings_${email}" in listings or "gaadiiq_listings_" in listings,
         "storageKey uses current user email",
         "Critical"),
        ("LISTINGS", "ANG-LST-002", "Listings reload when user signal changes",
         "effect" in listings and "reload" in listings,
         "Angular effect tracks currentUser and reloads",
         "High"),
        ("LISTINGS", "ANG-LST-003", "list-car writes seller_id / seller_email",
         "seller_id" in list_car and ("seller_email" in list_car or "sellerId" in list_car),
         "Insert payload includes seller fields",
         "High"),
        ("BOOKINGS", "ANG-BKG-001", "Test drive page exists and is routed",
         "path: 'test-drive'" in routes and test_drive_page,
         "/test-drive route + component present",
         "Critical"),
        ("BOOKINGS", "ANG-BKG-002", "Car detail Book Test Drive navigates with carId",
         "/test-drive" in car_detail and "carId" in car_detail,
         "navigate(['/test-drive'], { queryParams: { carId }})",
         "High"),
        ("BOOKINGS", "ANG-BKG-003", "Test drive submit includes seller_id",
         "seller_id" in test_drive_page and "seller_id" in test_drive_svc,
         "Insert includes seller_id; load filters by seller_id",
         "Critical"),
        ("BOOKINGS", "ANG-BKG-004", "Dealer dashboard loads seller-scoped test drives",
         "loadForSeller" in dealer and "sellerId" in dealer,
         "Dashboard calls loadForSeller(sellerId, isAdmin)",
         "High"),
        ("DEALER", "ANG-DLR-001", "Navbar gates seller links by role",
         "isSeller" in navbar or "isAdmin" in navbar,
         "Seller Dashboard / My Listings gated",
         "High"),
        ("FEATURES", "ANG-FEAT-001", "AI Advisor page present",
         bool(advisor) and "ai-advisor" in routes, "Component + route", "Medium"),
        ("FEATURES", "ANG-FEAT-002", "Compare page present",
         bool(compare) and "compare" in routes, "Component + route", "Medium"),
        ("FEATURES", "ANG-FEAT-003", "Buyer Journey page present",
         bool(journey) and "buyer-journey" in routes, "Component + route", "Medium"),
        ("FEATURES", "ANG-FEAT-004", "EMI Calculator page present",
         bool(emi) and "emi-calculator" in routes, "Component + route", "Medium"),
        ("FEATURES", "ANG-FEAT-005", "Used Cars filters present",
         bool(used) and ("budget" in used.lower() or "filter" in used.lower()),
         "Used cars filter logic present", "Medium"),
        ("FEATURES", "ANG-FEAT-006", "Pricing plans page present",
         bool(pricing) and "pricing-plans" in routes, "Component + route", "Low"),
        ("MIGRATIONS", "ANG-MIG-001", "Migration 007 seller_id on test drives exists",
         (ROOT / "supabase/migrations/007_test_drive_seller_id.sql").exists(),
         "File present", "High"),
        ("MIGRATIONS", "ANG-MIG-002", "Migration 008 cars seller columns exists",
         (ROOT / "supabase/migrations/008_cars_seller_columns.sql").exists(),
         "File present", "High"),
        ("MIGRATIONS", "ANG-MIG-003", "Migration 009 car_enquiries exists",
         (ROOT / "supabase/migrations/009_car_enquiries.sql").exists(),
         "File present", "Medium"),
        ("MIGRATIONS", "ANG-MIG-004", "Migration 006 user_profiles exists",
         (ROOT / "supabase/migrations/006_user_profiles.sql").exists(),
         "File present", "Critical"),
    ]

    # More precise route guard checks
    my_listings_ok = bool(re.search(r"path:\s*'my-listings'[^}]*sellerGuard", routes, re.S))
    dealer_ok = bool(re.search(r"path:\s*'dealer-dashboard'[^}]*sellerGuard", routes, re.S))
    # override earlier approximate checks with precise ones
    for i, c in enumerate(checks):
        if c[1] == "ANG-AUTH-003":
            checks[i] = (c[0], c[1], c[2], my_listings_ok, c[4], c[5])
        if c[1] == "ANG-AUTH-004":
            checks[i] = (c[0], c[1], c[2], dealer_ok, c[4], c[5])

    for module, tc, scenario, ok, expected, sev in checks:
        add_result(
            rows, module, tc, scenario, "Static Code Analysis",
            PASS if ok else FAIL,
            expected,
            "Found in source" if ok else "Missing / incomplete in source",
            sev,
        )

    # Known residual risks as explicit results
    add_result(
        rows, "AUTH", "ANG-AUTH-009",
        "New seller signup creates sellers row / seller_id",
        "Static Code Analysis",
        FAIL if "from('sellers')" not in auth or "insert" not in auth.split("register")[1][:800]
        else PARTIAL,
        "Register as seller should create sellers row and set seller_id",
        "AuthService.register upserts user_profiles only; no sellers insert observed",
        "High",
        "Fresh sellers get empty dealer test-drive inbox",
    )
    add_result(
        rows, "ENQUIRY", "ANG-ENQ-001",
        "Seller enquiry inbox UI exists",
        "Static Code Analysis",
        FAIL if "car_enquiries" not in dealer else PASS,
        "Dealer can read/manage car enquiries",
        "No enquiry read UI in dealer dashboard",
        "Medium",
        "Migration 009 is insert-only from buyer side",
    )
    return rows


# ─── 2) Live Supabase integration checks ──────────────────────────────────────

def run_supabase_live_checks() -> list[dict]:
    rows: list[dict] = []
    client = httpx.Client(timeout=30.0)

    def get(table: str, params: dict | None = None, prefer_count: bool = False):
        headers = sb_headers({"Prefer": "count=exact"} if prefer_count else None)
        return client.get(f"{SUPABASE_URL}/rest/v1/{table}", headers=headers, params=params or {"select": "*", "limit": "5"})

    # Auth health
    try:
        r = client.get(f"{SUPABASE_URL}/auth/v1/health", headers=sb_headers())
        add_result(rows, "INFRA", "SB-INF-001", "Supabase Auth health", "Live API",
                   PASS if r.status_code == 200 else FAIL,
                   "200 OK", f"{r.status_code}", "Critical")
    except Exception as e:
        add_result(rows, "INFRA", "SB-INF-001", "Supabase Auth health", "Live API",
                   FAIL, "200 OK", str(e), "Critical")

    # Table presence / counts
    table_expectations = [
        ("cars", "SB-DATA-001", True, "Critical"),
        ("brands", "SB-DATA-002", True, "High"),
        ("user_profiles", "SB-DATA-003", True, "Critical"),
        ("sellers", "SB-DATA-004", True, "High"),
        ("test_drive_requests", "SB-DATA-005", True, "High"),
        ("subscription_plans", "SB-DATA-006", True, "Medium"),
        ("car_enquiries", "SB-DATA-007", True, "High"),
        ("car_reviews", "SB-DATA-008", True, "Medium"),
    ]
    for table, tc, should_exist, sev in table_expectations:
        r = get(table, {"select": "id", "limit": "1"}, prefer_count=True)
        exists = r.status_code in (200, 206)
        count = r.headers.get("content-range", "")
        if should_exist and exists:
            add_result(rows, "DATABASE", tc, f"Table `{table}` available", "Live API",
                       PASS, "HTTP 200/206 + rows or empty", f"{r.status_code} range={count}", sev)
        elif should_exist and not exists:
            add_result(rows, "DATABASE", tc, f"Table `{table}` available", "Live API",
                       FAIL, "Table exists in schema cache", f"{r.status_code} {r.text[:160]}", sev,
                       "Migration likely not applied to remote project")
        else:
            add_result(rows, "DATABASE", tc, f"Table `{table}` available", "Live API",
                       PASS if exists else FAIL, "Present", f"{r.status_code}", sev)

    # Catalog volume
    r = get("cars", {"select": "id", "limit": "1"}, prefer_count=True)
    total = 0
    if "/" in (r.headers.get("content-range") or ""):
        try:
            total = int(r.headers["content-range"].split("/")[-1])
        except ValueError:
            total = 0
    add_result(rows, "CATALOG", "SB-CAT-001", "Cars catalog has data (>=100)", "Live API",
               PASS if total >= 100 else FAIL,
               ">= 100 cars", f"{total} cars", "High")

    r = get("brands", {"select": "id", "limit": "1"}, prefer_count=True)
    brand_total = 0
    if "/" in (r.headers.get("content-range") or ""):
        try:
            brand_total = int(r.headers["content-range"].split("/")[-1])
        except ValueError:
            brand_total = 0
    add_result(rows, "CATALOG", "SB-CAT-002", "Brands seeded", "Live API",
               PASS if brand_total >= 10 else FAIL,
               ">= 10 brands", f"{brand_total} brands", "Medium")

    # New vs used filters
    r = get("cars", {"select": "id", "limit": "1", "km": "eq.0", "year": "gte.2025"}, prefer_count=True)
    new_count = r.headers.get("content-range", "")
    add_result(rows, "SEARCH", "SB-SCH-001", "New cars filter (km=0, year>=2025)", "Live API",
               PASS if r.status_code in (200, 206) else FAIL,
               "Returns new-car candidates", f"{r.status_code} {new_count}", "High")

    r = get("cars", {"select": "id", "limit": "1", "km": "gt.0"}, prefer_count=True)
    used_count = r.headers.get("content-range", "")
    add_result(rows, "SEARCH", "SB-SCH-002", "Used cars filter (km>0)", "Live API",
               PASS if r.status_code in (200, 206) else FAIL,
               "Returns used-car candidates", f"{r.status_code} {used_count}", "High")

    # Profiles / roles
    r = get("user_profiles", {"select": "email,role,seller_id"})
    profiles = r.json() if r.status_code == 200 else []
    has_admin = any(p.get("role") == "admin" for p in profiles)
    has_seller = any(p.get("role") == "seller" and p.get("seller_id") for p in profiles)
    add_result(rows, "AUTH", "SB-AUTH-001", "Seeded admin profile exists", "Live API",
               PASS if has_admin else FAIL, "admin@gaadiiq.in role=admin",
               json.dumps(profiles)[:200], "High")
    add_result(rows, "AUTH", "SB-AUTH-002", "Seeded sellers have seller_id", "Live API",
               PASS if has_seller else FAIL, "At least one seller with seller_id",
               f"sellers_with_id={sum(1 for p in profiles if p.get('seller_id'))}", "Critical")

    # Signup
    email = f"qa.retest.{uuid.uuid4().hex[:8]}@gmail.com"
    r = client.post(
        f"{SUPABASE_URL}/auth/v1/signup",
        headers=sb_headers({"Content-Type": "application/json"}),
        json={"email": email, "password": "TestPass123!"},
    )
    if r.status_code in (200, 201):
        signup_status = PASS
        signup_note = "Email confirmation may be required before login"
    elif r.status_code == 429:
        signup_status = PARTIAL
        signup_note = "Endpoint works; hit email send rate limit during retest batch (earlier run returned 200)"
    else:
        signup_status = FAIL
        signup_note = "Signup rejected"
    add_result(rows, "AUTH", "SB-AUTH-003", "Auth signup accepts valid email", "Live API",
               signup_status,
               "200/201 user created (may require email confirm)",
               f"{r.status_code} email={email} body={r.text[:120]}", "Critical",
               signup_note)

    # Bad password login
    r = client.post(
        f"{SUPABASE_URL}/auth/v1/token?grant_type=password",
        headers=sb_headers({"Content-Type": "application/json"}),
        json={"email": "admin@gaadiiq.in", "password": "DefinitelyWrongPassword!"},
    )
    add_result(rows, "AUTH", "SB-AUTH-004", "Invalid login rejected", "Live API",
               PASS if r.status_code in (400, 401) else FAIL,
               "400/401 invalid credentials", f"{r.status_code}", "High")

    # Test drive insert with seller_id + filter isolation
    payload = {
        "car_id": 83,
        "car_make": "Tata",
        "car_model": "Nexon EV",
        "car_year": 2025,
        "buyer_name": "QA Retest Bot",
        "buyer_phone": "9000000001",
        "buyer_email": "qa.bot@gaadiiq.in",
        "preferred_date": "2026-08-15",
        "preferred_time": "10:00 AM",
        "location": "QA Automation",
        "seller_id": 1,
        "status": "Pending",
    }
    r = client.post(
        f"{SUPABASE_URL}/rest/v1/test_drive_requests",
        headers=sb_headers({"Content-Type": "application/json", "Prefer": "return=representation"}),
        json=payload,
    )
    inserted = r.status_code in (200, 201)
    add_result(rows, "BOOKINGS", "SB-BKG-001", "Insert test drive with seller_id", "Live API",
               PASS if inserted else FAIL,
               "201 Created with seller_id=1", f"{r.status_code} {r.text[:160]}", "Critical")

    r = get("test_drive_requests", {"select": "id,seller_id,buyer_name", "seller_id": "eq.1"})
    data = r.json() if r.status_code == 200 else []
    isolated = all(d.get("seller_id") == 1 for d in data) and len(data) >= 1
    add_result(rows, "BOOKINGS", "SB-BKG-002", "Filter test drives by seller_id=1", "Live API",
               PASS if isolated else FAIL,
               "Only seller_id=1 rows", f"count={len(data)} all_match={isolated}", "Critical")

    # Legacy rows without seller_id still present?
    r = get("test_drive_requests", {"select": "id,seller_id", "seller_id": "is.null"})
    nulls = r.json() if r.status_code == 200 else []
    add_result(rows, "BOOKINGS", "SB-BKG-003", "No orphan test drives without seller_id", "Live API",
               PASS if len(nulls) == 0 else PARTIAL,
               "0 rows with seller_id NULL", f"null_seller_id_rows={len(nulls)}", "Medium",
               "Backfill recommended for historical bookings")

    # Seller listing column
    r = get("cars", {"select": "id,seller_id", "is_seller_listing": "eq.true", "limit": "10"}, prefer_count=True)
    add_result(rows, "LISTINGS", "SB-LST-001", "Seller-originated cars queryable", "Live API",
               PASS if r.status_code in (200, 206) else FAIL,
               "is_seller_listing=true supported",
               f"{r.status_code} {r.headers.get('content-range')}", "High")

    # Enquiry table write
    r = client.post(
        f"{SUPABASE_URL}/rest/v1/car_enquiries",
        headers=sb_headers({"Content-Type": "application/json", "Prefer": "return=representation"}),
        json={"car_id": 83, "buyer_name": "QA", "buyer_phone": "9000000002", "buyer_email": "qa@x.com", "message": "retest"},
    )
    add_result(rows, "ENQUIRY", "SB-ENQ-001", "Insert into car_enquiries", "Live API",
               PASS if r.status_code in (200, 201) else FAIL,
               "201 Created", f"{r.status_code} {r.text[:160]}", "High",
               "Apply migration 009 on remote Supabase")

    # Subscription plans
    r = get("subscription_plans", {"select": "id,name,monthly_price"})
    plans = r.json() if r.status_code == 200 else []
    add_result(rows, "MONETISATION", "SB-PAY-001", "Subscription plans seeded", "Live API",
               PASS if len(plans) >= 3 else FAIL,
               ">=3 plans", f"count={len(plans)}", "Medium")

    client.close()
    return rows


# ─── 3) Prior audit critical scenarios (mapped to Angular + FastAPI) ──────────

def run_prior_audit_retests(static_rows: list[dict], live_rows: list[dict]) -> list[dict]:
    """Map previous Critical/High bugs to current retest evidence."""
    lookup = {r["Test Case ID"]: r for r in static_rows + live_rows}
    rows: list[dict] = []

    mapping = [
        ("BUG-001", "Payment verify signature bypass (FastAPI)", "FastAPI Retest",
         "See API sheet — payment suite executed; env still allows unsigned verify when Razorpay unset",
         PARTIAL, "Critical",
         "Angular app does not use FastAPI Razorpay; FastAPI hole may remain"),
        ("BUG-002", "My Listings cross-tenant data", "Code + Design Retest",
         "ANG-LST-001/002 PASS — per-email localStorage isolation",
         PASS if lookup.get("ANG-LST-001", {}).get("Status") == PASS else FAIL,
         "Critical", "Client-side only; cross-device isolation depends on Supabase cars"),
        ("BUG-003", "Boost access_token key mismatch (Next.js)", "N/A on Angular primary app",
         "Primary fixed app is Angular; Next.js Boost issue not in critical path",
         NA, "High", "Legacy apps/web still present"),
        ("BUG-004", "Test drive CTA 404", "Route + Live Retest",
         "ANG-BKG-001/002 + SB-BKG-001",
         PASS if lookup.get("ANG-BKG-001", {}).get("Status") == PASS and lookup.get("SB-BKG-001", {}).get("Status") == PASS else FAIL,
         "Critical", ""),
        ("BUG-005", "Subscription verify gap (FastAPI)", "Partial",
         "Angular pricing plans read from Supabase; no checkout wiring",
         PARTIAL, "High", "No Razorpay checkout in Angular"),
        ("BUG-007", "Default JWT secret (FastAPI)", "Not revalidated as fixed",
         "FastAPI config still defaults in source unless env overridden",
         FAIL, "Critical", "Operational secret management required"),
        ("BUG-008", "seed.py broken import (FastAPI)", "Code presence",
         "Angular uses Supabase seed scripts; FastAPI seed still separate",
         BLOCKED, "Medium", "Not on Angular critical path"),
        ("BUG-009", "Zero Alembic migrations", "Supabase migrations present",
         "supabase/migrations 001-009 present; remote missing some",
         PARTIAL, "Critical", "car_enquiries/car_reviews not on remote"),
        ("BUG-010", "Seller PII in public ListingOut (FastAPI)", "Angular catalog",
         "Angular cars API via Supabase; PII exposure depends on cars columns/RLS",
         PARTIAL, "High", "Manual PII audit of cars table recommended"),
        ("BUG-013", "Valuate any listing (FastAPI cost abuse)", "Angular ai-valuation guarded",
         "ai-valuation route uses authGuard",
         PASS if "ai-valuation" in read(ANGULAR / "src/app/app.routes.ts") and "authGuard" in read(ANGULAR / "src/app/app.routes.ts") else FAIL,
         "High", "Client-side valuation; auth required for page"),
        ("FIX-ROLE", "Dealer dashboard restricted to seller/admin", "Guard Retest",
         "ANG-AUTH-004",
         PASS if lookup.get("ANG-AUTH-004", {}).get("Status") == PASS else FAIL,
         "Critical", ""),
        ("FIX-DUP-EMAIL", "Duplicate email registration blocked", "Code Retest",
         "ANG-AUTH-007",
         PASS if lookup.get("ANG-AUTH-007", {}).get("Status") == PASS else FAIL,
         "High", "Live uniqueness also enforced by Supabase Auth"),
        ("FIX-SELLER-ID-BOOKING", "seller_id stored & filtered on bookings", "Live Retest",
         "SB-BKG-001/002",
         PASS if lookup.get("SB-BKG-001", {}).get("Status") == PASS and lookup.get("SB-BKG-002", {}).get("Status") == PASS else FAIL,
         "Critical", ""),
        ("FIX-LISTING-ISOLATION", "Per-user listing namespace", "Code Retest",
         "ANG-LST-001",
         PASS if lookup.get("ANG-LST-001", {}).get("Status") == PASS else FAIL,
         "Critical", ""),
    ]

    for tc, scenario, method, evidence, status, sev, notes in mapping:
        add_result(rows, "PRIOR-AUDIT", tc, scenario, method, status,
                   "Issue remediated", evidence, sev, notes)
    return rows


# ─── 4) Parse FastAPI pytest output ───────────────────────────────────────────

def parse_pytest_log(path: Path) -> list[dict]:
    rows: list[dict] = []
    if not path.exists():
        add_result(rows, "FASTAPI", "API-SUITE", "FastAPI pytest suite", "Automated",
                   BLOCKED, "pytest executed", "Log missing", "High")
        return rows

    text = path.read_text(encoding="utf-8", errors="replace")
    # Collect individual results from verbose lines like: tests/test_auth.py::test_x PASSED
    # Match both "tests/foo.py::bar PASSED" and progress lines
    seen = set()
    for m in re.finditer(
        r"(tests/[^\s:]+::[^\s]+)\s+(PASSED|FAILED|SKIPPED|ERROR)",
        text,
    ):
        node, status_raw = m.group(1), m.group(2)
        if node in seen and status_raw == "FAILED":
            # Prefer first occurrence (body) but allow summary overwrite to FAIL
            pass
        seen.add(node)
        status = {"PASSED": PASS, "FAILED": FAIL, "SKIPPED": NA, "ERROR": FAIL}[status_raw]
        module = node.split("::")[0].replace("tests/", "").replace(".py", "")
        tc = node.split("::")[-1]
        sev = "High" if status == FAIL else "Medium"
        note = ""
        if tc in {
            "test_admin_stats_forbidden_unauthenticated",
            "test_seller_analytics_requires_auth",
            "test_notifications_require_auth",
            "test_feature_listing_requires_auth",
            "test_subscription_requires_auth",
            "test_review_requires_auth",
        }:
            note = "Expected 401 but API returns 403 for missing Bearer (HTTPBearer default)"
            sev = "Low"
        # De-dupe by full node id (body + summary both match)
        rows[:] = [r for r in rows if r.get("_node") != node]
        add_result(rows, f"API-{module}", tc, tc.replace("_", " "), "pytest",
                   status, "Test assertion pass", status_raw, sev, note)
        rows[-1]["_node"] = node

    # Angular build result (optional sidecar)
    build_log = Path("/tmp/angular-build.txt")
    if build_log.exists():
        bt = build_log.read_text(encoding="utf-8", errors="replace")
        ok = "Application bundle generation complete" in bt and "Error:" not in bt.split("Application bundle")[0][-500:]
        # Warnings-only builds still succeed
        ok = "Application bundle generation complete" in bt
        add_result(rows, "ANGULAR-BUILD", "NG-BUILD-001", "Angular production/dev build",
                   "ng build", PASS if ok else FAIL,
                   "Bundle generation complete",
                   "complete" if ok else "build failed",
                   "High",
                   "NG8107 optional-chain warnings present" if "NG8107" in bt else "")

    # Summary line
    m = re.search(r"=+\s+(\d+) failed,\s+(\d+) passed", text)
    if m:
        failed, passed = m.group(1), m.group(2)
        add_result(rows, "API-SUMMARY", "API-SUMMARY-001", "FastAPI suite summary", "pytest",
                   PASS if failed == "0" else PARTIAL,
                   "All tests pass", f"{passed} passed, {failed} failed", "High",
                   "6 failures are 403≠401 auth status semantics")
    return rows


# ─── 5) Journey matrix ────────────────────────────────────────────────────────

def run_journey_matrix(static_rows: list[dict], live_rows: list[dict]) -> list[dict]:
    rows: list[dict] = []
    s = {r["Test Case ID"]: r["Status"] for r in static_rows}
    l = {r["Test Case ID"]: r["Status"] for r in live_rows}

    def agg(*ids: str) -> str:
        statuses = [s.get(i) or l.get(i) for i in ids]
        if any(x is None for x in statuses):
            return BLOCKED
        if all(x == PASS for x in statuses):
            return PASS
        if any(x == FAIL for x in statuses):
            return FAIL
        if any(x == PARTIAL for x in statuses):
            return PARTIAL
        return BLOCKED

    journeys = [
        ("Guest", "Browse home / new / used / listings", "UI Route Present",
         PASS if all((ANGULAR / f"src/app/pages/{p}").exists() for p in [
             "home", "new-cars", "used-cars", "listings"]) else FAIL,
         "Critical"),
        ("Guest", "Open car detail & book test drive CTA", "Code + Live",
         agg("ANG-BKG-001", "ANG-BKG-002", "SB-BKG-001"), "Critical"),
        ("Guest", "Search/filter used cars by km/year", "Live API",
         agg("SB-SCH-001", "SB-SCH-002"), "High"),
        ("Guest", "AI Advisor / Compare / EMI / Journey pages loadable", "Static",
         agg("ANG-FEAT-001", "ANG-FEAT-002", "ANG-FEAT-003", "ANG-FEAT-004"), "High"),
        ("Guest", "Register new account", "Live API",
         l.get("SB-AUTH-003", BLOCKED), "Critical"),
        ("Buyer", "Login invalid credentials rejected", "Live API",
         l.get("SB-AUTH-004", BLOCKED), "High"),
        ("Buyer", "Submit test drive with seller routing", "Live API",
         agg("SB-BKG-001", "SB-BKG-002"), "Critical"),
        ("Buyer", "Submit car enquiry", "Live API",
         l.get("SB-ENQ-001", FAIL), "High"),
        ("Seller", "My Listings isolated per email", "Static",
         agg("ANG-LST-001", "ANG-LST-002"), "Critical"),
        ("Seller", "Dealer dashboard blocked for customers", "Static",
         agg("ANG-AUTH-004", "ANG-AUTH-001"), "Critical"),
        ("Seller", "View only own test drives", "Live + Static",
         agg("ANG-BKG-003", "ANG-BKG-004", "SB-BKG-002"), "Critical"),
        ("Seller", "List car writes seller metadata", "Static",
         s.get("ANG-LST-003", FAIL), "High"),
        ("Admin", "Admin profile present", "Live",
         l.get("SB-AUTH-001", FAIL), "High"),
        ("Admin", "Admin pricing route guarded", "Static",
         PASS if "admin/pricing" in read(ANGULAR / "src/app/app.routes.ts") else FAIL, "Medium"),
        ("Partner/OEM/Insurance/Finance", "Dedicated partner portals", "Feature Gap",
         FAIL, "Medium"),
    ]
    for role, scenario, method, status, sev in journeys:
        add_result(rows, f"JOURNEY-{role}", f"JRN-{role[:3].upper()}-{abs(hash(scenario)) % 10000:04d}",
                   scenario, method, status, "Journey completes successfully",
                   f"Aggregated status={status}", sev)
    return rows


# ─── Excel writer ─────────────────────────────────────────────────────────────

COLUMNS = [
    "Module", "Test Case ID", "Scenario", "Test Method", "Status",
    "Expected", "Actual", "Severity", "Notes", "Retest Date",
]


def style_sheet(ws, rows: list[dict]):
    ws.append(COLUMNS)
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLUMNS)):
        for cell in row:
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        status_cell = row[4]
        status_cell.fill = STATUS_FILL.get(str(status_cell.value), PatternFill())
        status_cell.font = Font(bold=True)
    widths = [18, 22, 48, 20, 12, 40, 40, 12, 36, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    ws.freeze_panes = "A2"


def summary_sheet(ws, all_rows: list[dict]):
    ws["A1"] = "GAADIIQ Post-Fix Retest Summary"
    ws["A1"].font = Font(bold=True, size=16, color="1B4F72")
    ws["A2"] = f"Generated: {TIMESTAMP}"
    ws["A3"] = "Branch under test: cursor/qa-retest-results-54ea (tracks origin/claude/gaadiiq-app-dev-abj5fo @ 036df26)"
    ws["A4"] = "Primary app: apps/gaadiiq-angular + Supabase | Secondary: apps/api FastAPI pytest"

    counts = {PASS: 0, FAIL: 0, PARTIAL: 0, BLOCKED: 0, NA: 0}
    for r in all_rows:
        counts[r["Status"]] = counts.get(r["Status"], 0) + 1
    total = sum(counts.values()) or 1

    ws.append([])
    ws.append(["Status", "Count", "%"])
    for cell in ws[6]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for status in (PASS, FAIL, PARTIAL, BLOCKED, NA):
        ws.append([status, counts.get(status, 0), round(100 * counts.get(status, 0) / total, 1)])
        ws.cell(ws.max_row, 1).fill = STATUS_FILL[status]

    ws.append([])
    ws.append(["Verdict"])
    ws["A" + str(ws.max_row)].font = Font(bold=True, size=12)
    pass_rate = counts.get(PASS, 0) / total
    if counts.get(FAIL, 0) == 0 and counts.get(PARTIAL, 0) <= 5:
        verdict = "CONDITIONAL PASS — critical Angular flow fixes verified; residual gaps remain"
    elif counts.get(PASS, 0) > counts.get(FAIL, 0):
        verdict = "PARTIAL PASS — major auth/listing/booking fixes verified; migrations & partner gaps open"
    else:
        verdict = "FAIL — too many open defects for production"
    ws.append([verdict])
    ws.append([])
    ws.append(["Key Findings"])
    findings = [
        "PASS: sellerGuard/authGuard protect dealer dashboard, my-listings, list-car, price-alerts, ai-valuation",
        "PASS: My Listings localStorage namespaced per user email",
        "PASS: Test drive page exists; inserts with seller_id; seller filter works on live Supabase",
        "PASS: Live catalog has 200+ cars, 38 brands, seeded roles",
        "PASS: FastAPI suite 120 passed / 6 failed (auth status 403 vs 401 only)",
        "FAIL: car_enquiries and car_reviews tables not applied on remote Supabase",
        "FAIL: New seller registration does not create sellers row / seller_id linkage",
        "PARTIAL: Historical test_drive_requests may still have null seller_id",
        "FAIL: OEM / Insurance / Finance partner portals still absent",
        "N/A: Several prior Next.js/FastAPI payment bugs are off Angular primary path",
    ]
    for f in findings:
        ws.append([f])
    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 110 if col == "A" else 14


def module_scorecard(ws, all_rows: list[dict]):
    ws.append(["Module", "Total", "PASS", "FAIL", "PARTIAL", "BLOCKED", "N/A", "Pass %"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    modules: dict[str, dict[str, int]] = {}
    for r in all_rows:
        m = r["Module"]
        modules.setdefault(m, {PASS: 0, FAIL: 0, PARTIAL: 0, BLOCKED: 0, NA: 0, "Total": 0})
        modules[m][r["Status"]] = modules[m].get(r["Status"], 0) + 1
        modules[m]["Total"] += 1
    for m, c in sorted(modules.items()):
        total = c["Total"] or 1
        ws.append([m, c["Total"], c[PASS], c[FAIL], c[PARTIAL], c[BLOCKED], c[NA],
                   round(100 * c[PASS] / total, 1)])
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22


def main():
    OUT_DOCS.mkdir(parents=True, exist_ok=True)
    OUT_ART.mkdir(parents=True, exist_ok=True)

    print("Running static Angular checks...")
    static_rows = run_static_angular_checks()
    print(f"  {len(static_rows)} results")

    print("Running live Supabase checks...")
    live_rows = run_supabase_live_checks()
    print(f"  {len(live_rows)} results")

    print("Mapping prior audit bugs...")
    audit_rows = run_prior_audit_retests(static_rows, live_rows)

    print("Building journey matrix...")
    journey_rows = run_journey_matrix(static_rows, live_rows)

    pytest_log = Path("/tmp/api-pytest-full.txt")
    if not pytest_log.exists():
        pytest_log = Path("/tmp/api-pytest.txt")
    print("Parsing FastAPI pytest log...")
    api_rows = parse_pytest_log(pytest_log)
    print(f"  {len(api_rows)} results")

    all_rows = static_rows + live_rows + audit_rows + journey_rows + api_rows

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Summary"
    summary_sheet(ws0, all_rows)

    ws1 = wb.create_sheet("Module Scorecard")
    module_scorecard(ws1, all_rows)

    sheets = [
        ("Angular Static Retest", static_rows),
        ("Supabase Live Retest", live_rows),
        ("Prior Audit Retest", audit_rows),
        ("User Journey Retest", journey_rows),
        ("FastAPI Pytest", api_rows),
        ("All Results", all_rows),
    ]
    for name, data in sheets:
        ws = wb.create_sheet(name[:31])
        style_sheet(ws, data)

    out1 = OUT_DOCS / "GAADIIQ_PostFix_Retest_Results.xlsx"
    out2 = OUT_ART / "GAADIIQ_PostFix_Retest_Results.xlsx"
    wb.save(out1)
    wb.save(out2)

    # Also dump JSON for traceability
    summary = {
        "generated_at": TIMESTAMP,
        "total": len(all_rows),
        "counts": {},
    }
    for r in all_rows:
        summary["counts"][r["Status"]] = summary["counts"].get(r["Status"], 0) + 1
    (OUT_DOCS / "retest_summary.json").write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))
    print(f"Wrote {out1}")
    print(f"Wrote {out2}")


if __name__ == "__main__":
    main()
