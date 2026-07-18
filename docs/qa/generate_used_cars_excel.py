#!/usr/bin/env python3
"""Generate Used Cars E2E workbook + embed Claude prompt."""
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = Path("/opt/cursor/artifacts/used-cars/used-cars-scenarios.json")
OUT = Path(__file__).resolve().parent / "GAADIIQ_UsedCars_Test_Results.xlsx"

wb = Workbook()
header = PatternFill("solid", fgColor="1B3A6B")
pass_f = PatternFill("solid", fgColor="C6EFCE")
fail_f = PatternFill("solid", fgColor="FFC7CE")
partial_f = PatternFill("solid", fgColor="FFEB9C")
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

data = json.loads(SRC.read_text()) if SRC.exists() else {"passed": 0, "failed": 0, "total": 0, "results": [], "gaps": []}

ws = wb.active
ws.title = "Summary"
ws["A1"] = "GAADIIQ Used Cars — Full Scenario Retest"
ws["A1"].font = Font(bold=True, size=16, color="1B3A6B")
ws["A2"] = "2026-07-18 | Angular /used-cars (screenshot) + API listing_type=used"
ws.merge_cells("A1:E1")

rows = [
    ("Suite", "PASS", "FAIL", "TOTAL"),
    ("Scenario matrix (logic + wiring + API)", data["passed"], data["failed"], data["total"]),
]
for r_i, row in enumerate(rows, start=4):
    for c_i, v in enumerate(row, start=1):
        cell = ws.cell(r_i, c_i, v)
        cell.border = thin
        if r_i == 4:
            cell.fill = header
            cell.font = Font(bold=True, color="FFFFFF")
        elif r_i > 4 and c_i == 2:
            cell.fill = pass_f
        elif r_i > 4 and c_i == 3 and v:
            cell.fill = fail_f

ws["A8"] = "Screenshot verdict"
ws["A8"].font = Font(bold=True)
ws["B8"] = (
    "Page shows 0+ Used Cars / 0 results while MODEL=RITZ and Make=All. "
    "Root cause: free-text model filter (browser autocomplete) + hero badge using FILTERED count; "
    "also possible empty Supabase used inventory. Marketing claims 50,000+ cars — contradictory. "
    "Filter logic itself (km/fuel/owner) works when data + clean search exist."
)
ws.merge_cells("B8:F8")
ws["B8"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[8].height = 60
for col, w in zip("ABCD", [45, 12, 12, 12]):
    ws.column_dimensions[col].width = w

# Results
rw = wb.create_sheet("All Scenarios")
rw.append(["ID", "Scenario", "Status", "Detail", "Gap"])
for c in rw[1]:
    c.fill = header
    c.font = Font(bold=True, color="FFFFFF")
for r in data.get("results", []):
    rw.append([r["id"], r["name"], r["status"], r.get("detail"), r.get("gap") or ""])
    st = rw.cell(rw.max_row, 3)
    st.fill = pass_f if r["status"] == "PASS" else fail_f
for col, w in zip("ABCDE", [12, 48, 10, 50, 55]):
    rw.column_dimensions[col].width = w

# Product scorecard (user-facing)
sw = wb.create_sheet("Product Scorecard")
sw.append(["Feature", "Status", "Notes"])
for c in sw[1]:
    c.fill = header
    c.font = Font(bold=True, color="FFFFFF")
score = [
    ("Nav → /used-cars", "PASS", "Route + navbar OK"),
    ("Hero search City/Make/Model/Budget", "FAIL", "Model free-text + RITZ autocomplete → 0 results; city not from navbar"),
    ("Hero badge inventory count", "FAIL", "Shows filtered totalCount → 0+ when search empty"),
    ("Results list / sort / load more", "PARTIAL", "Works when data exists; empty in screenshot"),
    ("Sidebar filters (km/fuel/tx/body/owners/color/certified)", "PASS", "Logic verified in simulator"),
    ("Wishlist / recently viewed", "PASS", "localStorage"),
    ("View Details → /cars/:id", "PASS", "Wired"),
    ("AI-Verified price badge", "PARTIAL", "Heuristic vs avg — not real AI valuation"),
    ("Images / placeholder", "FAIL", "assets/placeholder.svg missing"),
    ("Marketing 50,000+", "FAIL", "Hardcoded; contradicts 0 count"),
    ("API listing_type=used (Next)", "PASS", "Create + filter + city + make"),
    ("E2E UI automation", "FAIL", "No Playwright coverage for /used-cars"),
]
for row in score:
    sw.append(list(row))
    st = sw.cell(sw.max_row, 2)
    st.fill = {"PASS": pass_f, "FAIL": fail_f, "PARTIAL": partial_f}[row[1]]
for col, w in zip("ABC", [48, 10, 60]):
    sw.column_dimensions[col].width = w

# Gaps for Claude
gw = wb.create_sheet("Gaps for Claude")
gw.append(["Priority", "Gap", "Evidence", "Fix"])
for c in gw[1]:
    c.fill = header
    c.font = Font(bold=True, color="FFFFFF")
gaps = [
    ("P0", "Empty results with MODEL=RITZ / Make=All", "Screenshot 0 cars", "Disable autocomplete; cascading make→model; Clear search; reset model on make change"),
    ("P0", "Hero badge shows filtered count (0+)", "totalCount() in badge", "Badge = unfiltered used inventory; results bar keeps filtered count"),
    ("P0", "Supabase used inventory may be empty", "0 cars even after clear possible", "Ensure seed/demo used cars (km>0 or year<2025 or seller listings) in Supabase"),
    ("P0", "Wrong image placeholder path", "assets/placeholder.svg missing", "Use assets/cars/placeholder.svg"),
    ("P0", "Hardcoded 50,000+ listed cars", "Stats banner", "Use live count or remove claim"),
    ("P1", "Navbar city not applied", "Kolkata in nav, hero All India", "Prefill heroCity from CityService"),
    ("P1", "Any Budget capped at ₹50L", "value=5000000", "Open ceiling (e.g. 2Cr) or null max"),
    ("P1", "yearTo max 2025", "yearOptions", "Extend to currentYear(+1)"),
    ("P1", "clearAllFilters misses heroBudgetMax", "UC-018/019 FAIL", "Reset heroBudgetMax signal"),
    ("P1", "AI-Verified is heuristic", "avgUsedPrice verdict", "Use ai_valuation or relabel"),
    ("P1", "No shareable ?make&model&city URL", "queryParams limited", "Read/write search params"),
    ("P1", "Align used definition with listings", "isSellerListing only on used-cars", "Single isUsedCar helper shared"),
    ("P2", "Add e2e coverage", "No tests", "Playwright: load, clear filters, assert count, city/budget"),
    ("P2", "Model datalist / typeahead", "Free text only", "Suggest models from inventory for selected make"),
]
for g in gaps:
    gw.append(list(g))
    gw.cell(gw.max_row, 1).fill = fail_f if g[0] == "P0" else partial_f
for col, w in zip("ABCD", [10, 48, 40, 55]):
    gw.column_dimensions[col].width = w

pw = wb.create_sheet("Claude Prompt")
prompt = """Fix GAADIIQ Angular Used Cars module (/used-cars) so users get the car list matching their request.
Read: docs/qa/UsedCars-E2E-Gaps.md and docs/qa/GAADIIQ_UsedCars_Test_Results.xlsx.

Screenshot bugs: badge "0+ Used Cars Available", MODEL=RITZ with Make=All, "0 used cars found", while banner claims "50,000+".

P0 — must fix:
1. SEARCH DEFAULTS — Stop empty-state from browser autocomplete. autocomplete="off" on model input; cascading Make→Model (reset model when make changes); Clear Search button that clears heroMake/heroModel/heroCity and budget.
2. HERO BADGE — Show UNFILTERED used inventory count in hero badge. Keep filtered count only in "X used cars found".
3. DATA — Ensure Supabase/demo has used cars (km>0 OR year<2025 OR is_seller_listing). If empty, show honest empty state ("No used cars listed yet") not "0+ Available" + fake 50,000+.
4. IMAGES — Replace assets/placeholder.svg → assets/cars/placeholder.svg on cards + recently viewed.
5. MARKETING — Replace hardcoded 50,000+ with live used count (or remove).

P1:
6. Prefill heroCity from CityService (navbar Kolkata).
7. Any Budget: no max (or ≥2Cr), not 50L cap.
8. yearTo through current year (+1); include in yearOptions.
9. clearAllFilters also resets heroBudgetMax.
10. Support query params make/model/city/minBudget; sync URL on search.
11. Relabel AI price verdict or wire ai_valuation.
12. Share isUsedCar helper with listings Used tab.

P2: Playwright e2e for /used-cars (load → clear → count>0 when seeded; city + budget filters).

Acceptance:
- Fresh /used-cars with no sticky filters shows used inventory count >0 when data exists (not 0+).
- MODEL empty by default; selecting Make filters models; RITZ alone cannot stick from autocomplete.
- Clear All restores full used list.
- City from navbar applies when set.
- No broken placeholder path; no false 50,000+ when count is 0.
"""
pw["A1"] = prompt
pw["A1"].alignment = Alignment(wrap_text=True, vertical="top")
pw.column_dimensions["A"].width = 100
pw.row_dimensions[1].height = 420

wb.save(OUT)
print("Wrote", OUT)
