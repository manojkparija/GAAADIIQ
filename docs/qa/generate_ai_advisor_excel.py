#!/usr/bin/env python3
"""Generate GAADIIQ_AIAdvisor_Test_Results.xlsx from scenario JSON + API smoke."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
QA = ROOT / "docs/qa"
ART = Path("/opt/cursor/artifacts/ai-advisor")

scenarios = json.loads((QA / "ai-advisor-scenarios.json").read_text())
api = json.loads((QA / "api-recommend-smoke.json").read_text())

wb = Workbook()

# ── Summary ──────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"
ws["A1"] = "GAADIIQ AI Advisor — Full Feature Retest"
ws["A1"].font = Font(bold=True, size=14)
ws["A3"] = "Primary surface (screenshot)"
ws["B3"] = "Angular /ai-advisor — 11-step quiz + client rule-engine"
ws["A4"] = "Secondary surface"
ws["B4"] = "Next.js /recommend — 4-step wizard"
ws["A5"] = "API"
ws["B5"] = "POST /recommend (rule-engine; no LLM; no ai-chat)"
ws["A6"] = "Generated"
ws["B6"] = scenarios["generated_at"]
ws["A7"] = "Seed cars"
ws["B7"] = scenarios["seed_cars"]
ws["A8"] = "Brand New filter count"
ws["B8"] = scenarios["brand_new_filter"]
ws["A9"] = "Certified Used filter count"
ws["B9"] = scenarios["certified_used_filter"]

counts = scenarios["counts"]
ws["A11"] = "Scenario matrix"
ws["B11"] = (
    f"{counts.get('PASS', 0)} PASS / {counts.get('FAIL', 0)} FAIL / "
    f"{counts.get('PARTIAL', 0)} PARTIAL  (total {scenarios['total']})"
)
api_pass = sum(1 for x in api if x["status"] == "PASS")
ws["A12"] = "API POST /recommend smoke"
ws["B12"] = f"{api_pass}/{len(api)} PASS"
ws["A13"] = "Verdict"
ws["B13"] = (
    "Quiz UX + client scoring work for common profiles, but product is NOT real AI: "
    "fake analyzing theater, unused answers (drivingMix/ev), Next ignores POST /recommend & usage, "
    "no ai-chat, dual divergent surfaces."
)
ws["B13"].alignment = Alignment(wrap_text=True)
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 100

# ── Results ──────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Scenario Results")
headers = ["Suite", "ID", "Name", "Status", "Severity", "Detail"]
fills = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
    "PARTIAL": PatternFill("solid", fgColor="FFEB9C"),
}
for col, h in enumerate(headers, 1):
    cell = ws2.cell(1, col, h)
    cell.font = Font(bold=True)
for i, r in enumerate(scenarios["results"], 2):
    ws2.cell(i, 1, r["suite"])
    ws2.cell(i, 2, r["id"])
    ws2.cell(i, 3, r["name"])
    cell = ws2.cell(i, 4, r["status"])
    cell.fill = fills.get(r["status"], PatternFill())
    ws2.cell(i, 5, r.get("severity", "—"))
    ws2.cell(i, 6, r["detail"])
for col, w in enumerate([14, 10, 48, 10, 10, 90], 1):
    ws2.column_dimensions[chr(64 + col)].width = w

# ── API ──────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("API Recommend Smoke")
for col, h in enumerate(["ID", "Status", "Payload", "Detail"], 1):
    ws3.cell(1, col, h).font = Font(bold=True)
for i, r in enumerate(api, 2):
    ws3.cell(i, 1, r["id"])
    cell = ws3.cell(i, 2, r["status"])
    cell.fill = fills.get(r["status"], PatternFill())
    ws3.cell(i, 3, json.dumps(r["payload"]))
    ws3.cell(i, 4, r["detail"])
for col, w in enumerate([12, 10, 55, 70], 1):
    ws3.column_dimensions[chr(64 + col)].width = w

# ── Gaps ─────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Gaps for Claude")
gaps = [
    ("P0", "Wire Next /recommend to POST /recommend", "Show match_score + reasons; stop GET /listings-only"),
    ("P0", "Use usage in Next + API path", "Today usage is collected and ignored on Next"),
    ("P0", "Honest AI labeling", "Replace fake 'Running AI recommendation engine…' or call real LLM"),
    ("P0", "Empty results state (Angular)", "When top=[] show empty CTA; today blank results page"),
    ("P0", "Hard budget filter when no fuel pref", "SC-04 Under ₹5L still returns ₹6–8L cars (budget leak)"),
    ("P0", "Align Brand New / Certified Used", "Use listing_type / km=0 / certified badge — not year>=cy-1 & km<15k only"),
    ("P0", "Implement or defer POST /recommend/ai-chat", "LLD promises SSE chat; missing"),
    ("P0", "Unify Angular vs Next advisors", "Shared engine or clearly brand one as primary"),
    ("P1", "Fix analytics bodyType key", "p['body'] → p['bodyType']"),
    ("P1", "Score drivingMix + ev answers", "Collected but unused"),
    ("P1", "Best Family badge bug", "needSeating(familySize as string) but value is string[]"),
    ("P1", "placehold.co → assets/cars/placeholder.svg", "Advisor results/compare images"),
    ("P1", "MPV vs MUV API mismatch", "Next sends body=mpv; DB enum is muv → 0 results"),
    ("P1", "Persist quiz progress", "localStorage/session so nav away does not wipe"),
    ("P1", "Wire Angular to API or shared scorer", "Local Supabase cars vs listings API diverge"),
    ("P2", "SEO 10 vs 11 questions", "Match copy to totalSteps()"),
    ("P2", "Dynamic 'Evaluating N vehicles'", "Stop hardcoding 58"),
    ("P2", "Add CNG to Next fuel options", "Angular + API support CNG"),
    ("P2", "City-aware recommendations", "Use navbar CityService"),
    ("P2", "Playwright e2e for /ai-advisor", "Full 11-step + results + compare"),
]
for col, h in enumerate(["Priority", "Gap", "Notes"], 1):
    ws4.cell(1, col, h).font = Font(bold=True)
for i, (p, g, n) in enumerate(gaps, 2):
    ws4.cell(i, 1, p)
    ws4.cell(i, 2, g)
    ws4.cell(i, 3, n)
ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 48
ws4.column_dimensions["C"].width = 70

# ── Claude prompt ────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Claude Prompt")
prompt = (QA / "AIAdvisor-E2E-Gaps.md").read_text() if (QA / "AIAdvisor-E2E-Gaps.md").exists() else ""
ws5["A1"] = "See AIAdvisor-E2E-Gaps.md section 'Claude fix prompt'"
ws5["A2"] = prompt[:32000] if prompt else "(markdown written alongside this workbook)"
ws5["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws5.column_dimensions["A"].width = 120
ws5.row_dimensions[2].height = 400

out1 = QA / "GAADIIQ_AIAdvisor_Test_Results.xlsx"
out2 = ART / "GAADIIQ_AIAdvisor_Test_Results.xlsx"
wb.save(out1)
wb.save(out2)
print("Wrote", out1)
