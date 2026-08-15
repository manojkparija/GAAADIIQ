#!/usr/bin/env python3
"""Generate AI Diagnosis BRD, E2E test matrix, gaps, Claude prompts."""
from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

QA = Path(__file__).resolve().parent
ART = Path("/opt/cursor/artifacts/ai-diagnosis-brd-qa")
QA.mkdir(parents=True, exist_ok=True)
ART.mkdir(parents=True, exist_ok=True)

# id, area, priority, requirement, status, evidence, gap, justification
REQS: list[tuple[str, str, str, str, str, str, str, str]] = [
    ("BR-VI-01", "Vehicle Info", "Must Have", "Manual vehicle information entry", "PASS", "Angular Step 1 form (make/model/year/fuel/transmission)", "—", "Baseline path for users who prefer typing"),
    ("BR-VI-02", "Vehicle Info", "Must Have", "Voice-based vehicle information entry", "PASS", "voice-mode overlay + Web Speech STT", "—", "Hands-free capture for drivers / literacy barriers"),
    ("BR-VI-03", "Vehicle Info", "Must Have", "Auto-extract vehicle details from speech", "PASS", "vehicle-info-extractor + POST /diagnosis/voice/extract", "—", "Reduce form friction after speech"),
    ("BR-VI-04", "Vehicle Info", "Must Have", "Auto-fill of vehicle information", "PASS", "onVoiceCompleted merges into form signals", "—", "Keep single source of truth for review step"),
    ("BR-VI-05", "Vehicle Info", "Must Have", "Intelligent validation", "PASS", "step1Valid + Pydantic DiagnoseRequest", "—", "Prevent invalid analyse payloads"),
    ("BR-VI-06", "Vehicle Info", "Must Have", "Missing information prompts", "PASS", "voice-mode _askMissingFields TTS loop", "—", "Complete mandatory vehicle fields before analyse"),
    ("BR-IR-01", "Issue Reporting", "Must Have", "Report issue using text", "PASS", "Step 2 textarea problem_description", "—", "Primary symptom capture"),
    ("BR-IR-02", "Issue Reporting", "Must Have", "Report issue using voice", "PASS", "Step 2 mic toolbar + overlay problem phase", "—", "Speak symptoms in native language"),
    ("BR-IR-03", "Issue Reporting", "Must Have", "Upload vehicle images", "PASS", "POST /upload/image → image_urls", "—", "Visual context for vision + mechanic handoff"),
    ("BR-IR-04", "Issue Reporting", "Must Have", "Upload audio recordings", "FAIL", "DB/schema audio_url exists; unused", "No POST /upload/audio or Step 2 UI", "Engine noise / rattles improve diagnosis"),
    ("BR-IR-05", "Issue Reporting", "Must Have", "Upload videos", "FAIL", "DB/schema video_url exists; unused", "No video upload API/UI", "Capture intermittent symptoms"),
    ("BR-IR-06", "Issue Reporting", "Must Have", "Select dashboard warning lights", "PASS", "WARNING_LIGHTS chips in Step 2", "—", "Structured symptom signals"),
    ("BR-IR-07", "Issue Reporting", "Should Have", "Add maintenance history", "FAIL", "Not in form/API/model", "Add maintenance_history JSON field", "Prior service context"),
    ("BR-ML-01", "Multilingual", "Must Have", "Automatic language detection", "PARTIAL", "detectLanguageFromText exists client-side", "Wire Auto-detect toggle into voice session", "Users may not know BCP-47 codes"),
    ("BR-ML-02", "Multilingual", "Must Have", "Support major Indian languages", "PASS", "11 VOICE_LANGUAGES in picker", "PROMPTS fully localized for 6 only", "India language coverage"),
    ("BR-ML-03", "Multilingual", "Must Have", "Preserve user language in conversation", "PASS", "gq_voice_lang + detected_language", "—", "Consistent journey language"),
    ("BR-ML-04", "Multilingual", "Must Have", "AI responses match user language", "PARTIAL", "_translate_diagnosis via Ollama", "Golden-set QA; English fallback banner", "Trust for non-English users"),
    ("BR-VA-01", "Voice AI", "Must Have", "Speech-to-Text", "PARTIAL", "Browser Web Speech API only", "Server STT + MediaRecorder for WebView", "Production Android/iOS reliability"),
    ("BR-VA-02", "Voice AI", "Must Have", "Text-to-Speech", "PARTIAL", "speechSynthesis only", "Optional server TTS; keep client fallback", "Playback on devices with weak voices"),
    ("BR-VA-03", "Voice AI", "Must Have", "Live transcription", "PASS", "interimText in voice overlay", "—", "User confidence while speaking"),
    ("BR-VA-04", "Voice AI", "Should Have", "Voice playback", "PASS", "Auto-speak report + TTS bar", "—", "Listen to diagnosis"),
    ("BR-VA-05", "Voice AI", "Should Have", "Replay", "PASS", "replaySpeaking()", "—", "Re-hear instructions"),
    ("BR-VA-06", "Voice AI", "Should Have", "Pause", "PASS", "pauseSpeaking()", "—", "Interruptible playback"),
    ("BR-VA-07", "Voice AI", "Should Have", "Stop", "PASS", "stop STT/TTS controls", "—", "Immediate silence"),
    ("BR-VA-08", "Voice AI", "Should Have", "Hands-free mode", "PARTIAL", "Auto-listen after TTS", "Always-on driving mode", "Parked/driving safety"),
    ("BR-VA-09", "Voice AI", "Should Have", "Driver-friendly interaction", "PARTIAL", "Large overlay UI", "Driving mode + RECORD_AUDIO", "Large targets, reduced glances"),
    ("BR-AI-01", "AI Diagnosis", "Must Have", "Ollama LLM integration", "PASS", "services/diagnosis.py run_diagnosis", "—", "Primary LLM path"),
    ("BR-AI-02", "AI Diagnosis", "Must Have", "RAG knowledge base", "PARTIAL", "Keyword match on repair_knowledge.json (12 cases)", "Vector RAG over KB (Qdrant/BGE unused)", "Grounded repair advice"),
    ("BR-AI-03", "AI Diagnosis", "Must Have", "Root cause prediction", "PASS", "possible_causes[]", "—", "Core diagnosis value"),
    ("BR-AI-04", "AI Diagnosis", "Must Have", "Severity prediction", "PASS", "risk_level", "—", "Urgency signalling"),
    ("BR-AI-05", "AI Diagnosis", "Must Have", "Safe-to-drive recommendation", "PASS", "safe_to_drive + UI warning", "—", "Safety-critical output"),
    ("BR-AI-06", "AI Diagnosis", "Must Have", "Estimated repair cost", "PASS", "cost_min_inr / cost_max_inr", "—", "Budget transparency"),
    ("BR-AI-07", "AI Diagnosis", "Must Have", "Estimated repair duration", "PASS", "repair_time_estimate", "—", "Time planning"),
    ("BR-AI-08", "AI Diagnosis", "Must Have", "Recommended next steps", "PASS", "recommended_steps + How to Fix", "—", "Actionable guidance"),
    ("BR-AI-09", "AI Diagnosis", "Should Have", "Preventive maintenance suggestions", "PASS", "preventive_maintenance", "—", "Upsell prevention"),
    ("BR-AI-10", "AI Diagnosis", "Should Have", "Follow-up questions when required", "PARTIAL", "Missing-field asks only", "Post-diagnosis follow_up_questions[]", "Improve low-confidence cases"),
    ("BR-UX-01", "UX", "Must Have", "Mobile-first design", "PARTIAL", "Responsive SCSS + layout e2e", "Dedicated mobile shell polish", "Primary usage is phone"),
    ("BR-UX-02", "UX", "Must Have", "Progress indicators", "PASS", "Steps 1–3 wizard progress", "—", "Orient user in flow"),
    ("BR-UX-03", "UX", "Must Have", "Conversation history", "PARTIAL", "Past Diagnoses panel + GET /history", "No multi-turn conversation DB/UI; open-by-id weak", "Revisit prior advice"),
    ("BR-UX-04", "UX", "Must Have", "Voice/Text switching", "PASS", "Mode selector Step 1", "—", "User preference"),
    ("BR-UX-05", "UX", "Must Have", "Error handling", "PASS", "STT + diagnosis error cards", "—", "Recoverable failures"),
    ("BR-UX-06", "UX", "Should Have", "Offline handling", "PARTIAL", "clientFallback when API down", "Offline banner; STT needs network", "Low-connectivity India"),
    ("BR-SEC-01", "Security", "Must Have", "Microphone permissions", "PARTIAL", "Browser not-allowed string only", "Android RECORD_AUDIO + Capacitor helper", "OS-level mic access"),
    ("BR-SEC-02", "Security", "Must Have", "Voice data encryption", "PARTIAL", "TLS in transit assumed", "At-rest encryption design for recordings", "Protect voice PII"),
    ("BR-SEC-03", "Security", "Must Have", "Secure storage", "PARTIAL", "Postgres + R2 for images", "Transcripts not stored securely", "Controlled retention"),
    ("BR-SEC-04", "Security", "Must Have", "User consent for voice processing", "FAIL", "Generic privacy pages only", "Pre-STT consent modal (purpose/retention)", "DPDP / lawful processing"),
    ("BR-SEC-05", "Security", "Must Have", "Delete recordings", "FAIL", "Nothing voice-persisted yet", "Retention + DELETE voice APIs", "User control of biometric-adjacent data"),
    ("BR-SEC-06", "Security", "Must Have", "Privacy compliance (DPDP)", "PARTIAL", "Disclaimer copy", "Diagnosis/voice delete UX", "India DPDP Act readiness"),
    ("BR-PERF-01", "Performance", "Must Have", "Fast response time", "PARTIAL", "Heuristic fallback fast", "Ollama timeout ≤8s → fallback", "Perceived performance"),
    ("BR-PERF-02", "Performance", "Should Have", "High STT accuracy", "PARTIAL", "Browser STT + noisy retry", "WER benchmarks per language", "Voice trust"),
    ("BR-PERF-03", "Performance", "Must Have", "High AI response accuracy", "PARTIAL", "KB+Ollama", "Golden-set fixtures in CI", "Safety + usefulness"),
    ("BR-API-01", "API", "Must Have", "STT API", "FAIL", "Browser-only", "POST /diagnosis/stt (multipart)", "WebView production path"),
    ("BR-API-02", "API", "Must Have", "TTS API", "FAIL", "Browser-only", "POST /diagnosis/tts optional", "Consistent voice quality"),
    ("BR-API-03", "API", "Should Have", "Language detection API", "FAIL", "Client-only heuristic", "POST /diagnosis/detect-language", "Server-side lang detect"),
    ("BR-API-04", "API", "Should Have", "Translation API", "PARTIAL", "Internal _translate_diagnosis", "Not standalone endpoint", "Reuse across modules"),
    ("BR-API-05", "API", "Must Have", "AI Diagnosis API", "PASS", "POST /diagnosis/analyse", "—", "Core backend"),
    ("BR-API-06", "API", "Must Have", "Image upload API", "PASS", "POST /upload/image", "—", "Media pipeline"),
    ("BR-API-07", "API", "Must Have", "Audio upload API", "FAIL", "Missing", "POST /upload/audio", "Persist engine audio"),
    ("BR-DB-01", "Database", "Must Have", "Voice transcript storage", "FAIL", "Missing", "voice_transcripts (consent-gated)", "Audit + improve STT"),
    ("BR-DB-02", "Database", "Must Have", "Conversation history storage", "FAIL", "In-session memory only", "diagnosis_conversations", "Multi-turn persistence"),
    ("BR-DB-03", "Database", "Must Have", "AI diagnosis history", "PASS", "vehicle_diagnoses + Past Diagnoses UI", "—", "User revisit"),
    ("BR-DB-04", "Database", "Should Have", "Audit logs", "FAIL", "Missing", "diagnosis_audit_events", "Compliance + debug"),
    ("BR-TEST-01", "Quality", "Must Have", "Automated diagnosis tests", "FAIL", "No test_diagnosis.py / extractor specs", "Add pytest + unit tests", "Prevent regressions"),
    ("BR-AND-01", "Mobile", "Must Have", "Android RECORD_AUDIO permission", "FAIL", "Manifest has CAMERA/location only", "Add RECORD_AUDIO + runtime request", "APK mic required"),
]


def ac_block(rid: str, title: str, status: str, given: str, when: str, then: str, pos: str, neg: str, edge: str) -> str:
    return f"""### AC-{rid} — {title} — **{status}**
| Field | Detail |
|-------|--------|
| Requirement ID | {rid} |
| User Story | As a user, I want {title.lower()} so the diagnosis journey meets BRD. |
| Preconditions | App on `/vehicle-diagnosis`; tip `claude/gaadiiq-app-dev-abj5fo` (or merged equivalent). |
| Given | {given} |
| When | {when} |
| Then | {then} |
| Positive | {pos} |
| Negative | {neg} |
| Edge | {edge} |
| Expected Result | Status **{status}** against current tip. |
"""


def main() -> None:
    status_c = Counter(r[4] for r in REQS)
    weights = {"PASS": 1.0, "PARTIAL": 0.5, "FAIL": 0.0}
    score = round(100 * sum(weights[r[4]] for r in REQS) / len(REQS))

    tests: list[dict] = []

    def T(suite: str, tid: str, name: str, status: str, detail: str, sev: str = "—") -> None:
        tests.append(
            {"suite": suite, "id": tid, "name": name, "status": status, "detail": detail, "severity": sev}
        )

    # Functional
    T("Functional", "TC-F-01", "Manual Step1→4 diagnose path", "PASS", "Wizard + analyse API")
    T("Functional", "TC-F-02", "Voice language picker → capture vehicle", "PASS", "voice-mode flow")
    T("Functional", "TC-F-03", "Extract Hindi/Bengali/Tamil brands", "PASS", "multilingual extractor")
    T("Functional", "TC-F-04", "Missing field voice re-ask", "PASS", "_askMissingFields")
    T("Functional", "TC-F-05", "Warning lights + severity", "PASS", "Step 2 chips")
    T("Functional", "TC-F-06", "Image upload then analyse", "PASS", "upload + image_urls")
    T("Functional", "TC-F-07", "Report cost/safe-drive/steps/disclaimer", "PASS", "Report UI")
    T("Functional", "TC-F-08", "TTS play/pause/stop/replay", "PASS", "voice-diagnosis.service")
    T("Functional", "TC-F-09", "Client KB fallback when API down", "PASS", "clientFallback")
    T("Functional", "TC-F-10", "Auth history endpoint", "PASS", "GET /diagnosis/history")
    T("Functional", "TC-F-11", "IDOR blocked on GET diagnosis", "PASS", "MOB-007 owner check")
    T("Functional", "TC-F-12", "Upload audio recording", "FAIL", "No UI/API", "P0")
    T("Functional", "TC-F-13", "Upload video", "FAIL", "No UI/API", "P0")
    T("Functional", "TC-F-14", "Maintenance history entry", "FAIL", "Not implemented", "P1")
    T("Functional", "TC-F-15", "Auto language detect without picker", "FAIL", "autoDetect unwired", "P1")
    T("Functional", "TC-F-16", "Full prompt localization kn/ml/gu/pa/or", "FAIL", "PROMPTS only 6 langs", "P1")
    T("Functional", "TC-F-17", "Angular Past Diagnoses history UI", "PASS", "history-panel on Claude tip")
    T("Functional", "TC-F-18", "Server STT for WebView/Safari", "FAIL", "Browser STT only", "P0")
    T("Functional", "TC-F-19", "Mic consent before STT", "FAIL", "Missing", "P0")
    T("Functional", "TC-F-20", "Android RECORD_AUDIO declared", "FAIL", "Manifest gap", "P0")
    T("Functional", "TC-F-21", "Persist conversation transcripts", "FAIL", "No DB", "P1")
    T("Functional", "TC-F-22", "Delete voice data (DPDP)", "FAIL", "Missing", "P1")
    T("Functional", "TC-F-23", "Post-diagnosis follow-up questions", "FAIL", "Not in analyse", "P2")
    T("Functional", "TC-F-24", "Vector RAG for diagnosis KB", "FAIL", "Keyword only", "P2")
    T("Functional", "TC-F-25", "Open past diagnosis by id from history", "PARTIAL", "List only; no detail navigation", "P2")
    T("Functional", "TC-F-26", "Guest can analyse without login", "PASS", "Public analyse + rate limit")
    T("Functional", "TC-F-27", "detected_language sent on analyse", "PASS", "Claude Angular + DiagnoseRequest")
    # Security
    T("Security", "TC-S-01", "Unauth GET diagnosis denied", "PASS", "owner auth")
    T("Security", "TC-S-02", "Prompt injection fencing", "PARTIAL", "User text in prompts", "P0")
    T("Security", "TC-S-03", "Rate limit analyse", "PASS", "documented limits")
    T("Security", "TC-S-04", "Audio upload validation", "FAIL", "No audio upload", "P1")
    T("Security", "TC-S-05", "Mic consent logged before STT", "FAIL", "No consent event", "P0")
    # Performance
    T("Performance", "TC-P-01", "Analyse p95 with fallback", "PARTIAL", "Ollama unbounded", "P1")
    T("Performance", "TC-P-02", "STT works offline", "FAIL", "Web Speech needs net", "P1")
    T("Performance", "TC-P-03", "Image upload ≤10MB reject", "PASS", "upload validation path")
    # A11y / Mobile / Regression / Integration / Usability
    T("A11y", "TC-A-01", "Voice overlay SR labels", "PARTIAL", "Audit incomplete", "P2")
    T("A11y", "TC-A-02", "aria-live for interim transcript", "FAIL", "Missing", "P2")
    T("Mobile", "TC-M-01", "No horizontal overflow on route", "PASS", "mobile-layout e2e")
    T("Mobile", "TC-M-02", "Voice overlay one-handed", "PARTIAL", "No driving mode", "P2")
    T("Regression", "TC-R-01", "API pytest diagnosis suite", "FAIL", "No test_diagnosis.py", "P0")
    T("Regression", "TC-R-02", "Voice extractor unit tests", "FAIL", "No *.spec.ts", "P0")
    T("Integration", "TC-I-01", "Ollama down → heuristic report", "PASS", "fallback path")
    T("Integration", "TC-I-02", "Vision on image_url", "PARTIAL", "vision.py untested quality", "P2")
    T("Integration", "TC-I-03", "voice/extract then autofill", "PASS", "Claude tip path")
    T("Usability", "TC-U-01", "Manual vs Voice mode clear", "PASS", "Step 1 selector")
    T("Usability", "TC-U-02", "Disclaimer before submit", "PASS", "Step 3")
    T("Usability", "TC-U-03", "safe_to_drive=false banner prominence", "PASS", "Report warning UI")

    tc = Counter(t["status"] for t in tests)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    iso = datetime.now(timezone.utc).isoformat()

    req_rows = "\n".join(
        f"| {a} | {b} | {c} | {d} | **{e}** | {g} |" for a, b, c, d, e, _f, g, _j in REQS
    )
    fail_tests = "\n".join(
        f"| {t['id']} | {t['status']} | {t['severity']} | {t['name']} — {t['detail']} |"
        for t in tests
        if t["status"] != "PASS"
    )
    just_rows = "\n".join(
        f"| {a} | {c} | {j} |" for a, _b, c, _d, _e, _f, _g, j in REQS
    )

    # Full AC for every requirement (compact table rows + detailed for critical)
    ac_sections = []
    ac_map = {
        "BR-VI-01": ("Manual vehicle entry", "User on Step 1 Manual mode", "enters make/model/year/fuel", "fields validate and Next enables", "Valid Maruti Swift 2019", "Empty make blocked", "Unknown variant allowed as optional"),
        "BR-VI-02": ("Voice vehicle entry", "Voice mode + language selected", "user speaks vehicle details", "STT captures transcript and advances extract", "Hindi Maruti speech", "Mic denied shows error", "No-speech retry"),
        "BR-VI-03": ("Auto-extract from speech", "Transcript available", "client/server extract runs", "manufacturer/model/year/fuel populated when present", "en/hi/bn/ta brands", "Gibberish → missing asks", "Mixed Hindi-English"),
        "BR-VI-04": ("Auto-fill form", "Extract result emitted", "onVoiceCompleted runs", "form signals match extract", "Partial extract fills only known", "Overwrite confirmed fields carefully", "User edits after fill"),
        "BR-VI-05": ("Intelligent validation", "Form incomplete", "user taps Next/Analyse", "validation errors shown; API rejects bad payloads", "All required OK", "Year out of range", "Odometer optional"),
        "BR-VI-06": ("Missing info prompts", "Extract missing year", "voice asks for year", "user reply fills year", "One missing field", "Repeated no-speech", "All fields missing"),
        "BR-IR-01": ("Text issue report", "Step 2", "types ≥10 char problem", "accepted for analyse", "Engine knock description", "Too short blocked", "Emoji-only"),
        "BR-IR-02": ("Voice issue report", "Step 2 mic or overlay", "speaks problem", "transcript fills textarea", "Tamil symptom speech", "Mic denied", "Noisy environment retry"),
        "BR-IR-03": ("Image upload", "Logged-in or upload path", "selects photo", "image_urls on analyse", "jpeg/png", "Corrupt file reject", "Max 5 images"),
        "BR-IR-04": ("Audio upload", "Step 2", "uploads/records audio", "audio_url persisted", "N/A — not built", "N/A", "≤60s when built"),
        "BR-IR-05": ("Video upload", "Step 2", "uploads video", "video_url persisted", "N/A — not built", "N/A", "Size limit TBD"),
        "BR-IR-06": ("Warning lights", "Step 2", "toggles light chips", "array sent to API", "Check engine", "None selected OK", "Multiple lights"),
        "BR-IR-07": ("Maintenance history", "Step 2", "enters last service", "sent/stored", "N/A — not built", "N/A", "Free text vs structured"),
        "BR-ML-01": ("Auto language detect", "Auto-detect on", "user speaks Hindi", "session language becomes hi-IN", "Currently PARTIAL/unwired", "English mixed", "Code-switch"),
        "BR-ML-02": ("Indian languages", "Language picker", "selects kn/ml/gu/pa/or", "STT locale set", "11 langs listed", "Unsupported locale", "Device missing voice"),
        "BR-ML-03": ("Preserve language", "hi-IN session", "completes journey", "TTS/prompts stay Hindi", "Full voice path", "User switches mid-way", "Reload restores lang"),
        "BR-ML-04": ("AI language match", "detected_language=ta-IN", "analyse returns", "report Tamil or EN banner", "Translation path exists", "Ollama fails translate", "Mixed scripts"),
        "BR-VA-01": ("STT", "Mic allowed", "speaks", "transcript appears", "Chrome Web Speech", "Android WebView often fails", "Server STT missing"),
        "BR-VA-02": ("TTS", "Report ready", "auto-play", "speech heard; pause/stop work", "Chrome voices", "Silent device", "No server TTS"),
        "BR-VA-03": ("Live transcription", "Listening", "user speaks", "interimText updates", "Continuous speech", "Empty interim", "Very long utterance"),
        "BR-VA-04": ("Voice playback", "Report", "TTS plays", "audio output", "Auto-play on", "Muted OS", "Long report"),
        "BR-VA-05": ("Replay", "TTS finished", "Replay tapped", "replays report", "Works", "Nothing to replay", "During pause"),
        "BR-VA-06": ("Pause", "TTS playing", "Pause", "speech pauses", "Works", "Already paused", "OS interrupt"),
        "BR-VA-07": ("Stop", "STT or TTS active", "Stop", "recognition/speech stops", "Works", "Idle stop noop", "Rapid toggle"),
        "BR-VA-08": ("Hands-free", "Voice session", "after TTS prompt", "listening resumes", "Partial auto-listen", "No driving mode", "Bluetooth headset"),
        "BR-VA-09": ("Driver-friendly", "Phone in car", "uses voice", "large targets / minimal typing", "Overlay large", "No driving mode", "RECORD_AUDIO gap"),
        "BR-AI-01": ("Ollama", "API up + Ollama", "analyse", "LLM fields populated", "Happy path", "Ollama down → fallback", "Timeout"),
        "BR-AI-02": ("RAG KB", "Known symptom in KB", "analyse", "KB case influences output", "Keyword hit", "Miss → generic", "Vector RAG missing"),
        "BR-AI-03": ("Root cause", "Valid payload", "analyse", "possible_causes non-empty", "Knocking", "Vague symptom", "Multi-cause"),
        "BR-AI-04": ("Severity", "Valid payload", "analyse", "risk_level set", "Critical overheat", "Cosmetic", "Unknown"),
        "BR-AI-05": ("Safe-to-drive", "Brake failure symptom", "analyse", "safe_to_drive false + warning", "Critical", "Minor rattle true", "Ambiguous"),
        "BR-AI-06": ("Repair cost", "Valid", "analyse", "cost_min/max INR shown", "Range present", "Zero range", "Currency INR"),
        "BR-AI-07": ("Repair duration", "Valid", "analyse", "repair_time_estimate shown", "Hours/days string", "Empty", "Complex jobs"),
        "BR-AI-08": ("Next steps", "Valid", "analyse", "recommended_steps listed", "Visit workshop", "Empty list", "DIY section"),
        "BR-AI-09": ("Preventive", "Valid", "analyse", "preventive_maintenance present", "Oil service tip", "Empty OK", "Fleet tips"),
        "BR-AI-10": ("Follow-ups", "Low confidence", "analyse", "follow_up_questions returned", "Not implemented", "High confidence skip", "Voice ask loop"),
        "BR-UX-01": ("Mobile-first", "390px viewport", "open diagnosis", "usable without H-scroll", "Layout e2e", "Desktop ok", "Landscape"),
        "BR-UX-02": ("Progress", "Wizard", "advance steps", "step indicator updates", "1→2→3", "Jump invalid", "Back nav"),
        "BR-UX-03": ("Conversation history", "Logged-in user", "opens Past Diagnoses", "list shows prior reports", "List PASS; conversation DB FAIL", "Guest empty/login", "Open-by-id PARTIAL"),
        "BR-UX-04": ("Voice/Text switch", "Step 1", "toggle Manual/Voice", "correct path starts", "Switch before start", "Mid-flow switch", "Unsupported STT hides voice"),
        "BR-UX-05": ("Error handling", "API 500", "analyse", "error card + retry/fallback", "Network fail", "Silent fail absent", "Partial response"),
        "BR-UX-06": ("Offline", "Airplane mode", "open page", "banner/queue guidance", "API fallback only", "STT offline fail", "Resume when online"),
        "BR-SEC-01": ("Mic permissions", "First mic use", "request permission", "OS prompt; denied guidance", "Browser path", "APK missing RECORD_AUDIO", "iOS WKWebView"),
        "BR-SEC-02": ("Voice encryption", "Recording stored", "at rest", "encrypted", "TLS only today", "Plain R2 audio risk", "Key rotation"),
        "BR-SEC-03": ("Secure storage", "Diagnosis saved", "DB write", "authz on read", "Owner only", "IDOR blocked", "Transcripts missing"),
        "BR-SEC-04": ("Mic consent", "First Use Voice", "tap Voice", "consent modal before STT", "FAIL missing", "Decline blocks STT", "Re-consent after revoke"),
        "BR-SEC-05": ("Delete recordings", "User has voice data", "Delete", "data removed", "FAIL missing", "Already deleted", "Cascade audit"),
        "BR-SEC-06": ("DPDP", "Privacy settings", "delete diagnosis", "erasure path", "PARTIAL", "Guest ephemeral", "Export request future"),
        "BR-PERF-01": ("Fast response", "Normal load", "analyse", "p95≤5s or fallback≤1s", "Fallback fast", "Ollama hang", "Concurrent users"),
        "BR-PERF-02": ("STT accuracy", "Clear speech hi-IN", "STT", "WER within target", "No benchmark", "Noisy street", "Accent variants"),
        "BR-PERF-03": ("AI accuracy", "Golden fixtures", "analyse", "expected causes match", "No CI set", "Adversarial text", "Vision assist"),
        "BR-API-01": ("STT API", "Multipart audio", "POST /diagnosis/stt", "transcript JSON", "FAIL missing", "Bad audio", "Auth/rate limit"),
        "BR-API-02": ("TTS API", "Text body", "POST /diagnosis/tts", "audio bytes/URL", "FAIL missing", "Empty text", "Lang param"),
        "BR-API-03": ("Lang detect API", "Text sample", "POST detect-language", "BCP-47 code", "FAIL missing", "Empty", "Code-switch"),
        "BR-API-04": ("Translation API", "EN report + ta-IN", "translate", "Tamil fields", "Internal only", "Fail → EN banner", "Standalone missing"),
        "BR-API-05": ("Diagnosis API", "Valid DiagnoseRequest", "POST /analyse", "201 + report", "Happy", "Validation 422", "Rate limit 429"),
        "BR-API-06": ("Image upload API", "Auth + image", "POST /upload/image", "URL returned", "jpeg", "Unauth 401", "SSRF on vision fetch"),
        "BR-API-07": ("Audio upload API", "Auth + audio", "POST /upload/audio", "URL returned", "FAIL missing", "Wrong MIME", "Max duration"),
        "BR-DB-01": ("Transcript storage", "Consent yes", "voice session ends", "row in voice_transcripts", "FAIL", "Consent no → no store", "Retention TTL"),
        "BR-DB-02": ("Conversation storage", "Multi-turn voice", "session completes", "turns persisted", "FAIL", "Cancel mid-way", "PII redaction"),
        "BR-DB-03": ("Diagnosis history", "Auth user", "after analyse", "row + history list", "PASS", "Guest no history", "Pagination"),
        "BR-DB-04": ("Audit logs", "Analyse/delete/consent", "event", "audit row", "FAIL", "PII minimization", "Immutable log"),
        "BR-TEST-01": ("Automated tests", "CI", "pytest/unit", "diagnosis suite green", "FAIL missing", "Flaky Ollama mocked", "E2E smoke"),
        "BR-AND-01": ("Android RECORD_AUDIO", "APK install", "Use Voice", "permission prompt works", "FAIL manifest", "Deny forever", "iOS mic usage string"),
    }
    for rid, (title, given, when, then, pos, neg, edge) in ac_map.items():
        st = next(r[4] for r in REQS if r[0] == rid)
        ac_sections.append(ac_block(rid, title, st, given, when, then, pos, neg, edge))
    ac_md = "\n".join(ac_sections)

    brd = f"""# GAADIIQ — AI Diagnosis Module
## Business Requirements Document (BRD) + E2E QA Results

| Field | Value |
|-------|-------|
| Product | GAADIIQ AI Vehicle Preliminary Diagnosis |
| Version | 1.1 (implementation-ready) |
| Date | {now} |
| Code tip audited | `claude/gaadiiq-app-dev-abj5fo` (voice multilingual: `733d258`…`14df98f`) |
| Docs branch | `cursor/ai-diagnosis-brd-qa-85e1` (docs-only; merge Claude tip before implementing voice ACs) |
| Status | **Partial vs full BRD — NOT production-ready for complete voice+media scope** |
| Audience | BA · Product · Engineering · UI/UX · QA |

---

## 1. Business Overview

### 1.1 Business problem
Indian vehicle owners delay repairs due to language barriers, opaque workshop advice, and fear of overcharging. They need fast, trustworthy, **multilingual** preliminary diagnosis before visiting a garage — including while hands are busy or literacy/English is limited.

### 1.2 Business objectives
1. Diagnose via **form or voice** in major Indian languages.
2. Return AI preliminary root cause, severity, safe-to-drive, cost/time, next steps, and preventive tips.
3. Support images (and audio/video per BRD) for richer context.
4. Be mobile-first, private (DPDP-aware), and production-reliable on Android/iOS WebView.
5. Reuse existing Ollama + repair KB engine with measurable quality and fallbacks.

### 1.3 Scope (in) — v1
- Manual 4-step diagnosis wizard
- Voice-first overlay (STT/TTS, language picker, extract, missing-field asks)
- Multilingual analyse responses (`detected_language`)
- Image upload + optional vision assist
- Diagnosis persistence + Past Diagnoses history (auth)
- Safety disclaimer + service-centre CTA
- Rate limits + IDOR protection on history detail

### 1.4 Out of scope (v1)
OBD-II / live sensors · certified medical-grade accuracy · full insurance claim filing · replacing licensed mechanic inspection · dealer DMS deep sync · always-on background listening without consent.

### 1.5 Assumptions
1. Users grant mic/camera when prompted (after consent UX ships).
2. Ollama may be unavailable; heuristic/KB fallback must still return a report.
3. Browser Web Speech is acceptable for **beta**; **GA requires server STT**.
4. Preliminary diagnosis always shows non-certified disclaimer.
5. Implementers work from Claude tip (or merge it) — master alone lacks voice stack.

### 1.6 Dependencies
| Dependency | Purpose |
|------------|---------|
| FastAPI `apps/api` diagnosis + upload routers | Analyse / history / media |
| Postgres `vehicle_diagnoses` | Persistence |
| Cloudflare R2 (or configured store) | Image (and future audio/video) |
| Ollama + `repair_knowledge.json` | LLM + keyword RAG |
| Capacitor Android/iOS | Mic/camera permissions |
| Web Speech / future Whisper|Google|Azure | STT |
| `speechSynthesis` / future server TTS | Playback |

---

## 2. User Personas

| Persona | Goals | Pain | Diagnosis needs |
|---------|-------|------|-----------------|
| Individual Car Owner | Understand issue before workshop | Overcharging fear | Clear cost + safe-to-drive |
| First-time Driver | Know if safe to continue | Panic / jargon | Simple language + TTS |
| Non-English Speaking User | Use Hindi/Tamil/etc. | English-only apps | Full journey in native language |
| Fleet Owner | Triage many vehicles | Downtime cost | History, severity, speed |
| Taxi Driver | Hands-busy, roadside | Can't type safely | Voice-first, driver-friendly |
| Dealer | Pre-inspect trade-ins | Inconsistent intake | Structured report + media |
| Service Advisor | Faster bay intake | Incomplete symptoms | Warning lights + photos + history |

---

## 3. Business Requirements — coverage vs code

**{status_c.get('PASS', 0)} PASS / {status_c.get('PARTIAL', 0)} PARTIAL / {status_c.get('FAIL', 0)} FAIL** ({len(REQS)} requirements)  
**Weighted BRD readiness: {score}/100**

| ID | Area | Priority | Requirement | Status | Gap |
|----|------|----------|-------------|--------|-----|
{req_rows}

### 3.1 Business justification (all IDs)

| ID | Priority | Justification |
|----|----------|---------------|
{just_rows}

---

## 4. Non-Functional Requirements (measurable)

| ID | Category | Target | Current | Status |
|----|----------|--------|---------|--------|
| NFR-P1 | Performance | Analyse p95 ≤ 5s; fallback ≤ 1s | Fallback OK; Ollama unbounded | PARTIAL |
| NFR-P2 | STT | WER ≤ 20% on clear hi-IN/en-IN lab set | No benchmark | FAIL |
| NFR-A1 | Availability | 99.5% report success via fallback | Heuristic fallback exists | PARTIAL |
| NFR-S1 | Scalability | 50 concurrent analyse with queue/timeout | Untested | FAIL |
| NFR-R1 | Reliability | Zero silent failures; always error or report | Error cards + fallback | PASS |
| NFR-SEC1 | Security | Consent before mic; authz on history | Authz OK; consent missing | PARTIAL |
| NFR-SEC2 | Privacy | DPDP delete within 72h of request | Delete UX missing | FAIL |
| NFR-A11Y | Accessibility | WCAG 2.1 AA on critical path | Partial labels | PARTIAL |
| NFR-M1 | Maintainability | Diagnosis pytest + extractor unit tests | Almost none | FAIL |
| NFR-O1 | Monitoring | Structured logs + latency metrics for analyse/STT | Partial API logs | PARTIAL |
| NFR-L1 | Logging | Audit consent, analyse, delete events | Missing audit table | FAIL |

---

## 5. UI/UX Requirements (screens)

| Screen | Required elements | Present on tip |
|--------|-------------------|----------------|
| Vehicle Details Form | Make/model/year/fuel/transmission; Manual/Voice mode | **Yes** |
| Voice Recording Interface | Lang picker, live transcript, mic states, cancel | **Yes** |
| AI Diagnosis / Review Screen | Confirm fields, disclaimer, submit | **Yes** |
| AI Result Screen | Causes, severity, safe-to-drive, cost, time, steps, DIY, preventive | **Yes** |
| Voice Playback Controls | Play / pause / stop / replay / mute | **Yes** |
| Conversation / Past History | List past diagnoses; open detail; conversation turns | **Partial** (list yes; turns/detail weak) |
| Media Upload | Images + audio + video | **Images only** |
| Maintenance History | Free text / structured prior service | **No** |
| Mic Consent Gate | Purpose, retention, accept/decline | **No** |
| Error Screens | Mic denied, STT fail, API fail, offline | **Partial** |

---

## 6. Process Flows

### 6.1 Manual diagnosis journey
1. Open `/vehicle-diagnosis` → choose Manual.  
2. Enter vehicle details → Next.  
3. Enter symptoms, warning lights, severity; optional photos → Next.  
4. Review + disclaimer → Analyse.  
5. View report (cost, safe-to-drive, steps) → optional TTS / service-centre CTA.  
6. If logged in, appears in Past Diagnoses.

### 6.2 Voice diagnosis journey
1. Choose Use Voice → **(gap)** consent modal → language picker.  
2. TTS greeting → speak vehicle → STT → extract (client + `/voice/extract`).  
3. Missing-field TTS asks until required fields complete.  
4. Speak problem → autofill form → optional lights/photos → review → analyse with `detected_language`.  
5. Server translates report fields when non-English → TTS reads report → pause/replay/stop.

### 6.3 AI processing workflow
Validate → optional vision on first `image_url` → keyword KB retrieve → Ollama (or heuristic) → translate → persist `vehicle_diagnoses` → return disclaimer-bound report.

### 6.4 Error handling workflow
| Failure | Expected UX |
|---------|-------------|
| Mic denied | Clear message + settings guidance (**APK RECORD_AUDIO gap**) |
| No-speech / low confidence | Retry with tip |
| Web Speech unsupported | **Should** fall back to MediaRecorder → `/diagnosis/stt` (**missing**) |
| API down | Client KB fallback report |
| Ollama timeout | Heuristic report (**timeout not enforced**) |
| Translation fail | English report + language fallback banner (**banner weak/unproven**) |

```mermaid
flowchart TD
  A[Open /vehicle-diagnosis] --> B{{Manual or Voice}}
  B -->|Manual| C[Vehicle form]
  B -->|Voice| D[Consent - MISSING]
  D --> E[Language + STT]
  E --> F[Extract + missing asks]
  C --> G[Symptoms + media]
  F --> G
  G --> H[Review + disclaimer]
  H --> I[POST /diagnosis/analyse]
  I --> J{{Ollama OK?}}
  J -->|Yes| K[Report + optional translate]
  J -->|No| L[Heuristic/KB fallback]
  K --> M[TTS + History]
  L --> M
```

---

## 7. API Requirements

| API | Method | Status | Notes |
|-----|--------|--------|-------|
| `/diagnosis/analyse` | POST | **Exists** | Public + rate limit; fields include `audio_url`/`video_url`/`detected_language` |
| `/diagnosis/voice/extract` | POST | **Exists** (Claude tip) | `{{ "transcript": str }}` |
| `/diagnosis/history` | GET | **Exists** | Auth required |
| `/diagnosis/{{id}}` | GET | **Exists** | Auth + owner check (MOB-007) |
| `/upload/image` | POST | **Exists** | Auth; R2 |
| `/upload/audio` | POST | **Missing** | Magic bytes; max ~60s; returns URL |
| `/upload/video` | POST | **Missing** | Size/duration limits |
| `/diagnosis/stt` | POST | **Missing** | Multipart → transcript + lang |
| `/diagnosis/tts` | POST | **Missing** | Optional server TTS |
| `/diagnosis/detect-language` | POST | **Missing** | Text → BCP-47 |
| `/diagnosis/{{id}}` DELETE | DELETE | **Missing** | DPDP erasure |
| `/diagnosis/voice-data` DELETE | DELETE | **Missing** | Transcripts/recordings |

### 7.1 DiagnoseRequest (current tip)
`manufacturer, model, variant?, model_year, fuel_type, transmission, odometer_km?, problem_description, warning_lights[], when_occurs[], severity, image_urls[]?, audio_url?, video_url?, user_id?, detected_language?`

---

## 8. Database Impact

| Object | Action | Fields / notes |
|--------|--------|----------------|
| `vehicle_diagnoses` | Keep | Wire `audio_url`/`video_url`; add `maintenance_history` JSON; optional `confidence` |
| `diagnosis_conversations` | **Create** | `id, user_id, diagnosis_id?, language, turns JSONB, created_at` |
| `voice_transcripts` | **Create** | `id, user_id, consent_id, transcript, lang, audio_url?, created_at, expires_at` |
| `diagnosis_audit_events` | **Create** | `event_type, actor_id, entity_id, meta JSONB, created_at` |
| Consent store | **Create** | Mic/voice processing consent timestamp + version |

---

## 9. Business Rules

1. Problem text ≥ 10 characters **or** non-empty voice transcript before analyse.  
2. Images: jpeg/png/webp/heic; max 5; auth for upload.  
3. Audio (when built): ≤ 60s; mime allowlist; consent required before capture.  
4. Video (when built): max duration/size TBD; consent if mic track present.  
5. Always show safety disclaimer; never claim certified/OBD diagnosis.  
6. If `safe_to_drive=false`, show prominent stop-driving warning before secondary CTAs.  
7. Response language = user language; on failure show English + explicit banner.  
8. Guest may analyse; history and delete require authentication.  
9. Low confidence → `follow_up_questions` or low-confidence banner (follow-ups not yet returned).  
10. Prompt-injection: fence user free text; never execute instructions from `problem_description`.  
11. Voice data retention default ≤ 30 days unless user deletes earlier (policy TBD).  
12. Android builds must declare `RECORD_AUDIO` before shipping voice to Play.

---

## 10. Acceptance Criteria (all requirements)

{ac_md}

---

## 11. Test Scenarios & Results

| Metric | Value |
|--------|------:|
| Scenarios | {len(tests)} |
| PASS | {tc.get('PASS', 0)} |
| PARTIAL | {tc.get('PARTIAL', 0)} |
| FAIL | {tc.get('FAIL', 0)} |
| Requirements | {status_c.get('PASS', 0)}P / {status_c.get('PARTIAL', 0)}~ / {status_c.get('FAIL', 0)}F |
| BRD readiness | **{score}/100** |
| Verdict | **NO-GO** full BRD; **Conditional Go** manual+image MVP |

### 11.1 Non-PASS scenarios

| ID | Status | Severity | Detail |
|----|--------|----------|--------|
{fail_tests}

### 11.2 Suite coverage intent
Functional · Integration · Regression · Usability · Security · Performance · Accessibility · Mobile.

---

## 12. Risks and Mitigation

| Risk | Impact | Mitigation |
|------|--------|------------|
| WebView STT dead on Android | Voice BRD blocked | Server STT + Capacitor mic + RECORD_AUDIO |
| LLM wrong safe-to-drive | Safety | Conservative prompts + disclaimer + golden tests |
| Almost no automated tests | Regressions | `test_diagnosis.py` + extractor specs (Wave 1) |
| Voice without consent | Legal/DPDP | Consent gate + audit + delete |
| Thin KB (12 cases) | Weak advice | Expand KB + embedding RAG |
| Translation quality | Trust loss | Golden multilingual fixtures + EN fallback banner |
| Docs branch ≠ voice code | Wrong baseline | Implement on Claude tip / merge first |
| Prompt injection | Unsafe output | Fence user text (MOB-008) |

---

## 13. Future Enhancements
OBD-II integration · real-time sensor diagnostics · predictive maintenance · connected vehicle telemetry · service appointment booking · dealer CRM integration · insurance claim assistance · AI maintenance reminders · Bluetooth OBD dongle · WhatsApp voice intake.

---

## 14. Production Readiness Verdict

| Scope | Verdict |
|-------|---------|
| Manual + images + AI report + disclaimer | **Conditional Go** (add tests + prompt fencing) |
| Browser multilingual voice (Chrome) | **Beta** (consent, Android mic, localization gaps) |
| Production Android/iOS voice + audio/video + DPDP | **NO-GO** |
| Full BRD | **NO-GO** — readiness **{score}/100** |

**Primary Claude entrypoint:** `docs/qa/diagnosis/Claude_Fix_Prompts_Diagnosis_BRD.md`
"""

    (QA / "GAADIIQ_AI_Diagnosis_BRD_and_QA.md").write_text(brd)
    (ART / "GAADIIQ_AI_Diagnosis_BRD_and_QA.md").write_text(brd)

    catalog = {
        "generated_at": iso,
        "brd_readiness_score": score,
        "code_tip": "claude/gaadiiq-app-dev-abj5fo",
        "requirement_counts": dict(status_c),
        "test_counts": dict(tc),
        "recommendation": "NO-GO for full BRD; CONDITIONAL for manual+image MVP",
        "requirements": [
            {
                "id": a,
                "area": b,
                "priority": c,
                "requirement": d,
                "status": e,
                "evidence": f,
                "gap": g,
                "justification": j,
            }
            for a, b, c, d, e, f, g, j in REQS
        ],
        "tests": tests,
    }
    (QA / "diagnosis-brd-qa.json").write_text(json.dumps(catalog, indent=2))
    (ART / "diagnosis-brd-qa.json").write_text(json.dumps(catalog, indent=2))

    with (QA / "diagnosis-requirements.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["id", "area", "priority", "requirement", "status", "evidence", "gap", "justification"])
        w.writerows(REQS)

    prompts = f"""# Claude Code Fix Prompts — AI Diagnosis BRD Gaps

Source: `docs/qa/diagnosis/GAADIIQ_AI_Diagnosis_BRD_and_QA.md`  
BRD readiness: **{score}/100**  
Code base for fixes: **`claude/gaadiiq-app-dev-abj5fo`** (voice stack). Do not implement voice against master-only tree.

---

## MASTER PROMPT (paste into Claude Code)

```
# ROLE
You are implementing GAADIIQ AI Diagnosis BRD gaps as a senior full-stack engineer.
Repo: monorepo with apps/gaadiiq-angular (Capacitor) + apps/api (FastAPI).

# BASELINE (KEEP WORKING — DO NOT REGRESS)
- Manual 4-step wizard on /vehicle-diagnosis
- Voice overlay: 11 Indian languages, Web Speech STT, speechSynthesis TTS
- vehicle-info-extractor + POST /diagnosis/voice/extract
- Missing-field TTS asks, autofill, detected_language → analyse translate
- Image upload POST /upload/image, Ollama + keyword KB + heuristic fallback
- Past Diagnoses history panel + GET /diagnosis/history
- MOB-007 IDOR owner check on GET /diagnosis/{{id}}
- safe_to_drive UI, disclaimers, TTS play/pause/stop/replay

# DO NOT BREAK
Hindi/Bengali/Tamil extraction, rate limits, disclaimer, clientFallback, history list.

# SOURCE OF TRUTH
Read docs/qa/diagnosis/GAADIIQ_AI_Diagnosis_BRD_and_QA.md and implement waves below.
After each wave: add/adjust tests; summarize PASS/FAIL for touched requirement IDs.

# WAVE 1 — P0 (implement first, one PR)
1) BR-SEC-04 + BR-SEC-01 + BR-AND-01 — Mic consent + Android RECORD_AUDIO
   - Pre-STT consent modal: purpose, retention, privacy link, Accept/Decline.
   - Persist consent version+timestamp (local + optional API).
   - AndroidManifest RECORD_AUDIO; Capacitor permission helper before getUserMedia/STT.
   - Denied → clear error + OS settings guidance (never silent fail).

2) BR-API-01 + BR-VA-01 — Server STT fallback
   - POST /diagnosis/stt (multipart audio) via env Whisper/Google/Azure (mockable).
   - If Web Speech unsupported OR WebView flag: MediaRecorder → /diagnosis/stt.
   - Keep browser STT when available.

3) BR-IR-04 + BR-API-07 — Audio upload
   - POST /upload/audio (auth, magic bytes, max 60s, allowlist mime).
   - Step 2 record/upload UI → DiagnoseRequest.audio_url (column already exists).
   - Do not send audio into LLM until validated; store URL on vehicle_diagnoses.

4) BR-TEST-01 — Tests
   - apps/api/tests/test_diagnosis.py: analyse happy/422, history auth, IDOR, voice/extract mocked, stt mocked.
   - Unit tests: vehicle-info-extractor (en + hi) + voice error map.
   - Harden prompt fencing for problem_description (MOB-008).

Acceptance Wave 1: consent on APK path; WebView STT path works with mock STT; pytest diagnosis green.

# WAVE 2 — Multilingual + history depth
- BR-ML-01: Auto-detect toggle wired (optional) + manual picker remains.
- BR-ML-02/04: Localize PROMPTS for kn/ml/gu/pa/or OR translate; show banner if translation fails.
- BR-UX-03: From Past Diagnoses, open GET /diagnosis/{{id}} detail; keep list working.
- BR-AI-10: Return follow_up_questions[] when confidence low; render + optional voice ask.

# WAVE 3 — Media, DB, privacy, RAG
- BR-IR-05: Video upload → video_url.
- BR-IR-07: maintenance_history on Step 2 → API/DB JSON.
- BR-DB-01/02/04: Alembic voice_transcripts, diagnosis_conversations, diagnosis_audit_events (consent-gated).
- BR-SEC-05/06: DELETE diagnosis + delete voice data (DPDP).
- BR-AI-02: Embedding RAG over repair_knowledge (reuse Qdrant/BGE if available) + keyword fallback.
- BR-API-02: Optional server TTS; client speechSynthesis fallback.
- NFR-P1: Ollama timeout (e.g. 8s) → fallback; structured latency logs.

# WAVE 4 — Polish
- BR-VA-08/09: Driving mode (larger targets, hands-free loop).
- BR-UX-01/06: Mobile shell + offline banner/queue.
- BR-PERF-02/03: Golden fixtures for STT/AI (CI informational).
- TC-A-02: aria-live for interim transcript.

# DONE DEFINITION
Requirement IDs in the wave move to PASS (or documented PARTIAL with test proof).
No regression on manual diagnose or Hindi voice extract.
```

---

## Wave 0 — Short context (if splitting chats)

```
Closing GAADIIQ AI Diagnosis BRD gaps on claude/gaadiiq-app-dev-abj5fo.
KEEP: voice overlay, extract, analyse, images, history panel, disclaimers, MOB-007.
DO NOT BREAK: hi/bn/ta extract, safe-to-drive, rate limits.
Implement only listed IDs; add tests for Must-Haves.
```

---

## Wave 1 — P0 only (compact)

```
Implement BR-SEC-04, BR-SEC-01, BR-AND-01, BR-API-01, BR-VA-01, BR-IR-04, BR-API-07, BR-TEST-01, MOB-008 fencing.
Details in docs/qa/diagnosis/Claude_Fix_Prompts_Diagnosis_BRD.md MASTER PROMPT Wave 1.
```

---

## Wave 2 — Multilingual + history depth

```
BR-ML-01 — Wire autoDetectLanguage (optional Auto-detect toggle) + manual picker.
BR-ML-02/04 — Localize PROMPTS for kn/ml/gu/pa/or OR translate; banner if translation fails.
BR-UX-03 — Open diagnosis detail from Past Diagnoses (GET /diagnosis/{{id}}); list already exists.
BR-AI-10 — Return follow_up_questions[] when confidence low; render + optional voice ask.
```

---

## Wave 3 — Media, DB, privacy, RAG

```
BR-IR-05 — Video upload → video_url.
BR-IR-07 — maintenance_history on Step 2 → API/DB JSON.
BR-DB-01/02/04 — Alembic: voice_transcripts, diagnosis_conversations, diagnosis_audit_events (consent-gated).
BR-SEC-05/06 — DELETE diagnosis + delete voice data (DPDP).
BR-AI-02 — Embedding RAG over repair_knowledge (reuse Qdrant if available) + heuristic fallback.
BR-API-02 — Optional server TTS; client speechSynthesis fallback.
NFR — Ollama timeout (e.g. 8s) → fallback; structured logs.
```

---

## Wave 4 — Polish

```
BR-VA-08/09 — Driving mode (larger targets, hands-free loop).
BR-UX-01/06 — Mobile shell + offline banner/queue.
BR-PERF-02/03 — Golden fixtures for STT/AI accuracy (CI informational).
A11y — aria-live for interim transcript.
```

---

## Single-issue template

```
Read docs/qa/diagnosis/GAADIIQ_AI_Diagnosis_BRD_and_QA.md requirement {{REQ_ID}}.
Implement ONLY {{REQ_ID}} on the Claude voice tip branch. Add tests.
Do not regress voice multilingual extract, Past Diagnoses list, or disclaimer.
```
"""
    (QA / "Claude_Fix_Prompts_Diagnosis_BRD.md").write_text(prompts)
    (ART / "Claude_Fix_Prompts_Diagnosis_BRD.md").write_text(prompts)

    summary = f"""# AI Diagnosis — E2E Test Architect Summary

**Date:** {now}  
**Code tip:** `claude/gaadiiq-app-dev-abj5fo`  
**BRD readiness:** **{score}/100**  
**Requirements:** {status_c.get('PASS', 0)} PASS / {status_c.get('PARTIAL', 0)} PARTIAL / {status_c.get('FAIL', 0)} FAIL  
**Test scenarios:** {tc.get('PASS', 0)} PASS / {tc.get('PARTIAL', 0)} PARTIAL / {tc.get('FAIL', 0)} FAIL  

**Verdict:** **NO-GO** for full BRD (server STT/TTS, audio/video, consent, DPDP, tests).  
**Conditional Go** for manual form + images + AI report with disclaimer.

## What works
Manual wizard · voice overlay (11 langs) · extract/autofill · images · Ollama+KB fallback · TTS controls · Past Diagnoses UI · history IDOR fix · disclaimers · `detected_language`.

## Top Claude priorities (Wave 1)
1. Mic consent + Android `RECORD_AUDIO`  
2. Server STT fallback for WebView  
3. Audio upload (`audio_url` column exists)  
4. `test_diagnosis.py` + extractor unit tests  
5. Prompt-injection fencing  

Then: auto language detect · full prompt localization · conversation DB · DPDP delete · video · vector RAG.

## Files
- `GAADIIQ_AI_Diagnosis_BRD_and_QA.md` — full BRD §§1–14 + AC for all requirements  
- `Claude_Fix_Prompts_Diagnosis_BRD.md` — **MASTER PROMPT** + Waves 0–4  
- `diagnosis-brd-qa.json` / `.csv` / `.xlsx`
"""
    (QA / "AI_Diagnosis_E2E_Summary.md").write_text(summary)
    (ART / "AI_Diagnosis_E2E_Summary.md").write_text(summary)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        fills = {
            "PASS": PatternFill("solid", "C6EFCE"),
            "PARTIAL": PatternFill("solid", "FFEB9C"),
            "FAIL": PatternFill("solid", "FFC7CE"),
        }
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive"
        for row in [
            ["GAADIIQ AI Diagnosis BRD QA"],
            ["Generated", now],
            ["Code tip", "claude/gaadiiq-app-dev-abj5fo"],
            ["BRD Score", score],
            ["Req PASS", status_c.get("PASS", 0)],
            ["Req PARTIAL", status_c.get("PARTIAL", 0)],
            ["Req FAIL", status_c.get("FAIL", 0)],
            ["Test PASS", tc.get("PASS", 0)],
            ["Test PARTIAL", tc.get("PARTIAL", 0)],
            ["Test FAIL", tc.get("FAIL", 0)],
            ["Recommendation", catalog["recommendation"]],
        ]:
            ws.append(row)
        ws["A1"].font = Font(bold=True, size=14)

        wr = wb.create_sheet("Requirements")
        wr.append(["ID", "Area", "Priority", "Requirement", "Status", "Evidence", "Gap", "Justification"])
        for r in REQS:
            wr.append(list(r))
            wr.cell(wr.max_row, 5).fill = fills[r[4]]

        wt = wb.create_sheet("TestCases")
        wt.append(["Suite", "ID", "Name", "Status", "Severity", "Detail"])
        for t in tests:
            wt.append([t["suite"], t["id"], t["name"], t["status"], t["severity"], t["detail"]])
            wt.cell(wt.max_row, 4).fill = fills.get(t["status"], PatternFill())

        wa = wb.create_sheet("AcceptanceCriteria")
        wa.append(["Requirement ID", "Status", "Title"])
        for rid, (title, *_rest) in ac_map.items():
            st = next(r[4] for r in REQS if r[0] == rid)
            wa.append([rid, st, title])
            wa.cell(wa.max_row, 2).fill = fills[st]

        out = QA / "GAADIIQ_AI_Diagnosis_BRD_QA.xlsx"
        wb.save(out)
        (ART / "GAADIIQ_AI_Diagnosis_BRD_QA.xlsx").write_bytes(out.read_bytes())
        print("xlsx", out)
    except Exception as e:
        print("xlsx", e)

    print(
        json.dumps(
            {
                "score": score,
                "reqs": dict(status_c),
                "tests": dict(tc),
                "n_reqs": len(REQS),
                "n_tests": len(tests),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
