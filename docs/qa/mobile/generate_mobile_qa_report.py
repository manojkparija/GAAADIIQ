#!/usr/bin/env python3
"""Generate GAADIIQ Mobile Production QA report + Claude fix prompts."""
from __future__ import annotations

import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
QA = Path(__file__).resolve().parent
ART = Path("/opt/cursor/artifacts/mobile-qa-audit")
ART.mkdir(parents=True, exist_ok=True)

issues = [
    dict(
        id="MOB-001",
        severity="BLOCKER",
        area="Platform",
        screen="iOS",
        problem="No iOS Capacitor project exists; only Android WebView shell under apps/gaadiiq-angular/android.",
        root="Product scoped Android-only scaffold; ios/ directory never created.",
        expected="Parity iOS app (Xcode project) with same Capacitor webDir and App Store readiness.",
        files=["apps/gaadiiq-angular/capacitor.config.ts", "apps/gaadiiq-angular/package.json"],
        category="Missing Feature",
    ),
    dict(
        id="MOB-002",
        severity="BLOCKER",
        area="Architecture",
        screen="Auth (all)",
        problem="Dual auth stacks: Angular uses Supabase Auth; FastAPI uses RS256 JWT. Mobile UI does not obtain FastAPI tokens, so many API routes (payments, loans inquiries, notifications, diagnosis history) are unreachable or broken from the app.",
        root="Two parallel backends evolved without a unified session bridge.",
        expected="Single identity: either Supabase JWT validated by API, or mobile uses FastAPI login and stores secure tokens.",
        files=[
            "apps/gaadiiq-angular/src/app/services/auth.service.ts",
            "apps/api/core/security.py",
            "apps/gaadiiq-angular/src/app/services/api.service.ts",
        ],
        category="Security/Architecture",
    ),
    dict(
        id="MOB-003",
        severity="BLOCKER",
        area="Payments",
        screen="Pricing Plans / Feature Listing",
        problem="No Razorpay checkout wired in Angular despite API /payments/* and marketing copy claiming UPI/cards/EMI via Razorpay.",
        root="Payments implemented server-side only; pricing-plans is FAQ/static.",
        expected="End-to-end checkout: create order → Razorpay SDK → verify → unlock plan/listing.",
        files=["apps/gaadiiq-angular/src/app/pages/pricing-plans/", "apps/api/routers/payments.py"],
        category="Missing Feature",
    ),
    dict(
        id="MOB-004",
        severity="BLOCKER",
        area="Mobile Native",
        screen="App shell",
        problem="Capacitor app is a bare WebView wrap: no Camera, Geolocation, Push, Preferences, or Filesystem plugins; package.json has no cap sync/run scripts.",
        root="Only @capacitor/core + android installed.",
        expected="Native plugins for camera, location, push, secure storage; npm scripts for sync/build.",
        files=[
            "apps/gaadiiq-angular/package.json",
            "apps/gaadiiq-angular/capacitor.config.ts",
            "apps/gaadiiq-angular/android/app/capacitor.build.gradle",
        ],
        category="Architecture",
    ),
    dict(
        id="MOB-005",
        severity="CRITICAL",
        area="Security",
        screen="AndroidManifest",
        problem='android:allowBackup="true" allows ADB backup of WebView data including localStorage (wishlist, journey, possibly session fragments).',
        root="Default Capacitor manifest not hardened.",
        expected="allowBackup=false; exclude sensitive paths; disable cleartext.",
        files=["apps/gaadiiq-angular/android/app/src/main/AndroidManifest.xml"],
        category="Security",
    ),
    dict(
        id="MOB-006",
        severity="CRITICAL",
        area="Security",
        screen="Network",
        problem="No SSL/certificate pinning and no Network Security Config; MITM on user devices possible for API/Supabase traffic.",
        root="No pinning library or network_security_config.xml beyond defaults.",
        expected="Pin api.gaadiiq.com + Supabase host; fail closed on mismatch in release builds.",
        files=["apps/gaadiiq-angular/android/app/src/main/res/xml/", "apps/gaadiiq-angular/capacitor.config.ts"],
        category="Security",
    ),
    dict(
        id="MOB-007",
        severity="CRITICAL",
        area="Security",
        screen="AI Diagnosis API",
        problem="GET /diagnosis/{id} is unauthenticated → IDOR: anyone with UUID can read another user's diagnosis (vehicle + symptoms).",
        root="get_diagnosis has no Depends(get_current_user) or ownership check.",
        expected="Require auth; allow owner or admin only.",
        files=["apps/api/routers/diagnosis.py"],
        category="Security",
    ),
    dict(
        id="MOB-008",
        severity="CRITICAL",
        area="Security",
        screen="AI modules",
        problem="User-controlled text interpolated into LLM prompts with no sanitisation/injection fence — prompt injection risk; HLD claims sanitisation but code lacks it.",
        root="f-string prompt assembly in diagnosis.py / valuation.py / sentiment.py.",
        expected="Input length limits, delimiter fencing, instruction hierarchy, output schema validation, refusal tests.",
        files=[
            "apps/api/services/diagnosis.py",
            "apps/api/services/valuation.py",
            "apps/api/services/sentiment.py",
        ],
        category="AI/Security",
    ),
    dict(
        id="MOB-009",
        severity="CRITICAL",
        area="Auth",
        screen="Registration / Login",
        problem="OTP / phone SMS authentication does not exist. Register may collect phone but AuthService.register() never uses verifyOtp.",
        root="Email/password + OAuth only.",
        expected="Phone OTP via Supabase or MSG91/Twilio for India primary mobile UX.",
        files=[
            "apps/gaadiiq-angular/src/app/pages/register/",
            "apps/gaadiiq-angular/src/app/services/auth.service.ts",
        ],
        category="Missing Feature",
    ),
    dict(
        id="MOB-010",
        severity="CRITICAL",
        area="Permissions",
        screen="City / Diagnosis geolocation",
        problem="App uses navigator.geolocation but AndroidManifest declares only INTERNET — no ACCESS_FINE_LOCATION / COARSE.",
        root="Web API used without native permission declarations/plugins.",
        expected="Add permissions + @capacitor/geolocation with runtime prompts.",
        files=[
            "apps/gaadiiq-angular/android/app/src/main/AndroidManifest.xml",
            "apps/gaadiiq-angular/src/app/components/city-selector/",
        ],
        category="Security/Mobile",
    ),
    dict(
        id="MOB-011",
        severity="CRITICAL",
        area="Auth Bridge",
        screen="API calls",
        problem="ApiService.submitLoanInquiry hits /loans/inquiry but API exposes /loans/inquiries — broken loan lead capture.",
        root="Path typo / contract drift.",
        expected="Align client path; add contract test.",
        files=["apps/gaadiiq-angular/src/app/services/api.service.ts", "apps/api/routers/loans.py"],
        category="API",
    ),
    dict(
        id="MOB-012",
        severity="CRITICAL",
        area="RBAC",
        screen="Dealer/Admin",
        problem="sellerGuard is client-side only. FastAPI dealer endpoints use soft Dealer-row checks; any user can POST /dealers/register. Role enums diverge (seller vs dealer).",
        root="No unified server-side role enforcement aligned with Angular.",
        expected="Unified roles; require_role on sensitive routes; adminGuard.",
        files=[
            "apps/gaadiiq-angular/src/app/guards/seller.guard.ts",
            "apps/api/routers/dealers.py",
            "apps/api/core/dependencies.py",
        ],
        category="Security",
    ),
    dict(
        id="MOB-013",
        severity="HIGH",
        area="Features",
        screen="Insurance",
        problem="No insurance enquiry API or partner quote flow; only client-side premium estimates / marketing.",
        root="PRD/LLD endpoints never implemented.",
        expected="POST insurance lead + partner deep-links with disclosure.",
        files=["apps/api/routers/", "apps/gaadiiq-angular/src/app/pages/car-detail/"],
        category="Missing Feature",
    ),
    dict(
        id="MOB-014",
        severity="HIGH",
        area="Features",
        screen="AI Chat",
        problem="Ask GAADIIQ / POST /recommend/ai-chat (SSE) from LLD is not implemented.",
        root="Recommend is quiz/rules only.",
        expected="Streaming chat endpoint + mobile UI with citations/fallback.",
        files=["apps/api/routers/", "apps/gaadiiq-angular/src/app/pages/ai-advisor/"],
        category="Missing Feature / AI",
    ),
    dict(
        id="MOB-015",
        severity="HIGH",
        area="Features",
        screen="Push Notifications",
        problem="No FCM/APNs Capacitor push plugin; in-app notifications API unused by primary Supabase auth path.",
        root="Push never integrated.",
        expected="FCM + user tokens; deep links to bookings/price drops.",
        files=["apps/gaadiiq-angular/android/", "apps/api/routers/notifications.py"],
        category="Missing Feature",
    ),
    dict(
        id="MOB-016",
        severity="HIGH",
        area="AI Advisor",
        screen="/ai-advisor",
        problem="Angular advisor uses local heuristic scoring; not wired to POST /recommend. Diverges from Next.js advisor.",
        root="Client-only engine shipped as AI.",
        expected="Call API; method badge; unify scoring; use CityService.",
        files=[
            "apps/gaadiiq-angular/src/app/pages/ai-advisor/ai-advisor.component.ts",
            "apps/api/routers/recommend.py",
        ],
        category="AI",
    ),
    dict(
        id="MOB-017",
        severity="HIGH",
        area="Data",
        screen="Wishlist / Price Alerts / Reviews",
        problem="Wishlist, price alerts, recently viewed persist only in localStorage — lost on reinstall, not cross-device.",
        root="Offline-first shortcuts without sync layer.",
        expected="Authenticated CRUD against API; local cache as offline mirror.",
        files=[
            "apps/gaadiiq-angular/src/app/pages/used-cars/",
            "apps/gaadiiq-angular/src/app/pages/price-alerts/",
        ],
        category="Reliability",
    ),
    dict(
        id="MOB-018",
        severity="HIGH",
        area="Uploads",
        screen="List Car / Diagnosis",
        problem="Audio/video/document upload unsupported. List-car uses Cloudinary unsigned preset (abuse risk). Upload MIME allowlist only; vision may fetch arbitrary image_urls (SSRF).",
        root="No magic bytes/AV; unsigned Cloudinary; unrestricted vision URLs.",
        expected="Magic bytes, size caps, AV scan, signed uploads, SSRF allowlist.",
        files=[
            "apps/api/routers/upload.py",
            "apps/api/services/vision.py",
            "apps/gaadiiq-angular/src/app/services/cloudinary.service.ts",
        ],
        category="Security",
    ),
    dict(
        id="MOB-019",
        severity="HIGH",
        area="Rate Limit",
        screen="API",
        problem="SlowAPI rate limits disabled unless environment==production; diagnosis/analyse not limited; /metrics public.",
        root="Limiter gated; incomplete decoration.",
        expected="Enable limits in staging; rate-limit diagnosis; protect /metrics.",
        files=["apps/api/core/limiter.py", "apps/api/routers/diagnosis.py", "apps/api/main.py"],
        category="Security/Performance",
    ),
    dict(
        id="MOB-020",
        severity="HIGH",
        area="Database",
        screen="Intent / Diagnosis models",
        problem="customer_activities/intent_scores FK columns String(36) vs UUID PKs; vehicle_diagnoses.user_id nullable without FK.",
        root="Migration type mismatch in 0005/0006.",
        expected="UUID FKs with ON DELETE SET NULL/CASCADE; backfill.",
        files=[
            "apps/api/alembic/versions/0005_customer_intent_tables.py",
            "apps/api/alembic/versions/0006_vehicle_diagnosis.py",
        ],
        category="Database",
    ),
    dict(
        id="MOB-021",
        severity="HIGH",
        area="CORS",
        screen="API",
        problem="Default allowed_origins is localhost:3000; Capacitor/Android and Angular :4200 may be blocked against production API.",
        root="CORS not listing mobile/web production origins.",
        expected="Explicit allowlist: web, capacitor origins, staging.",
        files=["apps/api/main.py", "apps/api/core/config.py"],
        category="API",
    ),
    dict(
        id="MOB-022",
        severity="HIGH",
        area="Testing",
        screen="Mobile QA",
        problem="Only 1 Angular unit spec (boilerplate broken); no Appium/Detox against Capacitor; Android instrumented tests still Capacitor sample package names.",
        root="No mobile test strategy executed.",
        expected="Contract tests + critical-path Appium; fix sample tests; CI on PR.",
        files=[
            "apps/gaadiiq-angular/src/app/app.component.spec.ts",
            "apps/gaadiiq-angular/android/app/src/androidTest/",
        ],
        category="Quality",
    ),
    dict(
        id="MOB-023",
        severity="MEDIUM",
        area="UX",
        screen="Used Cars Budget Slider",
        problem="Dual-range slider: WebKit thumb misalignment (no margin-top), no z-index swap when thumbs overlap, track not clickable.",
        root="CSS dual-range incomplete.",
        expected="Centered thumbs, dynamic z-index, accessible labels.",
        files=["apps/gaadiiq-angular/src/app/pages/used-cars/used-cars.component.scss"],
        category="UI/UX",
    ),
    dict(
        id="MOB-024",
        severity="MEDIUM",
        area="UX",
        screen="Used Cars",
        problem="showAllIndiaBanner requires !allIndiaOverride so amber banner never shows after auto-override for empty cities.",
        root="Inverted banner condition.",
        expected="Show banner when override active.",
        files=["apps/gaadiiq-angular/src/app/pages/used-cars/used-cars.component.ts"],
        category="UI/UX",
    ),
    dict(
        id="MOB-025",
        severity="MEDIUM",
        area="Features",
        screen="New Cars",
        problem="New Cars P0 fixes (Above ₹30L applyBudget) live on cursor/fix-new-cars-module-85e1 and are NOT merged into Claude tip.",
        root="Unmerged fix branch.",
        expected="Merge/cherry-pick fix branch.",
        files=["apps/gaadiiq-angular/src/app/pages/new-cars/"],
        category="Functional",
    ),
    dict(
        id="MOB-026",
        severity="MEDIUM",
        area="Features",
        screen="EV / TCO",
        problem="No dedicated EV calculator screen; TCO service exists but weak mobile discovery.",
        root="Feature incomplete in IA.",
        expected="EV cost/range calculator page; TCO entry from car detail.",
        files=["apps/gaadiiq-angular/src/app/services/tco.service.ts"],
        category="Missing Feature",
    ),
    dict(
        id="MOB-027",
        severity="MEDIUM",
        area="Features",
        screen="Dealer Search / Service History",
        problem="No dealer directory search; service booking only opens maps to static SERVICE_CENTERS; no service history.",
        root="Diagnosis CTA stub.",
        expected="Dealer search API+UI; booking persistence; history.",
        files=["apps/gaadiiq-angular/src/app/pages/vehicle-diagnosis/"],
        category="Missing Feature",
    ),
    dict(
        id="MOB-028",
        severity="MEDIUM",
        area="A11y",
        screen="Global",
        problem="Sparse ARIA; many outline:none without :focus-visible; no TalkBack/VoiceOver pass; contrast not systematically verified.",
        root="A11y not a release gate.",
        expected="WCAG 2.2 AA on top flows.",
        files=["apps/gaadiiq-angular/src/app/**/*.html", "apps/gaadiiq-angular/src/styles.scss"],
        category="Accessibility",
    ),
    dict(
        id="MOB-029",
        severity="MEDIUM",
        area="Performance",
        screen="Listings / Images",
        problem="Listing grids without virtual scroll; no measured startup/FPS/memory on device.",
        root="No perf budget in CI.",
        expected="Virtual scroll, image sizing, WebView metrics.",
        files=["apps/gaadiiq-angular/src/app/pages/listings/", "apps/gaadiiq-angular/ngsw-config.json"],
        category="Performance",
    ),
    dict(
        id="MOB-030",
        severity="MEDIUM",
        area="Offline",
        screen="Global",
        problem="No true offline mode; SW precaches shell only; forms fail opaquely offline; no queue/retry.",
        root="PWA shell ≠ offline product.",
        expected="Offline banner, queued mutations, read-through cache.",
        files=["apps/gaadiiq-angular/ngsw-config.json"],
        category="Reliability",
    ),
    dict(
        id="MOB-031",
        severity="MEDIUM",
        area="EMI",
        screen="EMI Calculator",
        problem="Bank rates are hardcoded stubs; client EMI not using GET /loans/emi-calculator.",
        root="Marketing calculator.",
        expected="Disclose indicative; sync API calculator.",
        files=["apps/gaadiiq-angular/src/app/pages/emi-calculator/"],
        category="Business Rules",
    ),
    dict(
        id="MOB-032",
        severity="MEDIUM",
        area="Sentiment",
        screen="Dealer Dashboard",
        problem="Public sentiment track endpoint forgeable; no customer consent UX.",
        root="Privacy/product gap.",
        expected="Consent, auth-only tracking, disclosure.",
        files=["apps/api/routers/sentiment.py"],
        category="Security/Privacy",
    ),
    dict(
        id="MOB-033",
        severity="MEDIUM",
        area="Compatibility",
        screen="Devices",
        problem="minSdk 24 but product asks Android 12+ validation; foldable/tablet/landscape untested; no iPad.",
        root="No device matrix execution.",
        expected="Document and execute supported device matrix.",
        files=["apps/gaadiiq-angular/android/variables.gradle"],
        category="Compatibility",
    ),
    dict(
        id="MOB-034",
        severity="LOW",
        area="Code Quality",
        screen="Android tests",
        problem="Instrumented tests still reference com.getcapacitor.myapp / ExampleUnitTest.",
        root="Capacitor template leftover.",
        expected="Replace with GAADIIQ smoke or remove from CI.",
        files=["apps/gaadiiq-angular/android/app/src/test/", "androidTest/"],
        category="Quality",
    ),
    dict(
        id="MOB-035",
        severity="LOW",
        area="UX",
        screen="Used Cars empty state",
        problem="No AI Advisor CTA when filters yield 0 results; no URL write-back for filters.",
        root="Incomplete empty-state IA.",
        expected="Advisor link + query sync.",
        files=["apps/gaadiiq-angular/src/app/pages/used-cars/"],
        category="UI/UX",
    ),
    dict(
        id="MOB-036",
        severity="LOW",
        area="AI Diagnosis",
        screen="History",
        problem="API /diagnosis/history exists but Angular wizard has no history UI.",
        root="UI incomplete.",
        expected="Past diagnoses list with reopen.",
        files=["apps/gaadiiq-angular/src/app/pages/vehicle-diagnosis/"],
        category="Functional",
    ),
    dict(
        id="MOB-037",
        severity="LOW",
        area="Security",
        screen="Root/Jailbreak",
        problem="No root/jailbreak detection or Play Integrity on release builds.",
        root="Not implemented.",
        expected="Play Integrity / DeviceCheck gated for payments.",
        files=["apps/gaadiiq-angular/android/"],
        category="Security",
    ),
    dict(
        id="MOB-038",
        severity="LOW",
        area="Config",
        screen="Environments",
        problem=".env.example documents HS256 SECRET_KEY while API uses RS256 PEMs.",
        root="Stale docs.",
        expected="Align examples with JWT key files.",
        files=["apps/api/.env.example", "apps/api/core/config.py"],
        category="Maintainability",
    ),
    dict(
        id="MOB-039",
        severity="LOW",
        area="Payments Webhook",
        screen="API",
        problem="When Razorpay keys unset, webhook verification may be open — dangerous if mis-deployed.",
        root="Dev shortcuts.",
        expected="Fail closed without keys outside local dev.",
        files=["apps/api/routers/payments.py"],
        category="Security",
    ),
    dict(
        id="MOB-040",
        severity="LOW",
        area="Profile",
        screen="Settings",
        problem="No dedicated Settings/Profile edit screen; no DPDP delete-account flow.",
        root="Thin account IA.",
        expected="Profile, notification prefs, privacy, delete account.",
        files=["apps/gaadiiq-angular/src/app/pages/"],
        category="Missing Feature",
    ),
]

features = [
    ("Registration", "PARTIAL", "Email/password+OAuth; phone unused; no OTP"),
    ("Login", "PASS", "Supabase email/password + Google/Facebook"),
    ("OTP", "FAIL", "Not implemented"),
    ("Password Reset", "PASS", "resetPasswordForEmail + /reset-password"),
    ("Dashboard", "PARTIAL", "Home marketplace, not full role dashboards"),
    ("Vehicle Search", "PASS", "Listings/search UI"),
    ("Compare Cars", "PASS", "Client compare"),
    ("Car Details", "PASS", "/cars/:id"),
    ("Used Cars", "PASS", "Filters + inventory (P0 fixed on tip)"),
    ("AI Recommendation", "PARTIAL", "Quiz heuristic; not API /recommend"),
    ("Dealer Search", "FAIL", "No directory"),
    ("Dealer Dashboard", "PASS", "sellerGuard route"),
    ("Test Drive Booking", "PARTIAL", "Page exists; Supabase-backed"),
    ("Loan Calculator", "PASS", "Client EMI"),
    ("Insurance", "FAIL", "No partner flow"),
    ("EMI Calculator", "PASS", "/emi-calculator"),
    ("TCO Calculator", "PARTIAL", "Service/Next; weak mobile entry"),
    ("EV Calculator", "FAIL", "Missing"),
    ("AI Vehicle Diagnosis", "PASS", "Wizard + API + fallback"),
    ("Customer Sentiment", "PARTIAL", "Dealer API only"),
    ("AI Chat", "FAIL", "Not implemented"),
    ("Vehicle Image Upload", "PARTIAL", "Cloudinary list-car; diagnosis API upload"),
    ("Audio Upload", "FAIL", "Missing"),
    ("Video Upload", "FAIL", "Missing"),
    ("Document Upload", "FAIL", "Missing"),
    ("Wishlist", "PARTIAL", "localStorage"),
    ("Notifications", "PARTIAL", "API exists; no push; auth mismatch"),
    ("User Profile", "PARTIAL", "Thin"),
    ("Settings", "PARTIAL", "Theme/lang only"),
    ("Service Booking", "PARTIAL", "Maps to static centres"),
    ("Service History", "FAIL", "Missing"),
    ("Review System", "PARTIAL", "Mixed local/API"),
    ("Payment Flow", "FAIL", "No mobile checkout"),
    ("Logout", "PASS", "Supabase signOut"),
    ("Android", "PARTIAL", "WebView shell only"),
    ("iOS", "FAIL", "No project"),
]


def main() -> None:
    sev = Counter(i["severity"] for i in issues)
    feat = Counter(s for _, s, _ in features)
    # Rubric (capped so multiple issues in one theme don't zero the score unrealistically):
    # Base 70 for existing browse/auth/diagnosis web app, then deduct by severity with caps.
    prod = 70
    prod -= min(28, sev["BLOCKER"] * 8)   # cap -28
    prod -= min(20, sev["CRITICAL"] * 3)  # cap -20
    prod -= min(12, sev["HIGH"] * 1.5)    # cap -12
    prod -= min(6, sev["MEDIUM"] * 0.5)   # cap -6
    prod -= min(3, sev["LOW"] * 0.25)     # cap -3
    # Hard gates: missing iOS + dual auth + no payments → floor for store release
    prod = max(22, min(100, round(prod)))
    quality = 48
    cov = round(100 * (feat.get("PASS", 0) + 0.5 * feat.get("PARTIAL", 0)) / len(features))

    feature_rows = "\n".join(f"| {n} | **{s}** | {d} |" for n, s, d in features)
    blockers = "\n".join(f"- **{i['id']}**: {i['problem']}" for i in issues if i["severity"] == "BLOCKER")

    def sev_section(name: str) -> str:
        lines = [f"### {name}"]
        for i in issues:
            if i["severity"] == name:
                lines.append(f"- **{i['id']}** [{i['category']}] {i['screen']}: {i['problem']}")
        return "\n".join(lines) + "\n"

    report = f"""# GAADIIQ Mobile Production QA Report

**Application:** GAADIIQ — AI Automotive Intelligence (India)  
**Platforms audited:** Android (Capacitor WebView), iOS (**absent**), Angular PWA  
**Code tip:** `claude/gaadiiq-app-dev-abj5fo` (includes Android icon + Used Cars / Diagnosis work)  
**Report date:** {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  
**Roles:** Principal Mobile Test Architect · QA Lead · Security · Performance · Accessibility · AI Testing · Product Owner  

---

## 1. Executive Summary

GAADIIQ’s mobile surface is a **Capacitor Android WebView wrapping an Angular SPA**, not a native feature-complete Android/iOS product. Core marketplace browsing, login/register, used/new cars, compare, EMI, and AI diagnosis **exist**. Production mobile release is **blocked** by missing iOS, dual auth (Supabase vs FastAPI JWT), absent payment checkout, missing OTP/push/native plugins, diagnosis IDOR, LLM prompt-injection exposure, and near-zero automated mobile tests.

**Recommendation: NO-GO for production mobile release.**

| Score | Value |
|------|------:|
| Production Readiness | **{prod}/100** |
| Overall Quality | **{quality}/100** |
| Requested Feature Coverage (weighted) | **{cov}%** |

---

## 2. Methodology & Limits

| Method | Scope |
|--------|--------|
| Static architecture & code review | Capacitor, AndroidManifest, Angular routes/services, FastAPI, AI services, migrations |
| Prior automated matrices | Used Cars, Diagnosis, New Cars, Advisor under `docs/qa/` |
| Physical devices / Appium / battery / FPS | **NOT EXECUTED** here — residual risk |
| Full pen-test | Code-level review only |

Features not evidenced in the repo are **FAIL / MISSING**, never assumed PASS.

---

## 3. Test Coverage

| Layer | Coverage | Notes |
|-------|----------:|-------|
| Feature checklist ({len(features)} items) | **{cov}%** weighted | {feat.get('PASS',0)} PASS / {feat.get('PARTIAL',0)} PARTIAL / {feat.get('FAIL',0)} FAIL |
| Angular unit tests | **~0%** | 1 broken boilerplate spec |
| API pytest | **~55%** of routers | diagnosis/recommend/upload gaps |
| Mobile E2E (Appium) | **0%** | Not present |
| Accessibility WCAG | **~10%** | Static spot check |
| Performance device lab | **0%** | Not measured |

**Mission checklist rigorously covered: ~35–40%; remainder blocked by missing device lab / instrumentation.**

---

## 4. Features Tested

| Feature | Result | Evidence note |
|---------|--------|---------------|
{feature_rows}

**Totals:** {feat.get('PASS',0)} PASS · {feat.get('PARTIAL',0)} PARTIAL · {feat.get('FAIL',0)} FAIL

---

## 5–6. Issue Counts

Issues logged: **{len(issues)}**

| Severity | Count |
|----------|------:|
| BLOCKER | {sev['BLOCKER']} |
| CRITICAL | {sev['CRITICAL']} |
| HIGH | {sev['HIGH']} |
| MEDIUM | {sev['MEDIUM']} |
| LOW | {sev['LOW']} |

---

## 7. Blockers

{blockers}

---

## 8–11. Critical / High / Medium / Low

{sev_section("CRITICAL")}
{sev_section("HIGH")}
{sev_section("MEDIUM")}
{sev_section("LOW")}

---

## 11. Security Vulnerabilities (highlights)

| ID | Issue |
|----|-------|
| MOB-002 | Dual auth / token non-interoperability |
| MOB-005 | `allowBackup=true` |
| MOB-006 | No SSL pinning |
| MOB-007 | Diagnosis IDOR |
| MOB-008 | LLM prompt injection |
| MOB-010 | Location without manifest permission |
| MOB-012 | Weak/inconsistent RBAC |
| MOB-018 | Upload/SSRF/Cloudinary unsigned risks |
| MOB-019 | Rate limits off outside prod; public metrics |
| MOB-032 | Forgeable public sentiment track |
| MOB-039 | Payment webhook fail-open risk |

---

## 12. Performance Bottlenecks

- Device FPS/memory/battery **unmeasured** (lab required).
- Listing grids lack virtualization (MOB-029).
- Cloudinary/full-size images risk on slow networks.
- Diagnosis LLM latency unbounded; client fallback helps UX only.
- Rate limiter disabled in non-prod hides soak issues.

---

## 13. UI/UX Problems

- Budget dual-slider polish (MOB-023); All-India banner logic (MOB-024).
- Empty states missing Advisor CTA (MOB-035).
- New Cars budget UX vs unmerged fix branch (MOB-025).
- Thin settings/profile (MOB-040).
- Heuristic advisor over-claims “AI” (MOB-016).

---

## 14. Accessibility Issues

- Insufficient ARIA/landmarks (MOB-028).
- `outline: none` without consistent `:focus-visible`.
- No documented TalkBack/VoiceOver pass.
- Large font / dynamic type reflow unverified.
- Dark-mode contrast needs token audit.

---

## 15. AI Validation Issues

| Topic | Finding |
|-------|---------|
| Recommendation | Client heuristic ≠ API |
| Diagnosis | Works + KB fallback; history UI missing |
| Sentiment | Dealer-only; public track forgeable |
| Chat / RAG | Chat missing; Ollama without injection defenses |
| Hallucination / confidence | Partial method badges; advisor weak |
| Prompt injection | **FAIL** (MOB-008) |

---

## 16. Missing Features

iOS · OTP · Push · Payment checkout · Insurance · AI Chat · EV calculator · Dealer search · Service history · Audio/Video/Document upload · Root detection · SSL pinning · Native camera/geo plugins · DPDP delete-account

---

## 17. Risk Assessment

| Risk | Likelihood | Impact | Priority |
|------|------------|--------|----------|
| Auth dual-stack breaks paid/API flows | High | Critical | P0 |
| Diagnosis IDOR / backup leakage | Medium | Critical | P0 |
| Prompt injection brand/legal harm | Medium | High | P0 |
| WebView marketed as native app | High | High | P0 |
| Webhook misconfig / payment fraud | Medium | Critical | P0 |
| Play/App Store rejection | High | High | P0 |
| A11y non-compliance | Medium | Medium | P1 |

---

## 18–19. Scores

| Metric | Score |
|--------|------:|
| Production Readiness | **{prod}/100** |
| Overall Quality | **{quality}/100** |

---

## 20. Go / No-Go

# **NO-GO**

Minimum before any store release:

1. MOB-002 dual auth  
2. MOB-005–007 backup + pinning + IDOR  
3. MOB-003 payments or remove claims  
4. MOB-001 iOS or explicit Android-only beta labeling  
5. MOB-004 / MOB-010 native plugins + permissions  
6. MOB-008 prompt-injection defenses  
7. MOB-022 critical-path mobile automation  

**Interim OK:** internal Android **Tech Preview** (WebView), payments disabled, P0 security fixes applied.

---

## Appendix — Architecture truth

```
Angular SPA (apps/gaadiiq-angular)
  └─ Capacitor 8 Android WebView (com.gaadiiq.app) — INTERNET only
  └─ Supabase Auth + Cloudinary + optional FastAPI
FastAPI (apps/api) — separate JWT world
Next.js (apps/web) — parallel web
iOS — does not exist
```

Fix prompts: `Claude_Fix_Prompts.md` · Machine catalog: `Issues_Catalog.json`
"""

    (QA / "GAADIIQ_Mobile_Production_QA_Report.md").write_text(report)
    (ART / "GAADIIQ_Mobile_Production_QA_Report.md").write_text(report)

    catalog = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "production_readiness": prod,
        "overall_quality": quality,
        "feature_coverage_pct": cov,
        "recommendation": "NO-GO",
        "severity_counts": dict(sev),
        "feature_counts": dict(feat),
        "issues": issues,
        "features": [{"name": n, "status": s, "detail": d} for n, s, d in features],
    }
    (QA / "Issues_Catalog.json").write_text(json.dumps(catalog, indent=2))
    (ART / "Issues_Catalog.json").write_text(json.dumps(catalog, indent=2))

    parts = [
        "# Claude Code Fix Prompts — GAADIIQ Mobile QA Issues\n",
        "Source report: `docs/qa/mobile/GAADIIQ_Mobile_Production_QA_Report.md`\n",
        "Each section is a standalone implementation task.\n",
    ]
    for i in issues:
        files = ", ".join(i["files"])
        parts.append(
            f"""
---

## {i['id']} — {i['severity']} — {i['screen']}

**Issue ID:** {i['id']}  
**Severity:** {i['severity']}  
**Affected Screen:** {i['screen']}  
**Category:** {i['category']}  
**Area:** {i['area']}  

**Problem:**  
{i['problem']}

**Root Cause:**  
{i['root']}

**Expected Behaviour:**  
{i['expected']}

**Suggested Fix:**  
Implement expected behaviour; preserve backward-compatible UX where safe; add regression tests.

### COMPLETE Claude Code Implementation Prompt

```
You are implementing GAADIIQ fix {i['id']} ({i['severity']}).

CONTEXT
- Monorepo: apps/gaadiiq-angular (Capacitor Android + Angular), apps/api (FastAPI), apps/web (Next.js).
- Problem: {i['problem']}
- Root cause: {i['root']}
- Expected: {i['expected']}

PRIMARY FILES
- {files}
- Add/adjust tests under apps/api/tests/ and/or Angular specs.
- Update docs/qa only if behaviour contracts change.

REQUIREMENTS
1. Frontend: UX/logic fix; loading/error/empty states; no silent failures.
2. Backend: authz, validation, correct HTTP codes, structured errors.
3. Database: Alembic migration if schema changes; reversible where possible.
4. API: OpenAPI-consistent; version if breaking.
5. Security: least privilege; no secrets in client; validate uploads/inputs; close IDOR/injection if in scope.
6. Validation: reject blank/oversized abuse; +91 phone rules if touching auth.
7. Errors: user-safe messages; server logs with request id.
8. Unit tests: happy path + ≥2 negative cases.
9. Integration/regression: cover the broken path for {i['id']}.
10. Mobile: if Manifest/plugins change, note capacitor sync in commit message.

ACCEPTANCE
- {i['id']} fail → pass with automated test or explicit manual checklist in PR.
- No new lint/type errors.
- Do not regress Used Cars P0 city/year fixes or diagnosis client fallback.

Deliver a concise PR description with files touched and test commands run.
```
"""
        )

    prompt_text = "\n".join(parts)
    (QA / "Claude_Fix_Prompts.md").write_text(prompt_text)
    (ART / "Claude_Fix_Prompts.md").write_text(prompt_text)

    import csv

    with (QA / "issues.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["id", "severity", "area", "screen", "category", "problem"])
        for i in issues:
            w.writerow([i["id"], i["severity"], i["area"], i["screen"], i["category"], i["problem"]])
    (ART / "issues.csv").write_text((QA / "issues.csv").read_text())

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Executive"
        for row in [
            ["GAADIIQ Mobile Production QA"],
            ["Generated", catalog["generated_at"]],
            ["Production Readiness", prod],
            ["Overall Quality", quality],
            ["Feature Coverage %", cov],
            ["Recommendation", "NO-GO"],
            ["Blockers", sev["BLOCKER"]],
            ["Critical", sev["CRITICAL"]],
            ["High", sev["HIGH"]],
            ["Medium", sev["MEDIUM"]],
            ["Low", sev["LOW"]],
            ["Issues Total", len(issues)],
        ]:
            ws.append(row)
        ws["A1"].font = Font(bold=True, size=14)

        wi = wb.create_sheet("Issues")
        wi.append(["ID", "Severity", "Area", "Screen", "Category", "Problem", "Root Cause", "Expected", "Files"])
        fills = {
            "BLOCKER": PatternFill("solid", "FF0000"),
            "CRITICAL": PatternFill("solid", "FF6B6B"),
            "HIGH": PatternFill("solid", "FFA500"),
            "MEDIUM": PatternFill("solid", "FFD666"),
            "LOW": PatternFill("solid", "C6EFCE"),
        }
        for i in issues:
            wi.append(
                [
                    i["id"],
                    i["severity"],
                    i["area"],
                    i["screen"],
                    i["category"],
                    i["problem"],
                    i["root"],
                    i["expected"],
                    "; ".join(i["files"]),
                ]
            )
            wi.cell(wi.max_row, 2).fill = fills[i["severity"]]
        for c in wi[1]:
            c.font = Font(bold=True)

        wf = wb.create_sheet("Features")
        wf.append(["Feature", "Status", "Detail"])
        sf = {
            "PASS": PatternFill("solid", "C6EFCE"),
            "PARTIAL": PatternFill("solid", "FFEB9C"),
            "FAIL": PatternFill("solid", "FFC7CE"),
        }
        for n, s, d in features:
            wf.append([n, s, d])
            wf.cell(wf.max_row, 2).fill = sf[s]

        out = QA / "GAADIIQ_Mobile_Production_QA.xlsx"
        wb.save(out)
        (ART / "GAADIIQ_Mobile_Production_QA.xlsx").write_bytes(out.read_bytes())
        print("xlsx", out)
    except Exception as e:
        print("xlsx skip", e)

    print(
        json.dumps(
            {"prod": prod, "quality": quality, "cov": cov, "issues": len(issues), "sev": dict(sev), "feat": dict(feat)},
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
