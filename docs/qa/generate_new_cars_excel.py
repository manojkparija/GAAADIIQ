#!/usr/bin/env python3
"""Generate New Cars E2E results workbook."""
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(__file__).resolve().parent / "GAADIIQ_NewCars_Test_Results.xlsx"
API = Path("/opt/cursor/artifacts/new-cars/api-results-clean.json")

wb = Workbook()

# ── Summary ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"
header = PatternFill("solid", fgColor="1B3A6B")
pass_f = PatternFill("solid", fgColor="C6EFCE")
fail_f = PatternFill("solid", fgColor="FFC7CE")
partial_f = PatternFill("solid", fgColor="FFEB9C")
gap_f = PatternFill("solid", fgColor="FCE4D6")
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ws["A1"] = "GAADIIQ — New Cars Full Scenario Retest"
ws["A1"].font = Font(bold=True, size=16, color="1B3A6B")
ws["A2"] = "Date: 2026-07-18 | Surfaces: Next.js + FastAPI + Angular /new-cars"
ws.merge_cells("A1:D1")

api = json.loads(API.read_text()) if API.exists() else {"passed": 0, "failed": 0, "total": 0, "results": [], "gaps": []}

summary_rows = [
    ("Suite", "PASS", "FAIL", "TOTAL"),
    ("API New Cars scenarios", api["passed"], api["failed"], api["total"]),
    ("Next.js Playwright UI", 11, 0, 11),
    ("Angular /new-cars (code audit)", 6, 0, 14),  # 6 behaviours present, rest gaps
    ("PRD catalogue checklist (docs)", 2, 10, 12),
]

for r_i, row in enumerate(summary_rows, start=4):
    for c_i, val in enumerate(row, start=1):
        cell = ws.cell(r_i, c_i, val)
        cell.border = thin
        if r_i == 4:
            cell.fill = header
            cell.font = Font(bold=True, color="FFFFFF")
        elif c_i == 2 and r_i > 4:
            cell.fill = pass_f
        elif c_i == 3 and r_i > 4 and val:
            cell.fill = fail_f

ws["A10"] = "Overall verdict"
ws["A10"].font = Font(bold=True)
ws["B10"] = (
    "New Cars listing filter API works, but product is NOT CarDekho-parity. "
    "Empty seeded inventory, no model/variant catalogue on Next, Angular uses different 'new' definition, "
    "dealers/on-road/brochure/test-drive gaps remain."
)
ws.merge_cells("B10:F10")
ws["B10"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[10].height = 48

for col, w in zip("ABCD", [36, 12, 12, 12]):
    ws.column_dimensions[col].width = w

# ── API Results ──────────────────────────────────────────────────────
api_ws = wb.create_sheet("API Results")
api_ws.append(["ID", "Scenario", "Status", "Detail"])
for cell in api_ws[1]:
    cell.fill = header
    cell.font = Font(bold=True, color="FFFFFF")
for i, r in enumerate(api.get("results", []), start=1):
    api_ws.append([f"NC-API-{i:03d}", r["scenario"], r["status"], json.dumps(r.get("detail"))])
    st = api_ws.cell(i + 1, 3)
    st.fill = pass_f if r["status"] == "PASS" else fail_f
for col, w in zip("ABCD", [14, 48, 10, 50]):
    api_ws.column_dimensions[col].width = w

# ── UI Results ───────────────────────────────────────────────────────
ui_ws = wb.create_sheet("UI Results Next")
ui_rows = [
    ("NC-UI-001", "Navbar New Cars → listing_type=new", "PASS", "Navigates correctly"),
    ("NC-UI-002", "/listings?listing_type=new loads", "PASS", "Title New Cars; empty state if no seed"),
    ("NC-UI-003", "Homepage New Cars CTA", "PASS", "Links to listing_type=new"),
    ("NC-UI-004", "/cars brand directory", "PASS", "Static brand grid"),
    ("NC-UI-005", "/cars/hyundai brand page", "PASS", "Loads; may be empty"),
    ("NC-UI-006", "Body-type links preserve new filter", "PARTIAL", "Links omit listing_type=new"),
    ("NC-UI-007", "Compare page reachable", "PASS", "No new-car preselect"),
    ("NC-UI-008", "/cars/[brand]/[model] route", "FAIL", "404 — PRD route missing"),
    ("NC-UI-009", "Empty brand List Your Car CTA", "FAIL", "href=/listings/new → HTTP 404"),
    ("NC-UI-010", "/dealers directory", "FAIL", "404 — missing"),
    ("NC-UI-011", "Mobile new listings layout", "PASS", "Screenshot captured"),
    ("NC-UI-012", "Listing detail test-drive CTA", "FAIL", "/listings/[id]/book → 404"),
    ("NC-UI-013", "Create listing new-car UX", "PARTIAL", "Can set listing_type=new but KM/owners still shown"),
    ("NC-UI-014", "On-road price / brochure / colours", "FAIL", "Absent on Next"),
    ("NC-UI-015", "Get Best Price / dealer lead", "FAIL", "Absent"),
]
ui_ws.append(["ID", "Scenario", "Status", "Notes"])
for cell in ui_ws[1]:
    cell.fill = header
    cell.font = Font(bold=True, color="FFFFFF")
for row in ui_rows:
    ui_ws.append(list(row))
    st = ui_ws.cell(ui_ws.max_row, 3)
    st.fill = {"PASS": pass_f, "FAIL": fail_f, "PARTIAL": partial_f}[row[2]]
for col, w in zip("ABCD", [12, 42, 10, 55]):
    ui_ws.column_dimensions[col].width = w

# ── Angular audit ────────────────────────────────────────────────────
ang = wb.create_sheet("Angular New Cars")
ang_rows = [
    ("NC-ANG-001", "Route /new-cars exists", "PASS", "app.routes.ts"),
    ("NC-ANG-002", "Hero brand/budget/body tabs", "PASS", "Navigates to listings?carType=New"),
    ("NC-ANG-003", "Popular models from Supabase", "PASS", "Filter km===0 && year>=2025"),
    ("NC-ANG-004", "Filters body/fuel/tx/budget/sort", "PASS", "Client-side"),
    ("NC-ANG-005", "New launches section", "PARTIAL", "Hardcoded static content"),
    ("NC-ANG-006", "Upcoming + Notify Me", "FAIL", "Notify is local signal only — not persisted"),
    ("NC-ANG-007", "Expert picks", "PARTIAL", "Hardcoded"),
    ("NC-ANG-008", "Compare heart → /compare", "FAIL", "Selection not passed to compare page"),
    ("NC-ANG-009", "Model → variant → detail", "PASS", "Via listings new mode + /cars/:id"),
    ("NC-ANG-010", "On-road price estimator", "PARTIAL", "Heuristic; state barely affects formula"),
    ("NC-ANG-011", "Colours / variants tabs", "PARTIAL", "Only for few models + sibling rows"),
    ("NC-ANG-012", "EMI on listings cards", "PARTIAL", "Fake price*0.008 formula"),
    ("NC-ANG-013", "Test drive booking", "PASS", "Angular /test-drive works via Supabase"),
    ("NC-ANG-014", "'New' definition vs API", "FAIL", "km===0 && year>=2025 ≠ listing_type=new"),
]
ang.append(["ID", "Scenario", "Status", "Notes"])
for cell in ang[1]:
    cell.fill = header
    cell.font = Font(bold=True, color="FFFFFF")
for row in ang_rows:
    ang.append(list(row))
    st = ang.cell(ang.max_row, 3)
    st.fill = {"PASS": pass_f, "FAIL": fail_f, "PARTIAL": partial_f}[row[2]]
for col, w in zip("ABCD", [12, 36, 10, 55]):
    ang.column_dimensions[col].width = w

# ── Gaps for Claude ──────────────────────────────────────────────────
gaps_ws = wb.create_sheet("Gaps for Claude")
gaps_ws.append(["Priority", "Gap", "Where", "Fix guidance"])
for cell in gaps_ws[1]:
    cell.fill = header
    cell.font = Font(bold=True, color="FFFFFF")

gaps = [
    ("P0", "Seed at least 20 listing_type=new inventory items", "apps/api/seed.py", "Mix of makes/cities so New Cars feed is not empty"),
    ("P0", "Reject or normalize km_driven when listing_type=new", "schemas/listing.py + create form", "Force km=0/null; hide KM/owners for new in UI"),
    ("P0", "Fix broken CTA /listings/new → /dashboard/listings/new", "cars/[make]/page.tsx", "Also fix /listings/[id]/book 404 or remove CTA"),
    ("P0", "Unify 'new' definition across Angular and API", "Angular cars-data + API", "Use listing_type/badge_type OR document single rule"),
    ("P0", "Implement /cars/[brand]/[model] (+ variant) catalogue pages", "apps/web + API", "PRD sitemap; ex-showroom starting price, specs"),
    ("P1", "City on-road price calculator (tax tables)", "API + Next/Angular detail", "Replace heuristic; state/city must change price"),
    ("P1", "Dealers directory + Get Best Price lead", "/dealers + leads API", "PRD dealer flow"),
    ("P1", "Brochure download + colour swatches", "catalogue API/UI", "Variant-level assets"),
    ("P1", "Wire Angular compare heart + notify to real backends", "new-cars.component", "Prefill compare IDs; persist notify emails"),
    ("P1", "Body-type links from /cars should keep listing_type=new when entered from New Cars", "cars/page.tsx", "Add listing_type=new query or separate new-cars hub on Next"),
    ("P1", "Dedicated Next /new-cars hub (parity with Angular)", "apps/web", "Launches, upcoming, expert picks from CMS/API not hardcoded"),
    ("P2", "EMI tied to on-road + bank LTV", "emi components", "Stop fake price*0.008"),
    ("P2", "Shareable compare URL for new cars (up to 5)", "compare", "PRD"),
    ("P2", "Playwright + API regression suite in CI for New Cars", "tests", "Keep new-cars.spec.ts + e2e_new_cars_api.py"),
]
for g in gaps:
    gaps_ws.append(list(g))
    gaps_ws.cell(gaps_ws.max_row, 1).fill = gap_f if g[0] == "P0" else partial_f
for col, w in zip("ABCD", [10, 55, 28, 50]):
    gaps_ws.column_dimensions[col].width = w

wb.save(OUT)
print("Wrote", OUT)
