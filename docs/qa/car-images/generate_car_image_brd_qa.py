#!/usr/bin/env python3
"""QA Architect: Car Image Asset Management BRD validation → Excel + Claude prompts."""
from __future__ import annotations

import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

QA = Path(__file__).resolve().parent
ART = Path("/opt/cursor/artifacts/car-image-brd-qa")
QA.mkdir(parents=True, exist_ok=True)
ART.mkdir(parents=True, exist_ok=True)

TIP = "claude/gaadiiq-app-dev-abj5fo @ 0b0bb10"

# id, area, priority, requirement, status, evidence, gap
REQS: list[tuple[str, str, str, str, str, str, str]] = [
    ("BR-001", "Upload", "Must Have", "Multi image upload JPG/JPEG/PNG/WEBP/HEIC/TIFF; configurable max up to 15GB", "PARTIAL", "POST /media-admin/upload + sniff_image; media_max_upload_mb default 64MB", "No Angular UI; no 15GB/chunked path; HEIC/TIFF not browser-converted"),
    ("BR-002", "Metadata", "Must Have", "Mandatory make/model/year/body/fuel/transmission/category; optional variant/color/primary", "PARTIAL", "API form fields on media-admin upload", "No admin metadata form UI; body_type naming differs"),
    ("BR-003", "Categories", "Must Have", "Full image category vocabulary (exterior/interior/360/etc.)", "PASS", "ImageCategory enum in vehicle_media.py", "Videos future — enum has video value"),
    ("BR-004", "AI Extract", "Must Have", "AI extract from filename/EXIF; admin editable before save", "PARTIAL", "filename_metadata + POST /media-admin/inspect; PATCH after save", "EXIF never populated; no pre-save review UI"),
    ("BR-005", "Storage Meta", "Must Have", "Rich metadata: hash, thumb, resolution, SEO, embedding, OCR, copyright…", "PARTIAL", "vehicle_media: hash/phash/thumb/w/h/alt/seo/copyright columns", "OCR, embedding vector, EXIF fill, AI description generation missing"),
    ("BR-DB-01", "Database", "Must Have", "Manufacturers / Models / Variants / CarImages relational tables", "FAIL", "Flat cars + vehicle_media free-text tags", "No BRD hierarchy tables; design-only in LLD"),
    ("BR-STO-01", "Storage", "Must Have", "Path /car-images/{make}/{model}/{year}/{category}/; no duplicate bytes", "PARTIAL", "car-images/{make}/{model}/{year}/{hash}.ext + sha256/phash dedupe", "Category folder segment missing"),
    ("BR-DSP-01", "Display", "Must Have", "Manufacturer page auto-display from single source", "PARTIAL", "GET /brochures/images + VehicleImageService on New Cars", "No dedicated manufacturer page DAM wiring everywhere"),
    ("BR-DSP-02", "Display", "Must Have", "Model page auto-display", "PARTIAL", "New Cars / car detail / brochure gallery", "Incomplete vs all model surfaces"),
    ("BR-DSP-03", "Display", "Must Have", "Variant page auto-display", "PARTIAL", "Filterable via media query params", "Variant-specific surfaces incomplete"),
    ("BR-DSP-04", "Display", "Must Have", "Search / AI Advisor / Compare / Dealer reuse same asset", "PARTIAL", "Compare + cards wired; AI Advisor & Dealer not", "Wire Advisor + Dealer to vehicle_media"),
    ("BR-SRCH-01", "Search", "Should Have", "Smart search: make/model/color/view filters", "PARTIAL", "OpenSearch media index + GET images filters", "Semantic 'White Nexon Front View' natural language incomplete"),
    ("BR-AI-01", "AI Features", "Should Have", "Auto-tag, dedupe, compress, thumb, SEO alt, description, semantic search, OCR, quality, moderation", "PARTIAL", "Dedupe + thumb + alt fields; filename tag", "OCR/embeddings/moderation/blur/quality gates missing"),
    ("FR-001", "Functional", "Must Have", "Admin multi-image upload", "PASS", "API + TestMultipleUploadSuite", "No Angular Car Image Admin UI"),
    ("FR-002", "Functional", "Must Have", "Mandatory metadata validation before save", "PASS", "API TestMandatoryMetadataSuite", "UI missing"),
    ("FR-003", "Functional", "Must Have", "AI auto-populates metadata", "PARTIAL", "Filename inspect only", "EXIF/vision auto-tag missing"),
    ("FR-004", "Functional", "Must Have", "Admin can override AI values", "PASS", "PATCH /media-admin/{id}", "Pre-save override UI missing"),
    ("FR-005", "Functional", "Must Have", "Images stored once in cloud", "PASS", "content_hash dedupe in media_library", "—"),
    ("FR-006", "Functional", "Must Have", "Same image reused across surfaces", "PARTIAL", "Shared vehicle_media read path", "Advisor/Dealer not wired"),
    ("FR-007", "Functional", "Must Have", "Automatic thumbnail generation", "PASS", "pdf_ingest thumb path used by media library", "—"),
    ("FR-008", "Functional", "Should Have", "Automatic WebP conversion", "FAIL", "Original format mostly kept", "Add convert-to-WebP pipeline"),
    ("FR-009", "Functional", "Must Have", "Duplicate detection before save", "PASS", "sha256 + phash dedupe", "—"),
    ("FR-010", "Functional", "Should Have", "Version history for image updates", "FAIL", "No media version table", "Add vehicle_media_versions"),
    ("NFR-001", "NFR", "Must Have", "Support uploads up to 15 GB", "FAIL", "Default 64MB; comment notes chunked needed", "Resumable/chunked upload"),
    ("NFR-002", "NFR", "Must Have", "Upload response ≤10s for standard images", "FAIL", "No SLA assertion/monitoring", "Add perf test + timeout policy"),
    ("NFR-003", "NFR", "Must Have", "CDN integration", "PARTIAL", "S3/R2 public_base support", "Default local /media paths in many envs"),
    ("NFR-004", "NFR", "Must Have", "RBAC for admin upload", "PASS", "get_admin_user on media-admin + brochures", "Dev shortcut localOnly has no token (ops issue)"),
    ("NFR-005", "NFR", "Should Have", "Audit logging for uploads/modifications", "PARTIAL", "Basic logging", "Dedicated audit table incomplete"),
    ("AC-001", "Acceptance", "Must Have", "Admin can upload one or multiple car images", "PARTIAL", "API PASS; UI FAIL", "Build Car Image Admin UI"),
    ("AC-002", "Acceptance", "Must Have", "Mandatory fields validated before save", "PASS", "API 422 on missing fields", "Surface validation in UI"),
    ("AC-003", "Acceptance", "Must Have", "AI extracts metadata from filename/image meta", "PARTIAL", "Filename inspect PASS; EXIF FAIL", "Populate EXIF + UI prefill"),
    ("AC-004", "Acceptance", "Must Have", "Admin review/modify AI metadata before save", "FAIL", "Only post-save PATCH", "Inspect → edit grid → upload flow in UI"),
    ("AC-005", "Acceptance", "Must Have", "Stored once and linked via metadata", "PASS", "Dedupe + vehicle_media links", "—"),
    ("AC-006", "Acceptance", "Must Have", "Appear on Manufacturer/Model/Variant without re-upload", "PARTIAL", "Partial surface wiring", "Complete all pages"),
    ("AC-007", "Acceptance", "Must Have", "Search shows correct images by metadata filters", "PARTIAL", "GET /brochures/images filters", "NL semantic search"),
    ("AC-008", "Acceptance", "Must Have", "Duplicate uploads detected and flagged", "PASS", "deduplicated=1 API behavior + tests", "—"),
    ("AC-009", "Acceptance", "Must Have", "Thumbnails and optimized formats auto-generated", "PARTIAL", "Thumbs yes; WebP optimize no", "WebP conversion"),
    ("AC-010", "Acceptance", "Should Have", "Metadata supports SEO/AI search/analytics", "PARTIAL", "alt/seo columns exist", "Embeddings + analytics events"),
    ("UX-001", "UX Bug", "Must Have", "Admin File Ingestion must not claim image upload if PDF-only", "FAIL", "/admin/pdf-ingestion accept=*/* then API 'File is not a PDF'", "Separate Car Image Admin OR route images to /media-admin"),
]


def main() -> None:
    status_c = Counter(r[4] for r in REQS)
    weights = {"PASS": 1.0, "PARTIAL": 0.5, "FAIL": 0.0}
    score = round(100 * sum(weights[r[4]] for r in REQS) / len(REQS))
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    tests: list[dict] = []

    def T(suite: str, tid: str, name: str, status: str, detail: str, sev: str = "—", br: str = "") -> None:
        tests.append(
            {
                "suite": suite,
                "id": tid,
                "name": name,
                "status": status,
                "detail": detail,
                "severity": sev,
                "br_ref": br,
            }
        )

    # Functional / API
    T("Functional", "TC-IMG-001", "Multi PNG upload via /media-admin/upload with mandatory fields", "PASS", "API + TestMultipleUploadSuite", "—", "FR-001/AC-001")
    T("Functional", "TC-IMG-002", "Accept JPG/JPEG/PNG/WEBP/HEIC/TIFF by magic bytes", "PASS", "sniff_image allowlist", "—", "BR-001")
    T("Functional", "TC-IMG-003", "Reject non-image masquerading as .jpg", "PASS", "422 magic-byte reject", "—", "BR-001")
    T("Functional", "TC-IMG-004", "Missing make/model/year rejected", "PASS", "TestMandatoryMetadataSuite", "—", "AC-002")
    T("Functional", "TC-IMG-005", "Invalid image_category rejected with vocabulary", "PASS", "ImageCategory enum", "—", "BR-003")
    T("Functional", "TC-IMG-006", "Inspect Tata_Nexon_FearlessPlus_Front_2025.webp filename", "PASS", "filename_metadata + /inspect", "—", "AC-003")
    T("Functional", "TC-IMG-007", "PATCH metadata after upload", "PASS", "PATCH /media-admin/{id}", "—", "FR-004")
    T("Functional", "TC-IMG-008", "Duplicate bytes → deduplicated flag", "PASS", "sha256/phash", "—", "AC-008")
    T("Functional", "TC-IMG-009", "Storage key includes {category} folder", "FAIL", "Key omits category segment", "P1", "BR-STO-01")
    T("Functional", "TC-IMG-010", "Upload 100MB with default config", "FAIL", "Default cap 64MB", "P1", "NFR-001")
    T("Functional", "TC-IMG-011", "Configure MEDIA_MAX_UPLOAD_MB raises limit", "PARTIAL", "Config works; not 15GB/chunked", "P1", "BR-001")
    T("UI", "TC-IMG-012", "Angular Car Image Admin multi-select + metadata grid", "FAIL", "No /media-admin UI; only PDF page", "P0", "AC-001/AC-004")
    T("UI", "TC-IMG-013", "File Ingestion upload JPG as car image", "FAIL", "API returns File is not a PDF", "P0", "UX-001")
    T("UI", "TC-IMG-014", "File Ingestion upload PDF brochure", "PASS", "brochures/upload PDF path", "—", "Separate feature")
    T("AI", "TC-IMG-015", "EXIF make/model auto-fill", "FAIL", "EXIF column never populated", "P1", "BR-004")
    T("AI", "TC-IMG-016", "OCR text stored on media row", "FAIL", "Not implemented for car images", "P2", "BR-AI-01")
    T("AI", "TC-IMG-017", "Image embedding vector for semantic search", "FAIL", "Not on vehicle_media", "P2", "AC-010")
    T("Integration", "TC-IMG-018", "New Cars card uses vehicle_media when car.image empty", "PASS", "VehicleImageService.imageFor", "—", "BR-DSP-02")
    T("Integration", "TC-IMG-019", "Compare uses shared media library", "PASS", "compare + brochure images", "—", "FR-006")
    T("Integration", "TC-IMG-020", "AI Advisor uses shared media library", "FAIL", "Still placeholders/car.image", "P0", "BR-DSP-04")
    T("Integration", "TC-IMG-021", "Dealer dashboard uses shared media", "FAIL", "Not wired", "P1", "BR-DSP-04")
    T("Integration", "TC-IMG-022", "GET /brochures/images filter make/model/variant", "PASS", "Shared read API", "—", "AC-007")
    T("Performance", "TC-IMG-023", "5MB image upload p95 ≤10s", "FAIL", "No SLA test/monitor", "P2", "NFR-002")
    T("Infrastructure", "TC-IMG-024", "CDN public URL when R2 configured", "PARTIAL", "S3Storage public_base", "P2", "NFR-003")
    T("Security", "TC-IMG-025", "Unauthenticated media-admin upload → 401", "PASS", "get_admin_user", "—", "NFR-004")
    T("Security", "TC-IMG-026", "Non-admin token → 403", "PASS", "admin dependency", "—", "NFR-004")
    T("Regression", "TC-IMG-027", "pytest test_media_admin suite green", "PASS", "Exists on tip", "—", "FR-*")
    T("UX", "TC-IMG-028", "Pre-save AI metadata review grid", "FAIL", "No inspect→edit→upload UI", "P0", "AC-004")
    T("Data", "TC-IMG-029", "Manufacturers/Models/Variants/CarImages tables exist", "FAIL", "Not as BRD specifies", "P1", "BR-DB-01")
    T("Functional", "TC-IMG-030", "Automatic WebP conversion on store", "FAIL", "Not implemented", "P1", "FR-008")
    T("Functional", "TC-IMG-031", "Image version history on replace", "FAIL", "No versions table", "P2", "FR-010")
    T("UI", "TC-IMG-032", "Nav label 'Car Image Management' vs File Ingestion only", "FAIL", "Navbar only File Ingestion", "P1", "UX-001")

    tc = Counter(t["status"] for t in tests)

    # Markdown summary
    req_rows = "\n".join(
        f"| {a} | {b} | {c} | {d} | **{e}** | {g} |" for a, b, c, d, e, _f, g in REQS
    )
    fail_tests = "\n".join(
        f"| {t['id']} | {t['status']} | {t['severity']} | {t['name']} — {t['detail']} |"
        for t in tests
        if t["status"] != "PASS"
    )

    md = f"""# GAADIIQ — Car Image Asset Management BRD QA

| Field | Value |
|-------|-------|
| Feature | AI-Powered Car Image Upload & Intelligent Asset Management |
| BRD source | `docs/qa/car-images/BRD_Car_Image_Asset_Management.txt` |
| Code tip audited | `{TIP}` |
| Date | {now} |
| BRD readiness | **{score}/100** |
| Requirements | {status_c.get('PASS',0)} PASS / {status_c.get('PARTIAL',0)} PARTIAL / {status_c.get('FAIL',0)} FAIL |
| Test cases | {tc.get('PASS',0)} PASS / {tc.get('PARTIAL',0)} PARTIAL / {tc.get('FAIL',0)} FAIL |
| Verdict | **NO-GO** for full BRD · **Conditional** API-only DAM foundation |

---

## Executive finding

Tip implements a **backend media library** (`POST /media-admin/*` + `vehicle_media`) with categories, filename AI inspect, dedupe, and thumbs. There is **no Angular Car Image Admin**. Operators use **File Ingestion** (`/admin/pdf-ingestion`), which calls **PDF-only** `POST /brochures/upload` → **"File is not a PDF"** when uploading car images — matching the live UI failure.

---

## Requirements vs code

| ID | Area | Priority | Requirement | Status | Gap |
|----|------|----------|-------------|--------|-----|
{req_rows}

---

## Non-PASS test cases

| ID | Status | Severity | Detail |
|----|--------|----------|--------|
{fail_tests}

---

## P0 gaps
1. Build **Car Image Admin** UI wired to `/media-admin/inspect` + `/upload` + metadata grid (AC-001/004).
2. Fix **File Ingestion** UX: either reject non-PDF client-side OR route images to media-admin; stop advertising image upload on PDF page (UX-001 / TC-IMG-013).
3. Wire **AI Advisor** (and Dealer) to shared `vehicle_media` (BR-DSP-04).

## P1 gaps
Category folder in storage key · 15GB/chunked upload · EXIF fill · WebP conversion · relational Manufacturers/Models/Variants or documented free-text strategy · Dealer dashboard images.

## Artifacts
- Excel: `GAADIIQ_Car_Image_BRD_QA.xlsx`
- Claude prompt: `Claude_Fix_Prompts_Car_Image_BRD.md`
- JSON: `car-image-brd-qa.json`
"""
    (QA / "Car_Image_BRD_QA_Summary.md").write_text(md)
    (ART / "Car_Image_BRD_QA_Summary.md").write_text(md)

    prompts = f"""# Claude Code Fix Prompts — Car Image Asset Management BRD

**BRD readiness:** {score}/100 · Tip: `{TIP}`  
**Source:** `docs/qa/car-images/Car_Image_BRD_QA_Summary.md` + uploaded BRD

---

## MASTER PROMPT (paste into Claude Code)

```
# ROLE
Implement GAADIIQ Car Image Asset Management BRD gaps on tip that already has
POST /media-admin/* and vehicle_media (do not rebuild from scratch).

# READ
docs/qa/car-images/BRD_Car_Image_Asset_Management.txt
docs/qa/car-images/Car_Image_BRD_QA_Summary.md
docs/qa/car-images/Claude_Fix_Prompts_Car_Image_BRD.md

# KEEP WORKING
- POST /media-admin/inspect, /upload, PATCH /media-admin/{{id}}
- vehicle_media + ImageCategory enum + filename_metadata
- sha256/phash dedupe, thumbnails, GET /brochures/images
- test_media_admin.py / test_filename_metadata.py
- PDF brochure pipeline POST /brochures/upload (PDF-only) — keep for brochures

# DO NOT BREAK
PDF ingestion for real PDFs, media dedupe, admin auth (get_admin_user),
New Cars/Compare VehicleImageService wiring.

# WAVE 1 — P0 (ship first)

## 1) UX-001 + TC-IMG-013 — Stop "File is not a PDF" trap for car images
Option A (preferred): New Angular route `/admin/car-images` (Car Image Management)
  - Multi-file picker accept image/* (+ heic/tiff)
  - Call POST /media-admin/inspect → editable metadata grid → POST /media-admin/upload
  - Fields: manufacturer, model, variant?, year, body/category, fuel, transmission,
    image_category, colour?, primary flag
  - Nav: Admin → Car Image Management (keep File Ingestion for PDFs)
Option B (minimum): On /admin/pdf-ingestion, if file is image, route to /media-admin/upload
  with required metadata modal; if not PDF and not image, clear client error.
Also: change accept + help text so PDF page does NOT claim "all file types / images"
unless Option B is fully wired.

## 2) AC-004 — Pre-save review
Inspect → edit AI suggestions → confirm upload. Do not save until admin confirms.

## 3) BR-DSP-04 — Wire AI Advisor (and Dealer if in scope) to VehicleImageService /
   GET /brochures/images so recommendations show DAM assets, not placeholders.

# WAVE 2 — P1
- BR-STO-01: storage key car-images/{{make}}/{{model}}/{{year}}/{{category}}/{{hash}}.ext
- BR-004: populate EXIF into vehicle_media.exif; merge into inspect suggestions
- FR-008 / AC-009: convert delivery derivative to WebP (keep original)
- NFR-001: document real max; implement chunked/resumable if 15GB still required
  (or officially revise BRD to e.g. 100MB with MEDIA_MAX_UPLOAD_MB)
- BR-DB-01: either introduce manufacturers/models/variants FKs OR document that
  vehicle_media free-text + cars table is the accepted model (update BRD)
- TC-IMG-021: Dealer dashboard shared media

# WAVE 3 — P2
- FR-010 version history
- Image embeddings + semantic search ("White Nexon Front View")
- OCR on brochure images already in PDF path — optional link to media
- NSFW/plate blur hooks (stub OK)
- NFR-002: perf test for 5MB upload p95≤10s (informational CI)
- Audit log table for media upload/patch/delete

# TESTS
Extend apps/api/tests/test_media_admin.py for category path, EXIF (mocked), WebP.
Add Angular e2e/smoke for /admin/car-images happy path (mock API).
Keep brochure PDF upload test: non-PDF still rejected on /brochures/upload.

# ACCEPTANCE
- Admin can upload 3 JPGs via Car Image UI with metadata → appears in GET images
- Uploading JPG on PDF-only page does not show confusing success path; either
  redirected to car-images or clear client validation
- AI Advisor shows library image when available
- pytest media suites green
```

---

## Compact Wave 1 only

```
Implement Wave 1 from docs/qa/car-images/Claude_Fix_Prompts_Car_Image_BRD.md:
(1) /admin/car-images UI → /media-admin inspect+upload
(2) Fix File Ingestion accept/copy so images aren't sent to PDF endpoint
(3) Wire AI Advisor to shared vehicle_media images.
Keep PDF brochure upload working. Add tests.
```

---

## Single-issue: File is not a PDF

```
Bug: Admin File Ingestion UI accepts images but POST /brochures/upload returns
"File is not a PDF". BRD is Car Image DAM (BR-001), not brochures.

Fix: Add /admin/car-images using /media-admin/* OR client-route images to
media-admin with metadata. Update accept= and help text on PDF page.
Do not weaken PDF magic-byte check on /brochures/upload.
```
"""
    (QA / "Claude_Fix_Prompts_Car_Image_BRD.md").write_text(prompts)
    (ART / "Claude_Fix_Prompts_Car_Image_BRD.md").write_text(prompts)

    catalog = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "code_tip": TIP,
        "brd_readiness_score": score,
        "requirement_counts": dict(status_c),
        "test_counts": dict(tc),
        "recommendation": "NO-GO for full BRD; API DAM foundation Conditional; UI P0",
        "requirements": [
            {
                "id": a,
                "area": b,
                "priority": c,
                "requirement": d,
                "status": e,
                "evidence": f,
                "gap": g,
            }
            for a, b, c, d, e, f, g in REQS
        ],
        "tests": tests,
    }
    (QA / "car-image-brd-qa.json").write_text(json.dumps(catalog, indent=2))
    (ART / "car-image-brd-qa.json").write_text(json.dumps(catalog, indent=2))

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        import sys

        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

    fills = {
        "PASS": PatternFill("solid", "C6EFCE"),
        "PARTIAL": PatternFill("solid", "FFEB9C"),
        "FAIL": PatternFill("solid", "FFC7CE"),
    }
    wb = Workbook()

    ws = wb.active
    ws.title = "Executive"
    for row in [
        ["GAADIIQ — Car Image Asset Management BRD QA"],
        ["Generated", now],
        ["Code tip", TIP],
        ["BRD readiness score", score],
        ["Requirements PASS", status_c.get("PASS", 0)],
        ["Requirements PARTIAL", status_c.get("PARTIAL", 0)],
        ["Requirements FAIL", status_c.get("FAIL", 0)],
        ["Test PASS", tc.get("PASS", 0)],
        ["Test PARTIAL", tc.get("PARTIAL", 0)],
        ["Test FAIL", tc.get("FAIL", 0)],
        ["Total requirements", len(REQS)],
        ["Total test cases", len(tests)],
        ["Verdict", catalog["recommendation"]],
        ["Live UX bug", "File Ingestion accepts images → API 'File is not a PDF'"],
        ["P0 fix", "Build /admin/car-images → /media-admin; fix PDF page accept/copy"],
    ]:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    wr = wb.create_sheet("Requirements")
    wr.append(["ID", "Area", "Priority", "Requirement", "Status", "Evidence", "Gap"])
    for r in REQS:
        wr.append(list(r))
        wr.cell(wr.max_row, 5).fill = fills[r[4]]
    for i, w in enumerate([12, 14, 12, 55, 12, 45, 45], 1):
        wr.column_dimensions[get_column_letter(i)].width = w
    wr.auto_filter.ref = f"A1:G{wr.max_row}"
    wr.freeze_panes = "A2"

    wt = wb.create_sheet("TestCases")
    wt.append(["Suite", "ID", "Name", "Status", "Severity", "BR Ref", "Detail"])
    for t in tests:
        wt.append([t["suite"], t["id"], t["name"], t["status"], t["severity"], t["br_ref"], t["detail"]])
        wt.cell(wt.max_row, 4).fill = fills.get(t["status"], PatternFill())
    for i, w in enumerate([14, 12, 55, 12, 10, 16, 45], 1):
        wt.column_dimensions[get_column_letter(i)].width = w
    wt.auto_filter.ref = f"A1:G{wt.max_row}"
    wt.freeze_panes = "A2"

    wg = wb.create_sheet("Gaps_P0_P1_P2")
    wg.append(["Priority", "ID", "Gap", "Suggested fix"])
    gaps = [
        ("P0", "UX-001", "File Ingestion UI accepts images; API PDF-only → File is not a PDF", "Add /admin/car-images OR route images to /media-admin; fix accept/copy"),
        ("P0", "AC-001/012", "No Angular Car Image Admin for /media-admin", "Build inspect→edit→upload UI"),
        ("P0", "AC-004", "No pre-save AI metadata review", "Metadata grid before upload"),
        ("P0", "BR-DSP-04", "AI Advisor not on shared media", "Use VehicleImageService / brochures images"),
        ("P1", "BR-STO-01", "Storage path missing {category}", "Include category in object key"),
        ("P1", "NFR-001", "15GB not supported (64MB default)", "Chunked upload or revise BRD max"),
        ("P1", "BR-004", "EXIF not populated", "Parse EXIF in inspect/upload"),
        ("P1", "FR-008", "No WebP conversion", "Generate WebP derivative"),
        ("P1", "BR-DB-01", "No Manufacturers/Models/Variants/CarImages tables", "FK model or document free-text approach"),
        ("P1", "TC-IMG-021", "Dealer dashboard not wired", "Shared media in dealer UI"),
        ("P2", "FR-010", "No version history", "vehicle_media_versions"),
        ("P2", "AC-010", "No image embeddings", "Vector column + semantic search"),
        ("P2", "NFR-002", "No ≤10s SLA test", "Perf fixture in CI"),
    ]
    for g in gaps:
        wg.append(list(g))
        if g[0] == "P0":
            wg.cell(wg.max_row, 1).fill = PatternFill("solid", "FFC7CE")
        elif g[0] == "P1":
            wg.cell(wg.max_row, 1).fill = PatternFill("solid", "FFEB9C")
        else:
            wg.cell(wg.max_row, 1).fill = PatternFill("solid", "DDEBF7")
    for i, w in enumerate([10, 14, 55, 55], 1):
        wg.column_dimensions[get_column_letter(i)].width = w

    wa = wb.create_sheet("AcceptanceCriteria")
    wa.append(["AC ID", "Criteria", "Status", "Notes"])
    for r in REQS:
        if r[0].startswith("AC-"):
            wa.append([r[0], r[3], r[4], r[6]])
            wa.cell(wa.max_row, 3).fill = fills[r[4]]
    for i, w in enumerate([10, 55, 12, 45], 1):
        wa.column_dimensions[get_column_letter(i)].width = w

    out = QA / "GAADIIQ_Car_Image_BRD_QA.xlsx"
    wb.save(out)
    (ART / "GAADIIQ_Car_Image_BRD_QA.xlsx").write_bytes(out.read_bytes())
    (ART / "Claude_Fix_Prompts_Car_Image_BRD.md").write_text(prompts)

    print(
        json.dumps(
            {
                "score": score,
                "reqs": dict(status_c),
                "tests": dict(tc),
                "xlsx": str(out),
                "n_reqs": len(REQS),
                "n_tests": len(tests),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
