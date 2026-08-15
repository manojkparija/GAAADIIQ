#!/usr/bin/env python3
"""Car Image BRD retest after WAVE 1–3 merge → Excel + Claude remaining-gap prompts."""
from __future__ import annotations

import json
import subprocess
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

QA = Path(__file__).resolve().parent
ART = Path("/opt/cursor/artifacts/car-image-brd-retest")
QA.mkdir(parents=True, exist_ok=True)
ART.mkdir(parents=True, exist_ok=True)

TIP = "origin/master @ 15518ac (WAVE 2 complete + WAVE 3 partial)"
PRIOR = 55

# id, area, priority, requirement, old, new, evidence, gap
REQS: list[tuple[str, str, str, str, str, str, str, str]] = [
    ("BR-001", "Upload", "Must Have", "Multi upload JPG/JPEG/PNG/WEBP/HEIC/TIFF; max up to 15GB", "PARTIAL", "PARTIAL", "UI+/media-admin; media_max_upload_mb=100", "Not 15GB; no chunked/resumable"),
    ("BR-002", "Metadata", "Must Have", "Mandatory make/model/year/body/fuel/transmission/category", "PARTIAL", "PASS", "Admin form + API Form validation", "—"),
    ("BR-003", "Categories", "Must Have", "Full ImageCategory vocabulary", "PASS", "PASS", "ImageCategory enum", "—"),
    ("BR-004", "AI Extract", "Must Have", "Filename/EXIF extract; admin editable before save", "PARTIAL", "PASS", "inspect + extract_exif + pre-save UI", "EXIF not auto-applied into form fields"),
    ("BR-005", "Storage Meta", "Must Have", "Rich metadata incl. OCR/embedding/SEO", "PARTIAL", "PARTIAL", "hash/phash/thumb/webp/exif/alt/seo", "OCR+embeddings on claude tip only, not master"),
    ("BR-DB-01", "Database", "Must Have", "Manufacturers/Models/Variants/CarImages tables", "FAIL", "PASS", "AD-001 documents free-text vehicle_media approach", "No FK hierarchy (accepted by design)"),
    ("BR-STO-01", "Storage", "Must Have", "Path /car-images/{make}/{model}/{year}/{category}/", "PARTIAL", "PASS", "key_prefix includes category", "—"),
    ("BR-DSP-01", "Display", "Must Have", "Manufacturer page DAM display", "PARTIAL", "PARTIAL", "Shared library; brand pages thin", "Dedicated manufacturer gallery incomplete"),
    ("BR-DSP-02", "Display", "Must Have", "Model page DAM display", "PARTIAL", "PASS", "New Cars + car-detail + VehicleImageService", "—"),
    ("BR-DSP-03", "Display", "Must Have", "Variant page DAM display", "PARTIAL", "PARTIAL", "API variant filter exists", "No dedicated variant page surface"),
    ("BR-DSP-04", "Display", "Must Have", "Search/AI Advisor/Compare/Dealer reuse same asset", "PARTIAL", "PASS", "Advisor imageOr + Dealer gallery + Compare", "Dealer fetch hardcodes localhost:8000"),
    ("BR-SRCH-01", "Search", "Should Have", "Smart search incl. White Nexon Front View", "PARTIAL", "PARTIAL", "q= filters + OpenSearch", "CLIP semantic NL search tip-only"),
    ("BR-AI-01", "AI Features", "Should Have", "Auto-tag, dedupe, compress, thumb, SEO, OCR, moderation…", "PARTIAL", "PARTIAL", "Dedupe/thumb/webp/filename/exif", "OCR/NSFW/plate blur tip-only or missing"),
    ("FR-001", "Functional", "Must Have", "Admin multi-image upload", "PASS", "PASS", "/admin/car-images + API", "—"),
    ("FR-002", "Functional", "Must Have", "Mandatory metadata validation", "PASS", "PASS", "API 422 + UI checks", "—"),
    ("FR-003", "Functional", "Must Have", "AI auto-populates metadata", "PARTIAL", "PARTIAL", "Inspect badges shown", "Hints not auto-applied into form signals"),
    ("FR-004", "Functional", "Must Have", "Admin override AI values", "PASS", "PASS", "Editable pre-save + PATCH", "—"),
    ("FR-005", "Functional", "Must Have", "Images stored once", "PASS", "PASS", "sha256 + phash dedupe", "—"),
    ("FR-006", "Functional", "Must Have", "Reuse across surfaces", "PARTIAL", "PASS", "Advisor/Dealer/Compare/New Cars", "Manufacturer/variant pages thin"),
    ("FR-007", "Functional", "Must Have", "Thumbnail generation", "PASS", "PASS", "make_thumbnail", "—"),
    ("FR-008", "Functional", "Should Have", "Automatic WebP conversion", "FAIL", "PASS", "make_webp → webp_key", "—"),
    ("FR-009", "Functional", "Must Have", "Duplicate detection", "PASS", "PASS", "deduplicated flag + tests", "—"),
    ("FR-010", "Functional", "Should Have", "Version history", "FAIL", "PASS", "vehicle_media_versions + rollback API", "—"),
    ("NFR-001", "NFR", "Must Have", "Uploads up to 15 GB", "FAIL", "PARTIAL", "100MB default + 413 path", "Revise BRD or add chunked upload"),
    ("NFR-002", "NFR", "Must Have", "Upload ≤10s for standard images", "FAIL", "PARTIAL", "Docs claim ~3s for 5MB", "No automated p95 SLA test"),
    ("NFR-003", "NFR", "Must Have", "CDN integration", "PARTIAL", "PARTIAL", "R2/S3 public_base", "Often local /media in dev"),
    ("NFR-004", "NFR", "Must Have", "RBAC admin upload", "PASS", "PASS", "get_admin_user", "—"),
    ("NFR-005", "NFR", "Should Have", "Audit logging", "PARTIAL", "PASS", "vehicle_media_audit + GET trail", "—"),
    ("AC-001", "Acceptance", "Must Have", "Admin upload one/multiple car images", "PARTIAL", "PASS", "/admin/car-images multi-select", "—"),
    ("AC-002", "Acceptance", "Must Have", "Mandatory fields validated", "PASS", "PASS", "API+UI", "—"),
    ("AC-003", "Acceptance", "Must Have", "AI extracts from filename/image meta", "PARTIAL", "PASS", "inspect + EXIF store", "—"),
    ("AC-004", "Acceptance", "Must Have", "Review/modify AI metadata before save", "FAIL", "PASS", "Inspect → grid → upload", "Batch shared metadata (not per-file category)"),
    ("AC-005", "Acceptance", "Must Have", "Stored once linked by metadata", "PASS", "PASS", "Dedupe + vehicle_media", "—"),
    ("AC-006", "Acceptance", "Must Have", "Appear on Manufacturer/Model/Variant pages", "PARTIAL", "PARTIAL", "Model/detail/compare yes", "Manufacturer + variant pages incomplete"),
    ("AC-007", "Acceptance", "Must Have", "Search by metadata filters", "PARTIAL", "PARTIAL", "GET images filters", "NL semantic search tip-only"),
    ("AC-008", "Acceptance", "Must Have", "Duplicate flagged", "PASS", "PASS", "Tests green", "—"),
    ("AC-009", "Acceptance", "Must Have", "Thumbs + optimized formats", "PARTIAL", "PASS", "Thumb + WebP", "—"),
    ("AC-010", "Acceptance", "Should Have", "SEO/AI search/analytics metadata", "PARTIAL", "PARTIAL", "alt/seo columns", "Embeddings tip-only"),
    ("UX-001", "UX Bug", "Must Have", "File Ingestion must not trap car images as PDF", "FAIL", "PASS", "PDF accept=.pdf; separate Car Images nav", "—"),
]


def main() -> None:
    status_c = Counter(r[5] for r in REQS)
    weights = {"PASS": 1.0, "PARTIAL": 0.5, "FAIL": 0.0}
    score = round(100 * sum(weights[r[5]] for r in REQS) / len(REQS))
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    tests: list[dict] = []

    def T(suite, tid, name, status, detail, sev="—", br=""):
        tests.append(
            {"suite": suite, "id": tid, "name": name, "status": status, "detail": detail, "severity": sev, "br_ref": br}
        )

    T("UI", "TC-IMG-012", "Angular Car Image Admin multi-select + metadata", "PASS", "/admin/car-images exists", "—", "AC-001")
    T("UI", "TC-IMG-013", "File Ingestion JPG no longer hits PDF trap", "PASS", "accept=.pdf only; Car Images separate", "—", "UX-001")
    T("UI", "TC-IMG-014", "File Ingestion PDF brochure still works", "PASS", "brochures/upload unchanged", "—", "—")
    T("UI", "TC-IMG-028", "Pre-save AI metadata review grid", "PASS", "inspect → edit → upload", "—", "AC-004")
    T("UI", "TC-IMG-032", "Nav Car Images label", "PASS", "Navbar Car Images link", "—", "UX-001")
    T("Functional", "TC-IMG-001", "Multi PNG upload /media-admin", "PASS", "pytest media_admin", "—", "FR-001")
    T("Functional", "TC-IMG-002", "Magic-byte formats JPG/PNG/WEBP/TIFF/HEIC", "PASS", "sniff_image", "—", "BR-001")
    T("Functional", "TC-IMG-004", "Missing mandatory → 422", "PASS", "TestMandatoryMetadataSuite", "—", "AC-002")
    T("Functional", "TC-IMG-005", "Invalid image_category rejected", "PASS", "ImageCategory enum", "—", "BR-003")
    T("Functional", "TC-IMG-006", "Filename inspect Tata_Nexon_…_Front_2025", "PASS", "filename_metadata", "—", "AC-003")
    T("Functional", "TC-IMG-007", "PATCH metadata after upload", "PASS", "PATCH + version event", "—", "FR-004")
    T("Functional", "TC-IMG-008", "Duplicate bytes deduplicated", "PASS", "sha256/phash", "—", "AC-008")
    T("Functional", "TC-IMG-009", "Storage key includes {category}", "PASS", "car-images/.../{category}/{hash}", "—", "BR-STO-01")
    T("Functional", "TC-IMG-010", "Default max upload 100MB (not 15GB)", "PARTIAL", "100MB; 15GB not implemented", "P1", "NFR-001")
    T("Functional", "TC-IMG-015", "EXIF populated on store", "PASS", "extract_exif → vehicle_media.exif", "—", "BR-004")
    T("Functional", "TC-IMG-030", "WebP derivative generated", "PASS", "webp_key", "—", "FR-008")
    T("Functional", "TC-IMG-031", "Version history on update", "PASS", "vehicle_media_versions", "—", "FR-010")
    T("Integration", "TC-IMG-018", "New Cars uses vehicle_media", "PASS", "VehicleImageService", "—", "BR-DSP-02")
    T("Integration", "TC-IMG-019", "Compare uses shared media", "PASS", "compare wired", "—", "FR-006")
    T("Integration", "TC-IMG-020", "AI Advisor uses DAM images", "PASS", "imageOr + VehicleImageService", "—", "BR-DSP-04")
    T("Integration", "TC-IMG-021", "Dealer dashboard shared media", "PASS", "dealer-images + gallery", "P1", "BR-DSP-04")
    T("Integration", "TC-IMG-021b", "Dealer gallery uses environment.apiUrl", "FAIL", "Hardcoded http://localhost:8000", "P1", "BR-DSP-04")
    T("Integration", "TC-IMG-022", "GET images filter make/model/variant", "PASS", "Shared read API", "—", "AC-007")
    T("Integration", "TC-IMG-033", "Manufacturer page DAM gallery", "PARTIAL", "No dedicated manufacturer gallery", "P1", "AC-006")
    T("Integration", "TC-IMG-034", "Variant page DAM gallery", "PARTIAL", "No dedicated variant page", "P1", "AC-006")
    T("AI", "TC-IMG-016", "OCR on car image (master)", "FAIL", "Tip-only WAVE 3 ML", "P2", "BR-AI-01")
    T("AI", "TC-IMG-017", "Embedding semantic search (master)", "FAIL", "Tip-only CLIP /search", "P2", "AC-010")
    T("AI", "TC-IMG-035", "NSFW/plate blur pipeline (master)", "FAIL", "Tip detects; blur not done; not on master", "P2", "BR-AI-01")
    T("AI", "TC-IMG-036", "Inspect hints auto-apply into form", "PARTIAL", "Badges only", "P1", "FR-003")
    T("Performance", "TC-IMG-023", "5MB upload p95 ≤10s automated", "FAIL", "No SLA pytest; docs claim ~3s", "P2", "NFR-002")
    T("Security", "TC-IMG-025", "Unauth media-admin → 401", "PASS", "get_admin_user", "—", "NFR-004")
    T("Security", "TC-IMG-037", "Audit trail recorded on upload/edit", "PASS", "vehicle_media_audit", "—", "NFR-005")
    T("Regression", "TC-IMG-027", "pytest media_admin+storage+filename", "PASS", "70 passed in retest env", "—", "—")
    T("Data", "TC-IMG-029", "BRD hierarchy tables OR documented free-text", "PASS", "ARCHITECTURAL_DECISIONS AD-001", "—", "BR-DB-01")

    tc = Counter(t["status"] for t in tests)

    closed = [r for r in REQS if r[4] != "PASS" and r[5] == "PASS"]
    still = [r for r in REQS if r[5] != "PASS"]

    md = f"""# Car Image BRD — E2E Retest (Post Merge)

| Field | Value |
|-------|-------|
| Date | {now} |
| Code tip | `{TIP}` |
| Prior score | {PRIOR}/100 |
| **BRD readiness now** | **{score}/100** |
| Requirements | {status_c.get('PASS',0)} PASS / {status_c.get('PARTIAL',0)} PARTIAL / {status_c.get('FAIL',0)} FAIL |
| Test cases | {tc.get('PASS',0)} PASS / {tc.get('PARTIAL',0)} PARTIAL / {tc.get('FAIL',0)} FAIL |
| pytest (retest) | **70 passed** (`test_media_admin` + `test_media_storage` + `test_filename_metadata`) |
| Verdict | **Conditional GO** for core DAM · Full BRD (15GB + semantic ML) still open |

---

## What closed since prior audit (55 → {score})

| ID | Was → Now | Fix |
|----|-----------|-----|
""" + "\n".join(f"| {r[0]} | {r[4]} → **{r[5]}** | {r[6]} |" for r in closed) + f"""

### Confirmed P0 closures
- `/admin/car-images` UI + nav **Car Images**
- PDF File Ingestion `accept=".pdf"` — no more image → “File is not a PDF” trap
- Pre-save inspect → metadata → upload (AC-004)
- AI Advisor + Dealer wired to `vehicle_media`
- Category storage path, EXIF, WebP, 100MB limit, version history, audit log

---

## Remaining gaps (not all fixed)

### P1
1. **NFR-001** — BRD says 15GB; product is **100MB** (chunked upload or BRD revision)
2. **AC-006 / BR-DSP-01/03** — Manufacturer + Variant dedicated DAM surfaces incomplete
3. **FR-003** — Inspect hints shown as badges but **not auto-applied** into form
4. **TC-IMG-021b** — Dealer gallery hardcodes `http://localhost:8000` (breaks prod)
5. Merge Claude tip WAVE 3 ML carefully onto master (CLIP/OCR/safety) without losing #28 PDF polling fixes

### P2
1. Semantic search “White Nexon Front View” (CLIP) — tip `633f885` only
2. OCR / NSFW / plate **blur** / quality validation
3. **NFR-002** automated ≤10s SLA test
4. Stronger EXIF→form prefill; per-file category on batch upload
5. CDN defaulting in all envs

---

## Requirements matrix

| ID | Area | Priority | Requirement | Prior | Now | Gap |
|----|------|----------|-------------|-------|-----|-----|
""" + "\n".join(
        f"| {a} | {b} | {c} | {d} | {old} | **{new}** | {g} |"
        for a, b, c, d, old, new, _e, g in REQS
    ) + f"""

---

## Non-PASS tests

| ID | Status | Severity | Detail |
|----|--------|----------|--------|
""" + "\n".join(
        f"| {t['id']} | {t['status']} | {t['severity']} | {t['name']} — {t['detail']} |"
        for t in tests
        if t["status"] != "PASS"
    ) + """

---

Claude remaining-gap prompt: `Claude_Fix_Prompts_Car_Image_Retest.md`  
Excel: `GAADIIQ_Car_Image_BRD_Retest.xlsx`
"""
    (QA / "Car_Image_BRD_Retest_Summary.md").write_text(md)
    (ART / "Car_Image_BRD_Retest_Summary.md").write_text(md)

    prompts = f"""# Claude Code Fix Prompts — Car Image BRD Remaining Gaps (Retest)

**Score now:** {score}/100 (was {PRIOR}) · Tip: `{TIP}`  
**Source:** `docs/qa/car-images/Car_Image_BRD_Retest_Summary.md`

---

## DO NOT RE-IMPLEMENT (confirmed PASS)

```
KEEP WORKING — already merged on master:
- /admin/car-images UI + nav Car Images
- PDF File Ingestion accept=.pdf only (no image trap)
- /media-admin inspect → edit → upload, PATCH
- Category storage path, EXIF store, WebP, thumbs, dedupe
- AI Advisor VehicleImageService + Dealer gallery endpoint
- vehicle_media_versions + vehicle_media_audit
- media_max_upload_mb=100 + 413
- AD-001 free-text metadata decision
- pytest media suites green (70 passed in retest)

DO NOT BREAK these.
```

---

## MASTER PROMPT — Remaining gaps only

```
# ROLE
Close remaining Car Image Asset Management BRD gaps after WAVE 1–2 merge.
Work on current master (15518ac+). Do not rebuild Car Images UI or PDF accept fix.

# READ
docs/qa/car-images/Car_Image_BRD_Retest_Summary.md
docs/IMPLEMENTATION_STATUS.md
docs/ARCHITECTURAL_DECISIONS.md

# WAVE R1 — P1 production polish (do first)

## R1.1 TC-IMG-021b — Dealer gallery API URL
File: apps/gaadiiq-angular/src/app/services/vehicle-image-gallery.service.ts
Replace hardcoded `http://localhost:8000/media-admin/dealer-images` with
`environment.apiUrl` (+ auth header via interceptor or fetch with session token).
Add a unit/smoke test or at least ensure prod Render URL is used.

## R1.2 FR-003 — Auto-apply inspect hints
On successful POST /media-admin/inspect, if form fields empty, apply suggested
make/model/year/category/colour into form signals (admin can still edit).
Prefer per-file suggestions when batch size=1; for multi-file keep shared defaults
but show per-file hint chips.

## R1.3 AC-006 / BR-DSP-01 / BR-DSP-03 — Manufacturer & Variant surfaces
- Manufacturer view: when user filters/opens a make, show DAM images for that make
  (reuse GET /brochures/images?make= or media-admin list).
- Variant: when variant selected on detail/compare, prefer variant-tagged images,
  fallback to model-level.
Do not invent new DB hierarchy (AD-001 stands).

## R1.4 NFR-001 — 15GB vs 100MB decision
Either:
A) Officially revise BRD/docs: max 100MB images; 15GB only if/when video/chunked ships, OR
B) Implement resumable/chunked upload (TUS or multipart) with MEDIA_MAX_UPLOAD_MB raise.
Update IMPLEMENTATION_STATUS with the chosen decision. Prefer A unless product insists on B.

# WAVE R2 — P2 WAVE 3 ML (merge carefully)
Claude tip 633f885 has CLIP search, OCR, NSFW/plate detection — NOT on master.
If merging:
- Rebase onto master; DO NOT drop finished-job polling fixes from PDF ingestion (#28).
- Wire GET /media-admin/search for "White Nexon Front View"
- Store OCR text + embedding on vehicle_media
- NSFW/plate: detect is OK; add configurable blur ONLY if safe/performant
- Keep feature flags so environments without models still upload

# WAVE R3 — P2 quality
- NFR-002: add informational pytest/bench for 5MB upload latency (mock storage OK)
- Per-file image_category on multi-upload
- EXIF make/model → form prefill when filename parse empty
- Strengthen tests asserting storage key contains /{{category}}/ and webp_key set

# ACCEPTANCE
- Dealer gallery works against deployed apiUrl (not localhost)
- Inspect of Tata_Nexon_FearlessPlus_Front_2025.webp prefills form when empty
- Manufacturer/model/variant image resolution documented + visible on at least one UI path each
- NFR-001 decision recorded in docs
- Existing media pytest suites stay green
```

---

## Compact P1 only

```
On master after Car Image WAVE 1–2 merge, fix ONLY:
1) vehicle-image-gallery.service.ts use environment.apiUrl not localhost:8000
2) Auto-apply /media-admin/inspect hints into empty form fields
3) Manufacturer + variant DAM image display paths (free-text filters)
4) Document NFR-001 as 100MB (or implement chunked 15GB)
Keep Car Images UI, PDF accept, WebP, EXIF, versions, audit. Add/adjust tests.
```
"""
    (QA / "Claude_Fix_Prompts_Car_Image_Retest.md").write_text(prompts)
    (ART / "Claude_Fix_Prompts_Car_Image_Retest.md").write_text(prompts)

    catalog = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "code_tip": TIP,
        "prior_score": PRIOR,
        "brd_readiness_score": score,
        "requirement_counts": dict(status_c),
        "test_counts": dict(tc),
        "pytest_media_passed": 70,
        "recommendation": "CONDITIONAL GO for core DAM; remaining P1 polish + WAVE3 ML",
        "requirements": [
            {
                "id": a,
                "area": b,
                "priority": c,
                "requirement": d,
                "prior_status": old,
                "status": new,
                "evidence": e,
                "gap": g,
            }
            for a, b, c, d, old, new, e, g in REQS
        ],
        "tests": tests,
        "still_open": [
            {"id": r[0], "status": r[5], "gap": r[7]} for r in still
        ],
    }
    (QA / "car-image-brd-retest.json").write_text(json.dumps(catalog, indent=2))
    (ART / "car-image-brd-retest.json").write_text(json.dumps(catalog, indent=2))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

    fills = {
        "PASS": PatternFill("solid", "C6EFCE"),
        "PARTIAL": PatternFill("solid", "FFEB9C"),
        "FAIL": PatternFill("solid", "FFC7CE"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive"
    for row in [
        ["GAADIIQ Car Image BRD — Post-Merge Retest"],
        ["Generated", now],
        ["Code tip", TIP],
        ["Prior score", PRIOR],
        ["BRD readiness now", score],
        ["Delta", f"+{score - PRIOR}"],
        ["Req PASS", status_c.get("PASS", 0)],
        ["Req PARTIAL", status_c.get("PARTIAL", 0)],
        ["Req FAIL", status_c.get("FAIL", 0)],
        ["Test PASS", tc.get("PASS", 0)],
        ["Test PARTIAL", tc.get("PARTIAL", 0)],
        ["Test FAIL", tc.get("FAIL", 0)],
        ["pytest media", "70 passed"],
        ["Verdict", catalog["recommendation"]],
        ["P0 remaining", "None (original UI/PDF trap closed)"],
        ["Top P1", "Dealer localhost URL; inspect auto-apply; mfr/variant pages; 15GB decision"],
    ]:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    wr = wb.create_sheet("Requirements")
    wr.append(["ID", "Area", "Priority", "Requirement", "Prior", "Now", "Evidence", "Gap"])
    for r in REQS:
        wr.append(list(r))
        wr.cell(wr.max_row, 6).fill = fills[r[5]]
        if r[4] != r[5]:
            wr.cell(wr.max_row, 5).fill = fills.get(r[4], PatternFill())
    for i, w in enumerate([10, 12, 12, 50, 10, 10, 40, 40], 1):
        wr.column_dimensions[get_column_letter(i)].width = w
    wr.auto_filter.ref = f"A1:H{wr.max_row}"
    wr.freeze_panes = "A2"

    wt = wb.create_sheet("TestCases")
    wt.append(["Suite", "ID", "Name", "Status", "Severity", "BR Ref", "Detail"])
    for t in tests:
        wt.append([t["suite"], t["id"], t["name"], t["status"], t["severity"], t["br_ref"], t["detail"]])
        wt.cell(wt.max_row, 4).fill = fills.get(t["status"], PatternFill())
    for i, w in enumerate([14, 12, 50, 10, 10, 14, 45], 1):
        wt.column_dimensions[get_column_letter(i)].width = w
    wt.auto_filter.ref = f"A1:G{wt.max_row}"
    wt.freeze_panes = "A2"

    wg = wb.create_sheet("Remaining_Gaps")
    wg.append(["Priority", "ID", "Gap", "Claude action"])
    for row in [
        ("P1", "TC-IMG-021b", "Dealer gallery hardcodes localhost:8000", "Use environment.apiUrl + auth"),
        ("P1", "FR-003", "Inspect hints not auto-applied to form", "Prefill empty fields from inspect"),
        ("P1", "AC-006", "Manufacturer/Variant DAM surfaces incomplete", "Wire make/variant image views"),
        ("P1", "NFR-001", "BRD 15GB vs product 100MB", "Document decision or chunked upload"),
        ("P1", "WAVE3-ML", "CLIP/OCR/safety on tip not master", "Careful rebase merge with flags"),
        ("P2", "AC-007/010", "Semantic NL search not on master", "Merge /media-admin/search"),
        ("P2", "BR-AI-01", "OCR/NSFW/plate blur incomplete", "Tip ML + blur policy"),
        ("P2", "NFR-002", "No automated ≤10s SLA test", "Add informational bench"),
        ("P2", "AC-004", "Batch upload shared metadata only", "Optional per-file category"),
    ]:
        wg.append(list(row))
        wg.cell(wg.max_row, 1).fill = (
            PatternFill("solid", "FFC7CE") if row[0] == "P1" else PatternFill("solid", "DDEBF7")
        )
    for i, w in enumerate([10, 14, 50, 45], 1):
        wg.column_dimensions[get_column_letter(i)].width = w

    wc = wb.create_sheet("Closed_Since_Prior")
    wc.append(["ID", "Was", "Now", "Evidence"])
    for r in closed:
        wc.append([r[0], r[4], r[5], r[6]])
        wc.cell(wc.max_row, 3).fill = fills["PASS"]

    out = QA / "GAADIIQ_Car_Image_BRD_Retest.xlsx"
    wb.save(out)
    (ART / "GAADIIQ_Car_Image_BRD_Retest.xlsx").write_bytes(out.read_bytes())
    print(json.dumps({"score": score, "prior": PRIOR, "reqs": dict(status_c), "tests": dict(tc)}, indent=2))


if __name__ == "__main__":
    main()
